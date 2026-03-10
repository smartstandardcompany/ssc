from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from typing import List, Optional
from datetime import datetime, timezone
import os
import shutil

from database import db, get_current_user, require_permission, get_branch_filter_with_global
from models import User, Supplier, SupplierCreate, SupplierPayment, SupplierPaymentCreate, SupplierCreditPayment

BILL_DIR = "/app/uploads/bills"
os.makedirs(BILL_DIR, exist_ok=True)

router = APIRouter()

@router.get("/suppliers")
async def get_suppliers(current_user: User = Depends(get_current_user)):
    require_permission(current_user, "suppliers", "read")
    query = get_branch_filter_with_global(current_user)
    
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    
    # Use MongoDB aggregation pipeline for total purchases (fast)
    pipeline = [
        {"$match": {"supplier_id": {"$exists": True, "$ne": None}}},
        {"$group": {"_id": "$supplier_id", "total": {"$sum": "$amount"}}}
    ]
    agg_result = await db.expenses.aggregate(pipeline).to_list(1000)
    purchase_totals = {r["_id"]: r["total"] for r in agg_result}
    
    for supplier in suppliers:
        if isinstance(supplier.get('created_at'), str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at']).isoformat()
        elif hasattr(supplier.get('created_at'), 'isoformat'):
            supplier['created_at'] = supplier['created_at'].isoformat()
        if supplier.get('updated_at') and hasattr(supplier['updated_at'], 'isoformat'):
            supplier['updated_at'] = supplier['updated_at'].isoformat()
        supplier['total_purchases'] = purchase_totals.get(supplier['id'], 0)
    
    return suppliers


@router.get("/suppliers/names")
async def get_supplier_names(current_user: User = Depends(get_current_user)):
    """Get just supplier names and IDs for dropdowns. No permission required beyond login."""
    # Use centralized filter that includes branch-specific AND global (no branch) suppliers
    query = get_branch_filter_with_global(current_user)
    
    suppliers = await db.suppliers.find(query, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    return [{"id": s["id"], "name": s["name"]} for s in suppliers]


@router.put("/suppliers/recalculate-all-balances")
async def recalculate_all_supplier_balances(current_user: User = Depends(get_current_user)):
    """Recalculate all supplier balances based on actual expenses and payments."""
    require_permission(current_user, "suppliers", "write")
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    all_credit_expenses = await db.expenses.find({"payment_mode": "credit", "supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(50000)
    all_payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(50000)
    
    results = []
    for supplier in suppliers:
        sid = supplier["id"]
        # Credit expenses add to balance
        total_credit = sum(e.get("amount", 0) for e in all_credit_expenses if e.get("supplier_id") == sid)
        # Cash/bank payments reduce balance
        total_paid = sum(p.get("amount", 0) for p in all_payments if p.get("supplier_id") == sid and p.get("payment_mode") in ["cash", "bank"])
        
        correct_balance = max(0, total_credit - total_paid)
        old_balance = supplier.get("current_credit", 0)
        
        if old_balance != correct_balance:
            await db.suppliers.update_one({"id": sid}, {"$set": {"current_credit": correct_balance}})
            results.append({
                "supplier_name": supplier["name"],
                "old_balance": old_balance,
                "new_balance": correct_balance
            })
    
    return {
        "suppliers_updated": len(results),
        "updates": results
    }


@router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "suppliers", "write")
    data = supplier_data.model_dump()
    for f in ['branch_id', 'category', 'sub_category', 'phone', 'email']:
        if data.get(f) == '': data[f] = None
    supplier = Supplier(**data)
    supplier_dict = supplier.model_dump()
    supplier_dict["created_at"] = supplier_dict["created_at"].isoformat()
    await db.suppliers.insert_one(supplier_dict)
    return supplier

@router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "suppliers", "write")
    result = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Supplier not found")
    update_data = supplier_data.model_dump()
    for f in ['branch_id', 'category', 'sub_category', 'phone', 'email']:
        if update_data.get(f) == '': update_data[f] = None
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Supplier(**updated)

@router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "suppliers", "write")
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    from routers.access_policies import check_delete_permission
    await check_delete_permission(current_user, "customers", supplier.get("created_at"), f"Supplier: {supplier.get('name', supplier_id[:8])}")
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    from routers.activity_logs import log_activity
    await log_activity(current_user, "delete", "suppliers", supplier_id)
    return {"message": "Supplier deleted successfully"}

@router.post("/suppliers/{supplier_id}/pay-credit")
async def pay_supplier_credit(supplier_id: str, payment: SupplierCreditPayment, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "supplier_payments", "write")
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    if payment.amount > supplier["current_credit"]:
        raise HTTPException(status_code=400, detail="Payment amount exceeds current credit")
    new_credit = supplier["current_credit"] - payment.amount
    await db.suppliers.update_one({"id": supplier_id}, {"$set": {"current_credit": new_credit}})
    payment_record = SupplierPayment(supplier_id=supplier_id, supplier_name=supplier["name"], amount=payment.amount, payment_mode=payment.payment_mode, branch_id=payment.branch_id, date=datetime.now(timezone.utc), notes=f"Credit payment - Remaining: SAR {new_credit:.2f}", created_by=current_user.id)
    payment_dict = payment_record.model_dump()
    payment_dict["date"] = payment_dict["date"].isoformat()
    payment_dict["created_at"] = payment_dict["created_at"].isoformat()
    await db.supplier_payments.insert_one(payment_dict)
    return {"message": "Credit payment recorded", "remaining_credit": new_credit}

# Supplier Payment Routes
@router.get("/supplier-payments")
async def get_supplier_payments(
    page: int = Query(1, ge=1),
    limit: int = Query(100, ge=1, le=500),
    supplier_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    require_permission(current_user, "supplier_payments", "read")
    query = {"supplier_id": {"$exists": True, "$ne": None}}
    if supplier_id: query["supplier_id"] = supplier_id
    if start_date:
        query["date"] = query.get("date", {}); query["date"]["$gte"] = start_date
    if end_date:
        query.setdefault("date", {}); query["date"]["$lte"] = end_date + "T23:59:59"
    total = await db.supplier_payments.count_documents(query)
    skip = (page - 1) * limit
    payments = await db.supplier_payments.find(query, {"_id": 0}).sort("date", -1).skip(skip).limit(limit).to_list(limit)
    for payment in payments:
        if isinstance(payment.get('date'), str): payment['date'] = datetime.fromisoformat(payment['date'])
        if isinstance(payment.get('created_at'), str): payment['created_at'] = datetime.fromisoformat(payment['created_at'])
    return {"data": payments, "total": total, "page": page, "limit": limit, "pages": (total + limit - 1) // limit}


@router.get("/supplier-payments/check-duplicate")
async def check_duplicate_supplier_payment(
    supplier_id: str = "",
    amount: float = 0,
    date: str = "",
    current_user: User = Depends(get_current_user)
):
    """Check if a supplier payment with same supplier and amount exists on the given date."""
    from datetime import timedelta
    d = datetime.strptime(date[:10], "%Y-%m-%d")
    next_day = (d + timedelta(days=1)).strftime("%Y-%m-%d")
    query = {"date": {"$gte": f"{date[:10]}T00:00:00", "$lt": f"{next_day}T00:00:00"}}
    if supplier_id:
        query["supplier_id"] = supplier_id
    existing = await db.supplier_payments.find(query, {"_id": 0, "id": 1, "amount": 1}).to_list(200)
    duplicates = [e for e in existing if abs(e.get("amount", 0) - amount) < 0.01]
    return {"has_duplicate": len(duplicates) > 0, "count": len(duplicates)}


@router.post("/supplier-payments", response_model=SupplierPayment)
async def create_supplier_payment(payment_data: SupplierPaymentCreate, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "supplier_payments", "write")
    supplier = await db.suppliers.find_one({"id": payment_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    if payment_data.payment_mode == "credit":
        credit_limit = supplier.get("credit_limit", 0); current_credit = supplier.get("current_credit", 0)
        new_credit = current_credit + payment_data.amount
        if credit_limit > 0 and new_credit > credit_limit:
            available = credit_limit - current_credit
            raise HTTPException(status_code=400, detail=f"Payment exceeds credit limit. Available: SAR {available:.2f}")
    payment = SupplierPayment(**payment_data.model_dump(), supplier_name=supplier["name"], created_by=current_user.id)
    payment_dict = payment.model_dump()
    payment_dict["date"] = payment_dict["date"].isoformat()
    payment_dict["created_at"] = payment_dict["created_at"].isoformat()
    await db.supplier_payments.insert_one(payment_dict)
    if payment_data.payment_mode == "credit":
        new_credit = supplier.get("current_credit", 0) + payment_data.amount
        await db.suppliers.update_one({"id": payment_data.supplier_id}, {"$set": {"current_credit": new_credit}})
    elif payment_data.payment_mode in ("cash", "bank"):
        # Cash/bank payment to supplier reduces their credit balance
        current_credit = supplier.get("current_credit", 0)
        new_credit = max(0, current_credit - payment_data.amount)
        await db.suppliers.update_one({"id": payment_data.supplier_id}, {"$set": {"current_credit": new_credit}})
    return payment

@router.delete("/supplier-payments/{payment_id}")
async def delete_supplier_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "supplier_payments", "write")
    
    payment = await db.supplier_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    from routers.access_policies import check_delete_permission
    await check_delete_permission(current_user, "supplier_payments", payment.get("date"), f"Payment to {payment.get('supplier_name', '')} - SAR {payment.get('amount', 0)}")
    
    supplier_id = payment.get("supplier_id")
    payment_mode = payment.get("payment_mode")
    amount = payment.get("amount", 0)
    
    if supplier_id:
        supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
        if supplier:
            current_credit = supplier.get("current_credit", 0)
            
            if payment_mode == "credit":
                # Credit payment was adding to supplier credit - reverse by reducing
                new_credit = max(0, current_credit - amount)
            elif payment_mode in ["cash", "bank"]:
                # Cash/Bank payment was reducing supplier credit (paying them back)
                # Reverse by increasing credit (you now owe them again)
                new_credit = current_credit + amount
            else:
                new_credit = current_credit
            
            await db.suppliers.update_one(
                {"id": supplier_id},
                {"$set": {"current_credit": new_credit}}
            )
    
    result = await db.supplier_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Payment deleted successfully", "supplier_balance_updated": True}


# =====================================================
# SUPPLIER RETURNS
# =====================================================

@router.post("/supplier-returns")
async def create_supplier_return(body: dict, current_user: User = Depends(get_current_user)):
    """Record a supplier return. Types: cash_refund, credit_return, full_invoice_return"""
    require_permission(current_user, "supplier_payments", "write")
    
    supplier_id = body.get("supplier_id")
    return_type = body.get("return_type", "credit_return")  # cash_refund, credit_return, full_invoice_return
    amount = float(body.get("amount", 0))
    invoice_ref = body.get("invoice_ref", "")
    reason = body.get("reason", "")
    branch_id = body.get("branch_id")
    date_str = body.get("date", datetime.now(timezone.utc).isoformat())
    bill_image_url = body.get("bill_image_url", "")
    
    if not supplier_id or amount <= 0:
        raise HTTPException(status_code=400, detail="Supplier and valid amount required")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    import uuid
    return_id = str(uuid.uuid4())
    
    return_record = {
        "id": return_id,
        "supplier_id": supplier_id,
        "supplier_name": supplier["name"],
        "return_type": return_type,
        "amount": amount,
        "invoice_ref": invoice_ref,
        "reason": reason,
        "branch_id": branch_id,
        "bill_image_url": bill_image_url,
        "date": date_str,
        "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.supplier_returns.insert_one(return_record)
    
    # Update supplier balance based on return type
    current_credit = supplier.get("current_credit", 0)
    
    if return_type == "cash_refund":
        # Supplier gives cash back - no credit change, but record as income/negative expense
        pass
    elif return_type == "credit_return":
        # Reduce the credit balance owed to supplier
        new_credit = max(0, current_credit - amount)
        await db.suppliers.update_one({"id": supplier_id}, {"$set": {"current_credit": new_credit}})
        return_record["balance_change"] = f"Credit reduced from SAR {current_credit:.2f} to SAR {new_credit:.2f}"
    elif return_type == "full_invoice_return":
        # Full invoice return - reduce credit balance
        new_credit = max(0, current_credit - amount)
        await db.suppliers.update_one({"id": supplier_id}, {"$set": {"current_credit": new_credit}})
        return_record["balance_change"] = f"Invoice returned. Credit reduced to SAR {new_credit:.2f}"
    
    return {k: v for k, v in return_record.items() if k != '_id'}


@router.get("/supplier-returns")
async def get_supplier_returns(supplier_id: str = None, current_user: User = Depends(get_current_user)):
    """Get all supplier returns, optionally filtered by supplier"""
    require_permission(current_user, "supplier_payments", "read")
    query = {}
    if supplier_id:
        query["supplier_id"] = supplier_id
    returns = await db.supplier_returns.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return returns


@router.delete("/supplier-returns/{return_id}")
async def delete_supplier_return(return_id: str, current_user: User = Depends(get_current_user)):
    """Delete a supplier return and reverse balance changes"""
    require_permission(current_user, "supplier_payments", "write")
    ret = await db.supplier_returns.find_one({"id": return_id}, {"_id": 0})
    if not ret:
        raise HTTPException(status_code=404, detail="Return not found")
    
    # Reverse the balance change
    if ret.get("return_type") in ["credit_return", "full_invoice_return"]:
        supplier = await db.suppliers.find_one({"id": ret["supplier_id"]}, {"_id": 0})
        if supplier:
            new_credit = supplier.get("current_credit", 0) + ret["amount"]
            await db.suppliers.update_one({"id": ret["supplier_id"]}, {"$set": {"current_credit": new_credit}})
    
    await db.supplier_returns.delete_one({"id": return_id})
    return {"message": "Return deleted and balance reversed"}


# =====================================================
# PURCHASE BILL IMAGE UPLOAD
# =====================================================

@router.post("/supplier-payments/upload-bill")
async def upload_bill_image(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Upload a bill/invoice image. Returns the URL to attach to a payment."""
    if not file.content_type or not (file.content_type.startswith("image/") or file.content_type == "application/pdf"):
        raise HTTPException(status_code=400, detail="File must be an image or PDF")
    
    import uuid as _uuid
    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "png"
    filename = f"bill_{_uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(BILL_DIR, filename)
    
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    return {"bill_url": f"/uploads/bills/{filename}", "message": "Bill uploaded"}


@router.get("/suppliers/{supplier_id}/payment-breakdown")
async def get_supplier_payment_breakdown(supplier_id: str, current_user: User = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    payments = await db.supplier_payments.find({"supplier_id": supplier_id}, {"_id": 0}).to_list(10000)
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    total_cash = sum(p["amount"] for p in payments if p.get("payment_mode") == "cash")
    total_bank = sum(p["amount"] for p in payments if p.get("payment_mode") == "bank")
    total_credit = sum(p["amount"] for p in payments if p.get("payment_mode") == "credit")
    by_branch = {}
    for p in payments:
        bid = p.get("branch_id"); bname = branches.get(bid, "No Branch") if bid else "No Branch"
        if bname not in by_branch: by_branch[bname] = {"cash": 0, "bank": 0, "credit": 0, "total": 0}
        mode = p.get("payment_mode", "cash"); by_branch[bname][mode] = by_branch[bname].get(mode, 0) + p["amount"]; by_branch[bname]["total"] += p["amount"]
    return {"supplier_id": supplier_id, "supplier_name": supplier["name"], "total_cash": total_cash, "total_bank": total_bank, "total_credit": total_credit, "total": total_cash + total_bank + total_credit, "by_branch": by_branch}

@router.get("/suppliers/payment-summaries")
async def get_all_supplier_summaries(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    result = {}
    for sup in suppliers:
        sid = sup["id"]; sp = [p for p in payments if p.get("supplier_id") == sid]
        by_branch = {}
        for p in sp:
            bid = p.get("branch_id"); bname = branches.get(bid, "No Branch") if bid else "No Branch"
            if bname not in by_branch: by_branch[bname] = {"cash": 0, "bank": 0}
            if p.get("payment_mode") == "cash": by_branch[bname]["cash"] += p["amount"]
            elif p.get("payment_mode") == "bank": by_branch[bname]["bank"] += p["amount"]
        result[sid] = {"cash": sum(p["amount"] for p in sp if p.get("payment_mode") == "cash"), "bank": sum(p["amount"] for p in sp if p.get("payment_mode") == "bank"), "by_branch": by_branch}
    return result


@router.post("/suppliers/{supplier_id}/recalculate-balance")
async def recalculate_supplier_balance(supplier_id: str, current_user: User = Depends(get_current_user)):
    """Recalculate supplier balance based on actual credit expenses and payments."""
    require_permission(current_user, "suppliers", "write")
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Get all credit expenses for this supplier
    credit_expenses = await db.expenses.find({
        "supplier_id": supplier_id,
        "payment_mode": "credit"
    }, {"_id": 0}).to_list(10000)
    total_credit_added = sum(e.get("amount", 0) for e in credit_expenses)
    
    # Get all credit payments made to this supplier (money paid back)
    credit_payments = await db.supplier_payments.find({
        "supplier_id": supplier_id
    }, {"_id": 0}).to_list(10000)
    # Only count cash/bank payments as they reduce credit
    total_paid = sum(p.get("amount", 0) for p in credit_payments if p.get("payment_mode") in ["cash", "bank"])
    
    # Calculate correct balance
    correct_balance = max(0, total_credit_added - total_paid)
    old_balance = supplier.get("current_credit", 0)
    
    await db.suppliers.update_one(
        {"id": supplier_id},
        {"$set": {"current_credit": correct_balance}}
    )
    
    return {
        "supplier_id": supplier_id,
        "supplier_name": supplier["name"],
        "old_balance": old_balance,
        "new_balance": correct_balance,
        "total_credit_expenses": total_credit_added,
        "total_payments": total_paid,
        "message": f"Balance updated from {old_balance} to {correct_balance}"
    }


@router.post("/suppliers/migrate-payments-to-bills")
async def migrate_payments_to_bills(current_user: User = Depends(get_current_user)):
    """
    Migrate supplier payments to purchase bills (expenses).
    This converts entries in supplier_payments to expenses with the correct supplier_id.
    After migration, supplier balances are recalculated.
    """
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    import uuid
    
    # Get all supplier payments
    payments = await db.supplier_payments.find({}, {"_id": 0}).to_list(50000)
    
    if not payments:
        return {"message": "No supplier payments to migrate", "migrated": 0}
    
    migrated_count = 0
    errors = []
    
    for payment in payments:
        try:
            # Create expense entry from payment
            expense = {
                "id": str(uuid.uuid4()),
                "amount": payment.get("amount", 0),
                "category": "Supplier Purchase",
                "description": payment.get("notes") or f"Migrated from supplier payment - {payment.get('supplier_name', 'Unknown')}",
                "payment_mode": payment.get("payment_mode", "credit"),  # Assume credit if buying on account
                "supplier_id": payment.get("supplier_id"),
                "branch_id": payment.get("branch_id", ""),
                "date": payment.get("date"),
                "created_by": payment.get("created_by"),
                "created_at": payment.get("created_at") or datetime.now(timezone.utc).isoformat(),
                "migrated_from": "supplier_payment",
                "original_payment_id": payment.get("id")
            }
            
            # Insert as expense
            await db.expenses.insert_one(expense)
            
            # Delete the old payment
            await db.supplier_payments.delete_one({"id": payment["id"]})
            
            migrated_count += 1
        except Exception as e:
            errors.append({"payment_id": payment.get("id"), "error": str(e)})
    
    # Recalculate all supplier balances
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    all_credit_expenses = await db.expenses.find(
        {"payment_mode": "credit", "supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}
    ).to_list(50000)
    all_payments = await db.supplier_payments.find(
        {"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}
    ).to_list(50000)
    
    balance_updates = []
    for supplier in suppliers:
        sid = supplier["id"]
        total_credit = sum(e.get("amount", 0) for e in all_credit_expenses if e.get("supplier_id") == sid)
        total_paid = sum(p.get("amount", 0) for p in all_payments if p.get("supplier_id") == sid and p.get("payment_mode") in ["cash", "bank"])
        
        new_balance = max(0, total_credit - total_paid)
        old_balance = supplier.get("current_credit", 0)
        
        await db.suppliers.update_one({"id": sid}, {"$set": {"current_credit": new_balance}})
        
        if old_balance != new_balance:
            balance_updates.append({
                "supplier": supplier["name"],
                "old": old_balance,
                "new": new_balance
            })
    
    return {
        "message": f"Migration complete! {migrated_count} payments converted to purchase bills.",
        "migrated_count": migrated_count,
        "errors": errors,
        "balance_updates": balance_updates,
        "note": "All supplier payments have been converted to expenses (purchase bills) with payment_mode=credit. Balances recalculated."
    }


@router.get("/suppliers/migration-preview")
async def preview_migration(current_user: User = Depends(get_current_user)):
    """Preview what will be migrated without actually doing it."""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    payments = await db.supplier_payments.find({}, {"_id": 0}).to_list(50000)
    
    # Group by supplier
    by_supplier = {}
    for p in payments:
        sname = p.get("supplier_name", "Unknown")
        if sname not in by_supplier:
            by_supplier[sname] = {"count": 0, "total": 0, "payments": []}
        by_supplier[sname]["count"] += 1
        by_supplier[sname]["total"] += p.get("amount", 0)
        by_supplier[sname]["payments"].append({
            "id": p.get("id"),
            "amount": p.get("amount"),
            "mode": p.get("payment_mode"),
            "date": p.get("date"),
            "notes": p.get("notes")
        })
    
    return {
        "total_payments": len(payments),
        "total_amount": sum(p.get("amount", 0) for p in payments),
        "by_supplier": by_supplier,
        "action": "These will be converted to purchase bills (expenses) with payment_mode matching original. Run POST /suppliers/migrate-payments-to-bills to execute."
    }



@router.get("/suppliers/{supplier_id}/ledger")
async def get_supplier_ledger(
    supplier_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Get supplier ledger with all purchases, payments, and running balance.
    Ledger shows complete transaction history for sharing with suppliers.
    """
    require_permission(current_user, "suppliers", "read")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Get branches for name lookup
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    
    # Build date query
    date_query = {}
    if start_date:
        date_query["$gte"] = start_date
    if end_date:
        date_query["$lte"] = end_date + "T23:59:59"
    
    # Get all expenses (purchases) from this supplier
    exp_query = {"supplier_id": supplier_id}
    if date_query:
        exp_query["date"] = date_query
    if branch_id:
        exp_query["branch_id"] = branch_id
    expenses = await db.expenses.find(exp_query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Get all payments to this supplier
    pay_query = {"supplier_id": supplier_id}
    if date_query:
        pay_query["date"] = date_query
    if branch_id:
        pay_query["branch_id"] = branch_id
    payments = await db.supplier_payments.find(pay_query, {"_id": 0}).sort("date", 1).to_list(10000)
    
    # Combine and sort by date
    ledger_entries = []
    
    for exp in expenses:
        bid = exp.get("branch_id", "")
        ledger_entries.append({
            "date": exp.get("date", ""),
            "type": "purchase",
            "description": exp.get("description") or exp.get("category", "Purchase"),
            "category": exp.get("category", ""),
            "debit": exp.get("amount", 0) if exp.get("payment_mode") == "credit" else 0,
            "credit": 0,
            "payment_mode": exp.get("payment_mode", ""),
            "reference": exp.get("id", "")[:8],
            "notes": exp.get("notes", ""),
            "branch_id": bid,
            "branch_name": branches.get(bid, "No Branch") if bid else "No Branch",
        })
    
    for pay in payments:
        bid = pay.get("branch_id", "")
        ledger_entries.append({
            "date": pay.get("date", ""),
            "type": "payment",
            "description": f"Payment ({pay.get('payment_mode', 'cash').title()})",
            "category": "",
            "debit": 0,
            "credit": pay.get("amount", 0),
            "payment_mode": pay.get("payment_mode", ""),
            "reference": pay.get("id", "")[:8],
            "notes": pay.get("notes", ""),
            "branch_id": bid,
            "branch_name": branches.get(bid, "No Branch") if bid else "No Branch",
        })
    
    # Sort by date
    ledger_entries.sort(key=lambda x: x["date"])
    
    # Calculate running balance
    running_balance = 0
    for entry in ledger_entries:
        running_balance += entry["debit"] - entry["credit"]
        entry["balance"] = running_balance
    
    # Calculate totals
    total_purchases = sum(e.get("amount", 0) for e in expenses)
    total_credit_purchases = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "credit")
    total_cash_purchases = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "cash")
    total_bank_purchases = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "bank")
    total_payments = sum(p.get("amount", 0) for p in payments)
    total_cash_payments = sum(p.get("amount", 0) for p in payments if p.get("payment_mode") == "cash")
    total_bank_payments = sum(p.get("amount", 0) for p in payments if p.get("payment_mode") == "bank")
    
    return {
        "supplier": {
            "id": supplier.get("id"),
            "name": supplier.get("name"),
            "phone": supplier.get("phone"),
            "email": supplier.get("email"),
            "category": supplier.get("category"),
            "bank_accounts": supplier.get("bank_accounts", []),
            "current_credit": supplier.get("current_credit", 0)
        },
        "period": {
            "start": start_date or "All time",
            "end": end_date or "Present"
        },
        "summary": {
            "total_purchases": total_purchases,
            "credit_purchases": total_credit_purchases,
            "cash_purchases": total_cash_purchases,
            "bank_purchases": total_bank_purchases,
            "total_payments": total_payments,
            "cash_payments": total_cash_payments,
            "bank_payments": total_bank_payments,
            "opening_balance": 0,
            "closing_balance": running_balance,
            "current_outstanding": supplier.get("current_credit", 0)
        },
        "entries": ledger_entries,
        "entry_count": len(ledger_entries)
    }


@router.get("/suppliers/{supplier_id}/ledger/export")
async def export_supplier_ledger(
    supplier_id: str,
    format: str = "pdf",
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Export supplier ledger as PDF or Excel.
    """
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    
    require_permission(current_user, "suppliers", "read")
    
    # Get ledger data
    ledger = await get_supplier_ledger(supplier_id, start_date, end_date, None, current_user)
    supplier = ledger["supplier"]
    entries = ledger["entries"]
    summary = ledger["summary"]
    
    if format == "excel":
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier Ledger"
        
        # Header
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Supplier Ledger - {supplier['name']}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Period: {ledger['period']['start']} to {ledger['period']['end']}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary
        ws['A4'] = "SUMMARY"
        ws['A4'].font = Font(bold=True)
        ws['A5'] = f"Total Purchases: SAR {summary['total_purchases']:,.2f}"
        ws['A6'] = f"Total Payments: SAR {summary['total_payments']:,.2f}"
        ws['A7'] = f"Outstanding Balance: SAR {summary['current_outstanding']:,.2f}"
        
        # Headers
        headers = ['Date', 'Type', 'Branch', 'Description', 'Debit', 'Credit', 'Balance', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid")
        
        # Data
        for row, entry in enumerate(entries, 10):
            ws.cell(row=row, column=1, value=entry['date'][:10] if entry['date'] else '')
            ws.cell(row=row, column=2, value=entry['type'].title())
            ws.cell(row=row, column=3, value=entry.get('branch_name', ''))
            ws.cell(row=row, column=4, value=entry['description'])
            ws.cell(row=row, column=5, value=entry['debit'] if entry['debit'] > 0 else '')
            ws.cell(row=row, column=6, value=entry['credit'] if entry['credit'] > 0 else '')
            ws.cell(row=row, column=7, value=entry['balance'])
            ws.cell(row=row, column=8, value=entry['reference'])
        
        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"ledger_{supplier['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    else:  # PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=10)
        elements.append(Paragraph(f"Supplier Ledger - {supplier['name']}", title_style))
        elements.append(Paragraph(f"Period: {ledger['period']['start']} to {ledger['period']['end']}", styles['Normal']))
        elements.append(Spacer(1, 10*mm))
        
        # Summary
        summary_data = [
            ['SUMMARY', ''],
            ['Total Purchases:', f"SAR {summary['total_purchases']:,.2f}"],
            ['Total Payments:', f"SAR {summary['total_payments']:,.2f}"],
            ['Outstanding:', f"SAR {summary['current_outstanding']:,.2f}"],
        ]
        summary_table = Table(summary_data, colWidths=[80*mm, 60*mm])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 10*mm))
        
        # Ledger entries
        ledger_data = [['Date', 'Type', 'Branch', 'Description', 'Debit', 'Credit', 'Balance']]
        for entry in entries:
            ledger_data.append([
                entry['date'][:10] if entry['date'] else '',
                entry['type'].title(),
                entry.get('branch_name', '-'),
                entry['description'][:25],
                f"{entry['debit']:,.2f}" if entry['debit'] > 0 else '',
                f"{entry['credit']:,.2f}" if entry['credit'] > 0 else '',
                f"{entry['balance']:,.2f}"
            ])
        
        ledger_table = Table(ledger_data, colWidths=[22*mm, 18*mm, 22*mm, 45*mm, 22*mm, 22*mm, 22*mm])
        ledger_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(ledger_table)
        
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"ledger_{supplier['name'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


@router.get("/suppliers/{supplier_id}/total-purchases")
async def get_supplier_total_purchases(
    supplier_id: str,
    current_user: User = Depends(get_current_user)
):
    """
    Get total purchase amount from a supplier.
    """
    require_permission(current_user, "suppliers", "read")
    
    expenses = await db.expenses.find({"supplier_id": supplier_id}, {"_id": 0}).to_list(50000)
    
    total = sum(e.get("amount", 0) for e in expenses)
    cash = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "cash")
    bank = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "bank")
    credit = sum(e.get("amount", 0) for e in expenses if e.get("payment_mode") == "credit")
    
    return {
        "supplier_id": supplier_id,
        "total_purchases": total,
        "cash_purchases": cash,
        "bank_purchases": bank,
        "credit_purchases": credit,
        "purchase_count": len(expenses)
    }


@router.post("/suppliers/{supplier_id}/share-statement")
async def share_supplier_statement(
    supplier_id: str,
    body: dict,
    current_user: User = Depends(get_current_user)
):
    """
    Share supplier ledger/statement via Email and/or WhatsApp.
    body: { channels: ["email", "whatsapp"], email: "...", phone: "+...", start_date, end_date }
    """
    require_permission(current_user, "suppliers", "read")
    
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    channels = body.get("channels", [])
    email_to = body.get("email", supplier.get("email"))
    phone_to = body.get("phone", supplier.get("phone"))
    start_date = body.get("start_date")
    end_date = body.get("end_date")
    
    if not channels:
        raise HTTPException(status_code=400, detail="At least one channel (email/whatsapp) required")
    
    # Get ledger data
    ledger = await get_supplier_ledger(supplier_id, start_date, end_date, None, current_user)
    supplier_info = ledger["supplier"]
    entries = ledger["entries"]
    summary = ledger["summary"]
    
    results = {"email": None, "whatsapp": None}
    
    # --- Send via Email with PDF attachment ---
    if "email" in channels:
        if not email_to:
            results["email"] = {"success": False, "error": "No email address provided"}
        else:
            try:
                import aiosmtplib
                from email.mime.multipart import MIMEMultipart
                from email.mime.text import MIMEText
                from email.mime.application import MIMEApplication
                
                email_settings = await db.email_settings.find_one({}, {"_id": 0})
                if not email_settings or not email_settings.get("smtp_host"):
                    results["email"] = {"success": False, "error": "Email not configured"}
                else:
                    # Generate PDF
                    pdf_buffer = _generate_ledger_pdf(supplier_info, entries, summary, ledger["period"])
                    
                    # Build email
                    msg = MIMEMultipart()
                    msg["Subject"] = f"Statement of Account - {supplier_info['name']}"
                    msg["From"] = email_settings.get("from_email", email_settings["username"])
                    msg["To"] = email_to
                    
                    period_text = f"{ledger['period']['start']} to {ledger['period']['end']}"
                    html_body = f"""
                    <h2>Statement of Account</h2>
                    <p><strong>Supplier:</strong> {supplier_info['name']}</p>
                    <p><strong>Period:</strong> {period_text}</p>
                    <p><strong>Total Purchases:</strong> SAR {summary['total_purchases']:,.2f}</p>
                    <p><strong>Total Payments:</strong> SAR {summary['total_payments']:,.2f}</p>
                    <p><strong>Outstanding Balance:</strong> SAR {summary['current_outstanding']:,.2f}</p>
                    <br>
                    <p>Please find the detailed statement attached as PDF.</p>
                    <p>Regards,<br>SSC Track</p>
                    """
                    msg.attach(MIMEText(html_body, "html"))
                    
                    # Attach PDF
                    pdf_attachment = MIMEApplication(pdf_buffer.read(), _subtype="pdf")
                    pdf_attachment.add_header("Content-Disposition", "attachment", 
                        filename=f"statement_{supplier_info['name'].replace(' ', '_')}.pdf")
                    msg.attach(pdf_attachment)
                    
                    await aiosmtplib.send(msg, 
                        hostname=email_settings["smtp_host"], 
                        port=email_settings["smtp_port"],
                        username=email_settings["username"], 
                        password=email_settings["password"],
                        use_tls=False, start_tls=True, timeout=30)
                    results["email"] = {"success": True, "sent_to": email_to}
            except Exception as e:
                results["email"] = {"success": False, "error": str(e)}
    
    # --- Send via WhatsApp ---
    if "whatsapp" in channels:
        if not phone_to:
            results["whatsapp"] = {"success": False, "error": "No phone number provided"}
        else:
            try:
                from routers.whatsapp import send_whatsapp_message
                
                # Build text summary
                period_text = f"{ledger['period']['start']} to {ledger['period']['end']}"
                wa_msg = (
                    f"*SSC Track - Supplier Statement*\n\n"
                    f"*{supplier_info['name']}*\n"
                    f"Period: {period_text}\n\n"
                    f"Total Purchases: SAR {summary['total_purchases']:,.2f}\n"
                    f"  - Cash/Bank: SAR {(summary['cash_purchases'] + summary['bank_purchases']):,.2f}\n"
                    f"  - Credit: SAR {summary['credit_purchases']:,.2f}\n\n"
                    f"Total Payments: SAR {summary['total_payments']:,.2f}\n"
                    f"*Outstanding Balance: SAR {summary['current_outstanding']:,.2f}*\n\n"
                    f"Transactions: {ledger['entry_count']}"
                )
                
                # Try sending to specific number
                config = await db.whatsapp_config.find_one({}, {"_id": 0})
                if not config or not config.get("account_sid"):
                    results["whatsapp"] = {"success": False, "error": "WhatsApp not configured"}
                else:
                    from twilio.rest import Client
                    client_tw = Client(config["account_sid"], config["auth_token"])
                    client_tw.messages.create(
                        from_=f'whatsapp:{config["phone_number"]}',
                        body=wa_msg,
                        to=f'whatsapp:{phone_to}'
                    )
                    results["whatsapp"] = {"success": True, "sent_to": phone_to}
            except Exception as e:
                results["whatsapp"] = {"success": False, "error": str(e)}
    
    return {
        "message": "Statement sharing completed",
        "supplier": supplier_info["name"],
        "results": results
    }


def _generate_ledger_pdf(supplier_info, entries, summary, period):
    """Generate a PDF buffer for the supplier ledger."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=10)
    elements.append(Paragraph(f"Statement of Account - {supplier_info['name']}", title_style))
    elements.append(Paragraph(f"Period: {period['start']} to {period['end']}", styles['Normal']))
    if supplier_info.get('phone'):
        elements.append(Paragraph(f"Phone: {supplier_info['phone']}", styles['Normal']))
    if supplier_info.get('email'):
        elements.append(Paragraph(f"Email: {supplier_info['email']}", styles['Normal']))
    elements.append(Spacer(1, 10*mm))
    
    summary_data = [
        ['SUMMARY', ''],
        ['Total Purchases:', f"SAR {summary['total_purchases']:,.2f}"],
        ['Cash/Bank Purchases:', f"SAR {(summary['cash_purchases'] + summary['bank_purchases']):,.2f}"],
        ['Credit Purchases:', f"SAR {summary['credit_purchases']:,.2f}"],
        ['Total Payments:', f"SAR {summary['total_payments']:,.2f}"],
        ['Outstanding:', f"SAR {summary['current_outstanding']:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[80*mm, 60*mm])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.95, 1.0)),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 10*mm))
    
    ledger_data = [['Date', 'Type', 'Description', 'Debit', 'Credit', 'Balance']]
    for entry in entries:
        ledger_data.append([
            entry['date'][:10] if entry['date'] else '',
            entry['type'].title(),
            entry['description'][:30] if entry.get('description') else '',
            f"{entry['debit']:,.2f}" if entry['debit'] > 0 else '',
            f"{entry['credit']:,.2f}" if entry['credit'] > 0 else '',
            f"{entry['balance']:,.2f}"
        ])
    
    ledger_table = Table(ledger_data, colWidths=[25*mm, 20*mm, 55*mm, 25*mm, 25*mm, 25*mm])
    ledger_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(ledger_table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


@router.get("/suppliers/aging-report")
async def get_supplier_aging_report(
    branch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """
    Supplier aging report - groups outstanding credit balances by age buckets:
    0-30 days, 31-60 days, 61-90 days, 90+ days
    """
    require_permission(current_user, "suppliers", "read")
    
    query = get_branch_filter_with_global(current_user)
    if branch_id:
        query["branch_id"] = branch_id
    
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    
    # Only suppliers with outstanding credit
    suppliers_with_credit = [s for s in suppliers if s.get("current_credit", 0) > 0]
    
    now = datetime.now(timezone.utc)
    aging_data = []
    totals = {"0_30": 0, "31_60": 0, "61_90": 0, "90_plus": 0, "total": 0}
    
    for supplier in suppliers_with_credit:
        sid = supplier["id"]
        
        # Get all credit purchases (unpaid) for this supplier
        credit_expenses = await db.expenses.find(
            {"supplier_id": sid, "payment_mode": "credit"},
            {"_id": 0, "id": 1, "date": 1, "amount": 1, "description": 1}
        ).sort("date", 1).to_list(5000)
        
        # Get all payments made to this supplier
        payments = await db.supplier_payments.find(
            {"supplier_id": sid, "payment_mode": {"$in": ["cash", "bank"]}},
            {"_id": 0, "amount": 1}
        ).to_list(5000)
        
        total_paid = sum(p.get("amount", 0) for p in payments)
        
        # Apply payments to oldest invoices first (FIFO)
        remaining_payment = total_paid
        buckets = {"0_30": 0, "31_60": 0, "61_90": 0, "90_plus": 0}
        unpaid_invoices = []
        
        for exp in credit_expenses:
            if remaining_payment >= exp["amount"]:
                remaining_payment -= exp["amount"]
                continue  # Fully paid
            
            unpaid_amount = exp["amount"] - remaining_payment
            remaining_payment = 0
            
            # Calculate age in days
            try:
                if isinstance(exp.get("date"), str):
                    exp_date = datetime.fromisoformat(exp["date"].replace("Z", "+00:00"))
                else:
                    exp_date = exp.get("date", now)
                if exp_date.tzinfo is None:
                    from datetime import timezone as tz
                    exp_date = exp_date.replace(tzinfo=tz.utc)
                age_days = (now - exp_date).days
            except:
                age_days = 0
            
            if age_days <= 30:
                buckets["0_30"] += unpaid_amount
            elif age_days <= 60:
                buckets["31_60"] += unpaid_amount
            elif age_days <= 90:
                buckets["61_90"] += unpaid_amount
            else:
                buckets["90_plus"] += unpaid_amount
            
            unpaid_invoices.append({
                "date": exp.get("date", ""),
                "amount": exp["amount"],
                "unpaid": unpaid_amount,
                "age_days": age_days,
                "description": exp.get("description", "")
            })
        
        supplier_total = sum(buckets.values())
        if supplier_total > 0:
            aging_data.append({
                "supplier_id": sid,
                "supplier_name": supplier["name"],
                "branch_id": supplier.get("branch_id"),
                "credit_limit": supplier.get("credit_limit", 0),
                "current_credit": supplier.get("current_credit", 0),
                "buckets": buckets,
                "total_outstanding": supplier_total,
                "unpaid_invoices": unpaid_invoices[:20],  # Limit for response size
            })
            
            for key in totals:
                if key != "total":
                    totals[key] += buckets.get(key, 0)
            totals["total"] += supplier_total
    
    # Sort by total outstanding descending
    aging_data.sort(key=lambda x: x["total_outstanding"], reverse=True)
    
    return {
        "report_date": now.isoformat(),
        "totals": totals,
        "supplier_count": len(aging_data),
        "suppliers": aging_data
    }


@router.get("/suppliers/aging-report/export")
async def export_aging_report(
    format: str = "pdf",
    branch_id: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    """Export aging report as PDF or Excel"""
    require_permission(current_user, "suppliers", "read")
    
    report = await get_supplier_aging_report(branch_id, current_user)
    
    if format == "excel":
        import openpyxl
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier Aging Report"
        
        # Header
        ws.append(["Supplier Aging Report", "", "", "", "", ""])
        ws.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
        ws.append([])
        
        # Totals
        ws.append(["SUMMARY"])
        ws.append(["0-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"])
        ws.append([report["totals"]["0_30"], report["totals"]["31_60"], report["totals"]["61_90"], report["totals"]["90_plus"], report["totals"]["total"]])
        ws.append([])
        
        # Detail
        ws.append(["Supplier", "Credit Limit", "Current Credit", "0-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total Outstanding"])
        for s in report["suppliers"]:
            ws.append([
                s["supplier_name"], s["credit_limit"], s["current_credit"],
                s["buckets"]["0_30"], s["buckets"]["31_60"], s["buckets"]["61_90"], s["buckets"]["90_plus"],
                s["total_outstanding"]
            ])
        
        # Format columns
        for col in range(2, 9):
            for row in range(6, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell.number_format = '#,##0.00'
        
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=aging_report.xlsx"})
    
    else:  # PDF
        from io import BytesIO
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from fastapi.responses import StreamingResponse
        
        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle('Title', parent=styles['Title'], fontSize=16, spaceAfter=10)
        elements.append(Paragraph("Supplier Aging Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 8*mm))
        
        # Summary table
        summary_data = [
            ["0-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total Outstanding"],
            [f"SAR {report['totals']['0_30']:,.2f}", f"SAR {report['totals']['31_60']:,.2f}",
             f"SAR {report['totals']['61_90']:,.2f}", f"SAR {report['totals']['90_plus']:,.2f}",
             f"SAR {report['totals']['total']:,.2f}"],
        ]
        st = Table(summary_data, colWidths=[50*mm]*5)
        st.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
        elements.append(st)
        elements.append(Spacer(1, 8*mm))
        
        # Detail table
        detail_data = [["Supplier", "Credit Limit", "0-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total"]]
        for s in report["suppliers"]:
            detail_data.append([
                s["supplier_name"][:20], f"SAR {s['credit_limit']:,.0f}",
                f"{s['buckets']['0_30']:,.2f}", f"{s['buckets']['31_60']:,.2f}",
                f"{s['buckets']['61_90']:,.2f}", f"{s['buckets']['90_plus']:,.2f}",
                f"{s['total_outstanding']:,.2f}",
            ])
        dt = Table(detail_data, colWidths=[45*mm, 30*mm, 30*mm, 30*mm, 30*mm, 30*mm, 35*mm])
        dt.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ]))
        elements.append(dt)
        
        doc.build(elements)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=aging_report.pdf"})
