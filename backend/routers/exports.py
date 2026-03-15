from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from database import db, get_current_user, get_tenant_filter, stamp_tenant
from models import User, ExportRequest

router = APIRouter()

@router.post("/export/data")
async def export_data(request: dict, current_user: User = Depends(get_current_user)):
    data_type = request.get("type", "sales")
    fmt = request.get("format", "excel")
    start_date = request.get("start_date")
    end_date = request.get("end_date")
    filters = request.get("filters", {})
    branches = await db.branches.find(get_tenant_filter(current_user), {"_id": 0}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}

    # Build date filter
    date_filter = {}
    if start_date:
        date_filter["$gte"] = start_date
    if end_date:
        date_filter["$lte"] = end_date + "T23:59:59"
    date_query = {"date": date_filter} if date_filter else {}

    # Date label for title
    date_label = ""
    if start_date and end_date:
        date_label = f" ({start_date} to {end_date})"
    elif start_date:
        date_label = f" (from {start_date})"

    buffer = BytesIO()
    if data_type == "sales":
        query = {**date_query}
        if filters.get("branch_id"):
            query["branch_id"] = filters["branch_id"]
        sales = await db.sales.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
        customers = await db.customers.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        cust_map = {c["id"]: c["name"] for c in customers if c.get("id")}
        rows = []
        for s in sales:
            modes = ", ".join(f'{p["mode"]}:${p["amount"]:.2f}' for p in s.get("payment_details", []))
            rows.append([
                datetime.fromisoformat(s["date"]).strftime("%Y-%m-%d") if isinstance(s["date"], str) else s["date"].strftime("%Y-%m-%d"),
                s["sale_type"].capitalize(), branch_map.get(s.get("branch_id"), "-"), cust_map.get(s.get("customer_id"), "-"),
                s["amount"], s.get("discount", 0), s.get("final_amount", s["amount"] - s.get("discount", 0)), modes,
                s.get("credit_amount", 0) - s.get("credit_received", 0)
            ])
        headers = ["Date", "Type", "Branch", "Customer", "Amount", "Discount", "Final", "Payments", "Credit Remaining"]
        title = f"Sales Report{date_label}"
    elif data_type == "expenses":
        query = {**date_query}
        if filters.get("branch_id"):
            query["branch_id"] = filters["branch_id"]
        if filters.get("category"):
            query["category"] = filters["category"]
        expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
        suppliers = await db.suppliers.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        sup_map = {s["id"]: s["name"] for s in suppliers}
        rows = []
        for e in expenses:
            rows.append([
                datetime.fromisoformat(e["date"]).strftime("%Y-%m-%d") if isinstance(e["date"], str) else e["date"].strftime("%Y-%m-%d"),
                e["category"].capitalize(), e["description"], sup_map.get(e.get("supplier_id"), "-"),
                branch_map.get(e.get("branch_id"), "-"), e["amount"], e["payment_mode"].capitalize()
            ])
        headers = ["Date", "Category", "Description", "Supplier", "Branch", "Amount", "Payment Mode"]
        title = f"Expenses Report{date_label}"
    elif data_type == "supplier-payments":
        query = {**date_query, "supplier_id": {"$exists": True, "$ne": None}}
        payments = await db.supplier_payments.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
        rows = []
        for p in payments:
            rows.append([
                datetime.fromisoformat(p["date"]).strftime("%Y-%m-%d") if isinstance(p["date"], str) else p["date"].strftime("%Y-%m-%d"),
                p["supplier_name"], branch_map.get(p.get("branch_id"), "-"), p["amount"], p["payment_mode"].capitalize(), p.get("notes", "")
            ])
        headers = ["Date", "Supplier", "Branch", "Amount", "Payment Mode", "Notes"]
        title = f"Supplier Payments Report{date_label}"
    elif data_type == "customers":
        customers = await db.customers.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        rows = [[c["name"], branch_map.get(c.get("branch_id"), "All Branches"), c.get("phone", "-"), c.get("email", "-")] for c in customers]
        headers = ["Name", "Branch", "Phone", "Email"]
        title = "Customers Report"
    elif data_type == "suppliers":
        suppliers = await db.suppliers.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        rows = [[s["name"], s.get("category", "-"), branch_map.get(s.get("branch_id"), "All"), s.get("phone", "-"), s.get("current_credit", 0), s.get("credit_limit", 0)] for s in suppliers]
        headers = ["Name", "Category", "Branch", "Phone", "Current Credit", "Credit Limit"]
        title = "Suppliers Report"
    elif data_type == "employees":
        employees = await db.employees.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        rows = []
        for emp in employees:
            rows.append([emp["name"], emp.get("position", "-"), emp.get("document_id", "-"), branch_map.get(emp.get("branch_id"), "-"), emp.get("salary", 0), emp.get("pay_frequency", "monthly"), datetime.fromisoformat(emp["document_expiry"]).strftime("%Y-%m-%d") if emp.get("document_expiry") else "-"])
        headers = ["Name", "Position", "Document ID", "Branch", "Salary", "Pay Frequency", "Doc Expiry"]
        title = "Employees Report"
    elif data_type == "loans":
        loans = await db.loans.find(get_tenant_filter(current_user), {"_id": 0}).sort("created_at", -1).to_list(5000)
        rows = []
        for l in loans:
            rows.append([l["employee_name"], l["loan_type"].replace("_", " ").capitalize(), l["amount"], l.get("monthly_installment", 0), l.get("total_installments", 0), l.get("paid_installments", 0), l.get("remaining_balance", 0), l["status"].capitalize(), datetime.fromisoformat(l["created_at"]).strftime("%Y-%m-%d") if isinstance(l.get("created_at"), str) else "-"])
        headers = ["Employee", "Loan Type", "Amount", "Monthly Installment", "Total Inst.", "Paid Inst.", "Remaining", "Status", "Created"]
        title = "Loans Report"
    elif data_type == "attendance":
        attendance = await db.attendance.find(get_tenant_filter(current_user), {"_id": 0}).sort("date", -1).to_list(10000)
        emps = await db.employees.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        emp_map = {e["id"]: e["name"] for e in emps}
        rows = []
        for a in attendance:
            tin = datetime.fromisoformat(a["time_in"]).strftime("%H:%M") if a.get("time_in") else "-"
            tout = datetime.fromisoformat(a["time_out"]).strftime("%H:%M") if a.get("time_out") else "-"
            hours = "-"
            if a.get("time_in") and a.get("time_out"):
                delta = (datetime.fromisoformat(a["time_out"]) - datetime.fromisoformat(a["time_in"])).total_seconds() / 3600
                hours = f"{delta:.1f}"
            rows.append([a.get("date", "-"), emp_map.get(a.get("employee_id"), a.get("employee_name", "-")), tin, tout, hours])
        headers = ["Date", "Employee", "Time In", "Time Out", "Hours"]
        title = "Attendance Report"
    elif data_type == "leaves":
        leaves = await db.leaves.find(get_tenant_filter(current_user), {"_id": 0}).sort("created_at", -1).to_list(5000)
        rows = []
        for l in leaves:
            start = datetime.fromisoformat(l["start_date"]).strftime("%Y-%m-%d") if isinstance(l.get("start_date"), str) else "-"
            end = datetime.fromisoformat(l["end_date"]).strftime("%Y-%m-%d") if isinstance(l.get("end_date"), str) else "-"
            rows.append([l.get("employee_name", "-"), l.get("leave_type", "-").capitalize(), start, end, l.get("days", 0), l.get("status", "-").capitalize(), l.get("reason", "-")])
        headers = ["Employee", "Type", "From", "To", "Days", "Status", "Reason"]
        title = "Leave Report"
    else:
        raise HTTPException(status_code=400, detail="Invalid data type")
    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(headers)
        for col in ws[1]:
            col.font = Font(bold=True, color="FFFFFF")
            col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
        for row in rows:
            ws.append(row)
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)
        # Build filename with dates
        date_suffix = ""
        if start_date and end_date:
            date_suffix = f"_{start_date}_to_{end_date}"
        elif start_date:
            date_suffix = f"_from_{start_date}"
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={data_type}_report{date_suffix}.xlsx"})
    else:
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#F5841F'), alignment=1)
        elements.append(Paragraph(title, title_style))
        # Add date range subtitle
        if start_date or end_date:
            date_info = f"Period: {start_date or 'All'} to {end_date or 'Present'}"
            sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=1)
            elements.append(Paragraph(date_info, sub_style))
        elements.append(Spacer(1, 0.2*inch))
        table_data = [headers] + [[str(c) for c in row] for row in rows[:50]]
        col_count = len(headers)
        col_width = 7.5 * inch / col_count
        t = Table(table_data, colWidths=[col_width] * col_count)
        t.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, -1), 7), ('GRID', (0, 0), (-1, -1), 0.5, colors.grey), ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')])]))
        elements.append(t)
        # Build filename with dates for PDF
        date_suffix = ""
        if start_date and end_date:
            date_suffix = f"_{start_date}_to_{end_date}"
        elif start_date:
            date_suffix = f"_from_{start_date}"
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={data_type}_report{date_suffix}.pdf"})

@router.post("/export/reports")
async def export_reports(export_request: ExportRequest, current_user: User = Depends(get_current_user)):
    try:
        sales = await db.sales.find(get_tenant_filter(current_user), {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find(get_tenant_filter(current_user), {"_id": 0}).to_list(10000)
        supplier_payments = await db.supplier_payments.find(get_tenant_filter(current_user), {"_id": 0}).to_list(10000)
        branches = await db.branches.find(get_tenant_filter(current_user), {"_id": 0}).to_list(100)
        customers = await db.customers.find(get_tenant_filter(current_user), {"_id": 0}).to_list(1000)
        if export_request.start_date:
            start = datetime.fromisoformat(export_request.start_date)
            sales = [s for s in sales if datetime.fromisoformat(s["date"]) >= start]
            expenses = [e for e in expenses if datetime.fromisoformat(e["date"]) >= start]
            supplier_payments = [p for p in supplier_payments if datetime.fromisoformat(p["date"]) >= start]
        if export_request.end_date:
            end = datetime.fromisoformat(export_request.end_date)
            sales = [s for s in sales if datetime.fromisoformat(s["date"]) <= end]
            expenses = [e for e in expenses if datetime.fromisoformat(e["date"]) <= end]
            supplier_payments = [p for p in supplier_payments if datetime.fromisoformat(p["date"]) <= end]
        if export_request.branch_id and export_request.branch_id != "all":
            sales = [s for s in sales if s.get("branch_id") == export_request.branch_id]
        if export_request.format == "pdf":
            return _generate_pdf_report(sales, expenses, supplier_payments, branches, customers)
        elif export_request.format == "excel":
            return _generate_excel_report(sales, expenses, supplier_payments, branches, customers)
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


def _generate_pdf_report(sales, expenses, supplier_payments, branches, customers):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, textColor=colors.HexColor('#F5841F'), spaceAfter=30, alignment=1)
    elements.append(Paragraph("SSC Track - Sales Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    total_sales = sum(s.get("final_amount", s["amount"] - s.get("discount", 0)) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    total_supplier = sum(p["amount"] for p in supplier_payments)
    net_profit = total_sales - total_expenses - total_supplier
    summary_data = [["Metric", "Amount"], ["Total Sales", f"SAR {total_sales:.2f}"], ["Total Expenses", f"SAR {total_expenses:.2f}"], ["Supplier Payments", f"SAR {total_supplier:.2f}"], ["Net Profit", f"SAR {net_profit:.2f}"]]
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 14), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph("Sales Transactions", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    sales_data = [["Date", "Type", "Amount", "Payment"]]
    for sale in sales[:20]:
        date_str = datetime.fromisoformat(sale["date"]).strftime("%Y-%m-%d")
        modes = ", ".join(p["mode"].capitalize() for p in sale.get("payment_details", []))
        if not modes: modes = sale.get("payment_mode", "N/A").capitalize()
        sales_data.append([date_str, sale["sale_type"].capitalize(), f"SAR {sale['amount']:.2f}", modes])
    sales_table = Table(sales_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    sales_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'CENTER'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('FONTSIZE', (0, 0), (-1, 0), 10), ('BOTTOMPADDING', (0, 0), (-1, 0), 8), ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(sales_table)
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=sales_report.pdf"})


def _generate_excel_report(sales, expenses, supplier_payments, branches, customers):
    buffer = BytesIO()
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary['A1'] = "SSC Track - Sales Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="F5841F")
    total_sales = sum(s.get("final_amount", s["amount"] - s.get("discount", 0)) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    total_supplier = sum(p["amount"] for p in supplier_payments)
    net_profit = total_sales - total_expenses - total_supplier
    ws_summary['A3'] = "Metric"; ws_summary['B3'] = "Amount"
    ws_summary['A3'].font = Font(bold=True); ws_summary['B3'].font = Font(bold=True)
    for idx, (metric, amount) in enumerate([("Total Sales", total_sales), ("Total Expenses", total_expenses), ("Supplier Payments", total_supplier), ("Net Profit", net_profit)], start=4):
        ws_summary[f'A{idx}'] = metric; ws_summary[f'B{idx}'] = f"SAR {amount:.2f}"
    ws_sales = wb.create_sheet("Sales")
    sales_headers = ["Date", "Type", "Branch/Customer", "Amount", "Payment Mode", "Status"]
    ws_sales.append(sales_headers)
    for col in ws_sales[1]:
        col.font = Font(bold=True, color="FFFFFF"); col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
    for sale in sales:
        date_str = datetime.fromisoformat(sale["date"]).strftime("%Y-%m-%d")
        branch_name = next((b["name"] for b in branches if b["id"] == sale.get("branch_id")), "-")
        customer_name = next((c["name"] for c in customers if c["id"] == sale.get("customer_id")), "-")
        ref = branch_name if sale["sale_type"] == "branch" else customer_name
        modes = ", ".join(p["mode"].capitalize() for p in sale.get("payment_details", []))
        if not modes: modes = sale.get("payment_mode", "N/A").capitalize()
        has_credit = sale.get("credit_amount", 0) > 0
        remaining = sale.get("credit_amount", 0) - sale.get("credit_received", 0)
        status = "Paid" if remaining <= 0 else ("Partial" if sale.get("credit_received", 0) > 0 else "Pending") if has_credit else "Received"
        ws_sales.append([date_str, sale["sale_type"].capitalize(), ref, sale["amount"], modes, status])
    ws_expenses = wb.create_sheet("Expenses")
    ws_expenses.append(["Date", "Category", "Description", "Amount", "Payment Mode"])
    for col in ws_expenses[1]:
        col.font = Font(bold=True, color="FFFFFF"); col.fill = PatternFill(start_color="F43F5E", end_color="F43F5E", fill_type="solid")
    for expense in expenses:
        date_str = datetime.fromisoformat(expense["date"]).strftime("%Y-%m-%d")
        ws_expenses.append([date_str, expense["category"].capitalize(), expense["description"], expense["amount"], expense["payment_mode"].capitalize()])
    ws_supplier = wb.create_sheet("Supplier Payments")
    ws_supplier.append(["Date", "Supplier Name", "Amount", "Payment Mode", "Notes"])
    for col in ws_supplier[1]:
        col.font = Font(bold=True, color="FFFFFF"); col.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
    for payment in supplier_payments:
        date_str = datetime.fromisoformat(payment["date"]).strftime("%Y-%m-%d")
        ws_supplier.append([date_str, payment["supplier_name"], payment["amount"], payment["payment_mode"].capitalize(), payment.get("notes", "")])
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"})
