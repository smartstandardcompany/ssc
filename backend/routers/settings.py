from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import uuid
import pandas as pd

from database import db, get_current_user, ROOT_DIR, require_permission
from models import User

router = APIRouter()

# Company Settings
@router.get("/settings/company")
async def get_company_settings(current_user: User = Depends(get_current_user)):
    settings = await db.company_settings.find_one({}, {"_id": 0})
    return settings or {"company_name": "Smart Standard Company", "address_line1": "", "address_line2": "", "city": "", "country": "", "phone": "", "email": "", "cr_number": "", "vat_number": "", "vat_enabled": False, "vat_rate": 15}

@router.post("/settings/company")
async def save_company_settings(body: dict, current_user: User = Depends(get_current_user)):
    require_permission(current_user, "settings", "write")
    existing = await db.company_settings.find_one({})
    data = {k: body.get(k, "") for k in ["company_name", "address_line1", "address_line2", "city", "country", "phone", "email", "cr_number", "vat_number"]}
    data["vat_enabled"] = body.get("vat_enabled", False)
    data["vat_rate"] = float(body.get("vat_rate", 15) or 15)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    if existing:
        await db.company_settings.update_one({}, {"$set": data})
    else:
        await db.company_settings.insert_one(data)
    return {"message": "Company settings saved"}

@router.post("/settings/upload-logo")
async def upload_logo(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    upload_dir = ROOT_DIR / "uploads" / "logos"
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / "company_logo.png"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"message": "Logo uploaded"}

# Email Settings
@router.get("/settings/email")
async def get_email_settings(current_user: User = Depends(get_current_user)):
    settings = await db.email_settings.find_one({}, {"_id": 0})
    if settings and settings.get("password"):
        settings["password"] = "••••••••"
    return settings

@router.post("/settings/email")
async def save_email_settings(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.email_settings.find_one({})
    data = {
        "smtp_host": body.get("smtp_host", ""),
        "smtp_port": int(body.get("smtp_port", 587)),
        "username": body.get("username", ""),
        "from_email": body.get("from_email", ""),
        "use_tls": body.get("use_tls", True),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if body.get("password") and body["password"] != "••••••••":
        data["password"] = body["password"]
    if existing:
        await db.email_settings.update_one({}, {"$set": data})
    else:
        data["password"] = body.get("password", "")
        await db.email_settings.insert_one(data)
    return {"message": "Email settings saved"}

@router.post("/settings/email/test")
async def test_email(body: dict, current_user: User = Depends(get_current_user)):
    import aiosmtplib
    from email.mime.text import MIMEText
    settings = await db.email_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("smtp_host"):
        raise HTTPException(status_code=400, detail="Email not configured. Save settings first.")
    try:
        to_email = body.get("to_email", current_user.email)
        msg = MIMEText("This is a test email from SSC Track. Your email settings are working correctly!")
        msg["Subject"] = "SSC Track - Test Email"
        msg["From"] = settings.get("from_email", settings["username"])
        msg["To"] = to_email
        await aiosmtplib.send(msg, hostname=settings["smtp_host"], port=settings["smtp_port"],
                              username=settings["username"], password=settings["password"],
                              use_tls=settings.get("use_tls", True))
        return {"message": f"Test email sent to {to_email}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")

# WhatsApp Settings
@router.get("/settings/whatsapp")
async def get_whatsapp_settings(current_user: User = Depends(get_current_user)):
    settings = await db.whatsapp_config.find_one({}, {"_id": 0})
    if settings and settings.get("auth_token"):
        settings["auth_token"] = "••••••••"
    return settings

@router.post("/settings/whatsapp")
async def save_whatsapp_settings(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.whatsapp_config.find_one({})
    data = {
        "account_sid": body.get("account_sid", ""),
        "phone_number": body.get("phone_number", ""),
        "recipient_number": body.get("recipient_number", ""),
        "enabled": body.get("enabled", True),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if body.get("auth_token") and body["auth_token"] != "••••••••":
        data["auth_token"] = body["auth_token"]
    if existing:
        await db.whatsapp_config.update_one({}, {"$set": data})
    else:
        data["auth_token"] = body.get("auth_token", "")
        await db.whatsapp_config.insert_one(data)
    return {"message": "WhatsApp settings saved"}

# Notification Preferences
@router.get("/settings/notifications")
async def get_notification_prefs(current_user: User = Depends(get_current_user)):
    prefs = await db.notification_prefs.find_one({}, {"_id": 0})
    return prefs or {
        "email_daily_sales": False,
        "email_document_expiry": True,
        "email_leave_updates": False,
        "whatsapp_daily_sales": False,
        "whatsapp_document_expiry": False,
    }

@router.post("/settings/notifications")
async def save_notification_prefs(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.notification_prefs.find_one({})
    data = {
        "email_daily_sales": body.get("email_daily_sales", False),
        "email_document_expiry": body.get("email_document_expiry", True),
        "email_leave_updates": body.get("email_leave_updates", False),
        "whatsapp_daily_sales": body.get("whatsapp_daily_sales", False),
        "whatsapp_document_expiry": body.get("whatsapp_document_expiry", False),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if existing:
        await db.notification_prefs.update_one({}, {"$set": data})
    else:
        await db.notification_prefs.insert_one(data)
    return {"message": "Notification preferences saved"}

# Database Backup
@router.get("/backup/database")
async def backup_database(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    import json as json_module
    collections = ["users", "branches", "customers", "suppliers", "sales", "expenses",
                    "supplier_payments", "salary_payments", "employees", "leaves", "documents",
                    "categories", "invoices", "items", "cash_transfers", "notifications",
                    "employee_requests", "recurring_expenses", "email_settings", "whatsapp_config",
                    "notification_prefs", "whatsapp_settings", "stock_entries", "stock_usage", "job_titles", "shifts", "shift_assignments"]
    backup_data = {"backup_date": datetime.now(timezone.utc).isoformat(), "app": "SSC Track", "collections": {}}
    for col_name in collections:
        try:
            col = db[col_name]
            docs = await col.find({}, {"_id": 0}).to_list(100000)
            backup_data["collections"][col_name] = docs
        except:
            backup_data["collections"][col_name] = []
    json_str = json_module.dumps(backup_data, default=str, indent=2)
    buffer = BytesIO(json_str.encode('utf-8'))
    buffer.seek(0)
    fname = f"dataentry_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(buffer, media_type="application/json",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})

@router.post("/backup/restore")
async def restore_database(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    return {"message": "Use the backup JSON file to restore. Contact support for restore assistance."}

# Data Import
@router.post("/import/data")
async def import_data(file: UploadFile = File(...), data_type: str = Form(...), current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    content = await file.read()
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {str(e)}")
    records = df.where(df.notna(), None).to_dict('records')
    imported = 0
    errors = []
    for i, row in enumerate(records):
        try:
            clean = {k: (None if v is None or str(v).strip() == '' else v) for k, v in row.items()}
            if data_type == "customers":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "branch_id": None, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.customers.insert_one(doc)
                    imported += 1
            elif data_type == "suppliers":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "category": str(clean.get("category", "")) if clean.get("category") else None, "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "branch_id": None, "credit_limit": float(clean.get("credit_limit", 0) or 0), "current_credit": 0, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.suppliers.insert_one(doc)
                    imported += 1
            elif data_type == "employees":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "document_id": str(clean.get("document_id", "")) if clean.get("document_id") else None, "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "position": str(clean.get("position", "")) if clean.get("position") else None, "salary": float(clean.get("salary", 0) or 0), "branch_id": None, "loan_balance": 0, "annual_leave_entitled": 30, "sick_leave_entitled": 15, "ticket_entitled": 1, "ticket_years": 2, "ticket_used": 0, "active": True, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.employees.insert_one(doc)
                    imported += 1
            elif data_type == "items":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "unit_price": float(clean.get("unit_price", clean.get("price", 0)) or 0), "category": str(clean.get("category", "")) if clean.get("category") else None, "active": True, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.items.insert_one(doc)
                    imported += 1
            elif data_type == "branches":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "location": str(clean.get("location", "")) if clean.get("location") else None, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.branches.insert_one(doc)
                    imported += 1
            elif data_type == "sales":
                amt = float(clean.get("amount", 0) or 0)
                disc = float(clean.get("discount", 0) or 0)
                mode = str(clean.get("payment_mode", "cash") or "cash").lower()
                doc = {"id": str(uuid.uuid4()), "sale_type": str(clean.get("sale_type", "branch")), "branch_id": None, "customer_id": None, "amount": amt, "discount": disc, "final_amount": amt - disc, "payment_details": [{"mode": mode, "amount": amt - disc}], "credit_amount": 0, "credit_received": 0, "date": str(clean.get("date", datetime.now(timezone.utc).isoformat())), "notes": str(clean.get("notes", "") or ""), "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()}
                await db.sales.insert_one(doc)
                imported += 1
            elif data_type == "expenses_import":
                amt = float(clean.get("amount", 0) or 0)
                doc = {"id": str(uuid.uuid4()), "category": str(clean.get("category", "other")), "description": str(clean.get("description", "")), "amount": amt, "payment_mode": str(clean.get("payment_mode", "cash") or "cash"), "branch_id": None, "supplier_id": None, "date": str(clean.get("date", datetime.now(timezone.utc).isoformat())), "notes": "", "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["description"]:
                    await db.expenses.insert_one(doc)
                    imported += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)[:50]}")
    return {"message": f"Imported {imported} records", "imported": imported, "total_rows": len(records), "errors": errors[:10]}

@router.get("/import/template/{data_type}")
async def download_import_template(data_type: str, current_user: User = Depends(get_current_user)):
    templates = {
        "customers": ["name", "phone", "email"],
        "suppliers": ["name", "category", "phone", "email", "credit_limit"],
        "employees": ["name", "document_id", "phone", "email", "position", "salary"],
        "items": ["name", "unit_price", "category"],
        "branches": ["name", "location"],
        "sales": ["date", "sale_type", "amount", "discount", "payment_mode", "notes"],
        "expenses_import": ["date", "category", "description", "amount", "payment_mode"],
    }
    if data_type not in templates:
        raise HTTPException(status_code=400, detail="Invalid type")
    wb = Workbook()
    ws = wb.active
    ws.title = data_type.capitalize()
    ws.append(templates[data_type])
    for col in ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
    ws.append(["Example " + templates[data_type][0]] + ["" for _ in templates[data_type][1:]])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={data_type}_import_template.xlsx"})



# ZATCA Phase 2 Settings
@router.get("/settings/zatca")
async def get_zatca_settings(current_user: User = Depends(get_current_user)):
    """Get ZATCA Phase 2 configuration"""
    settings = await db.zatca_settings.find_one({}, {"_id": 0})
    if settings:
        # Mask sensitive fields
        if settings.get("csid_secret"):
            settings["csid_secret"] = "••••••••"
        if settings.get("production_secret"):
            settings["production_secret"] = "••••••••"
        if settings.get("private_key"):
            settings["private_key"] = "••••••••"
    return settings or {
        "enabled": False,
        "environment": "sandbox",
        "otp": "",
        "csid": "",
        "csid_secret": "",
        "production_csid": "",
        "production_secret": "",
        "certificate": "",
        "private_key": "",
        "auto_submit": False,
        "invoice_counter": 1
    }


@router.post("/settings/zatca")
async def save_zatca_settings(body: dict, current_user: User = Depends(get_current_user)):
    """Save ZATCA Phase 2 configuration"""
    existing = await db.zatca_settings.find_one({})
    
    data = {
        "enabled": body.get("enabled", False),
        "environment": body.get("environment", "sandbox"),
        "otp": body.get("otp", ""),
        "csid": body.get("csid", ""),
        "production_csid": body.get("production_csid", ""),
        "certificate": body.get("certificate", ""),
        "auto_submit": body.get("auto_submit", False),
        "invoice_counter": int(body.get("invoice_counter", 1) or 1),
        "csid_expiry": body.get("csid_expiry", ""),
        "production_csid_expiry": body.get("production_csid_expiry", ""),
        "expiry_alert_days": int(body.get("expiry_alert_days", 30) or 30),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.id
    }
    
    # Only update secrets if they're not masked
    if body.get("csid_secret") and body.get("csid_secret") != "••••••••":
        data["csid_secret"] = body["csid_secret"]
    elif existing and existing.get("csid_secret"):
        data["csid_secret"] = existing["csid_secret"]
    
    if body.get("production_secret") and body.get("production_secret") != "••••••••":
        data["production_secret"] = body["production_secret"]
    elif existing and existing.get("production_secret"):
        data["production_secret"] = existing["production_secret"]
    
    if body.get("private_key") and body.get("private_key") != "••••••••":
        data["private_key"] = body["private_key"]
    elif existing and existing.get("private_key"):
        data["private_key"] = existing["private_key"]
    
    if existing:
        await db.zatca_settings.update_one({}, {"$set": data})
    else:
        await db.zatca_settings.insert_one(data)
    
    return {"message": "ZATCA settings saved successfully"}


@router.post("/settings/zatca/test")
async def test_zatca_connection(current_user: User = Depends(get_current_user)):
    """Test ZATCA API connection with stored credentials"""
    settings = await db.zatca_settings.find_one({}, {"_id": 0})
    
    if not settings:
        return {"success": False, "message": "No ZATCA settings configured"}
    
    if not settings.get("enabled"):
        return {"success": False, "message": "ZATCA integration is disabled"}
    
    environment = settings.get("environment", "sandbox")
    
    if environment == "sandbox":
        csid = settings.get("csid", "")
        has_secret = bool(settings.get("csid_secret", ""))
    else:
        csid = settings.get("production_csid", "")
        has_secret = bool(settings.get("production_secret", ""))
    
    if not csid:
        return {"success": False, "message": f"No CSID configured for {environment} environment"}
    
    if not has_secret:
        return {"success": False, "message": f"No secret configured for {environment} environment"}
    
    # For now, we validate the format of credentials
    # Actual API test would require the ZATCA SDK or direct API call
    if len(csid) < 10:
        return {"success": False, "message": "CSID appears to be invalid (too short)"}
    
    # Check if it's a valid Base64-encoded string
    try:
        import base64
        base64.b64decode(csid)
    except Exception:
        return {"success": False, "message": "CSID is not valid Base64 encoded"}
    
    return {
        "success": True,
        "message": f"Credentials validated for {environment} environment",
        "environment": environment,
        "csid_configured": bool(csid),
        "note": "Full API test requires ZATCA SDK integration. Credentials appear to be properly formatted."
    }


@router.get("/settings/zatca/status")
async def get_zatca_status(current_user: User = Depends(get_current_user)):
    """Get ZATCA integration status summary"""
    settings = await db.zatca_settings.find_one({}, {"_id": 0})
    company = await db.company_settings.find_one({}, {"_id": 0})
    
    # Count invoices with ZATCA data
    total_invoices = await db.invoices.count_documents({})
    zatca_invoices = await db.invoices.count_documents({"uuid": {"$exists": True, "$ne": None}})
    submitted_invoices = await db.invoices.count_documents({"zatca_status": "submitted"})
    
    # Check CSID expiry
    expiry_status = None
    if settings:
        environment = settings.get("environment", "sandbox")
        expiry_date_str = settings.get("production_csid_expiry") if environment == "production" else settings.get("csid_expiry")
        
        if expiry_date_str:
            try:
                expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d")
                days_until_expiry = (expiry_date - datetime.now()).days
                alert_days = settings.get("expiry_alert_days", 30)
                
                expiry_status = {
                    "expiry_date": expiry_date_str,
                    "days_until_expiry": days_until_expiry,
                    "is_expired": days_until_expiry < 0,
                    "needs_renewal": days_until_expiry <= alert_days,
                    "alert_days": alert_days
                }
            except ValueError:
                pass
    
    return {
        "enabled": settings.get("enabled", False) if settings else False,
        "environment": settings.get("environment", "sandbox") if settings else "sandbox",
        "vat_enabled": company.get("vat_enabled", False) if company else False,
        "vat_number": company.get("vat_number", "") if company else "",
        "csid_configured": bool(settings.get("csid") or settings.get("production_csid")) if settings else False,
        "auto_submit": settings.get("auto_submit", False) if settings else False,
        "invoice_counter": settings.get("invoice_counter", 1) if settings else 1,
        "expiry_status": expiry_status,
        "statistics": {
            "total_invoices": total_invoices,
            "zatca_ready": zatca_invoices,
            "submitted": submitted_invoices
        }
    }


@router.post("/settings/zatca/check-expiry")
async def check_csid_expiry(current_user: User = Depends(get_current_user)):
    """Check CSID expiry and send alerts if needed"""
    settings = await db.zatca_settings.find_one({}, {"_id": 0})
    
    if not settings or not settings.get("enabled"):
        return {"success": False, "message": "ZATCA not enabled"}
    
    environment = settings.get("environment", "sandbox")
    expiry_date_str = settings.get("production_csid_expiry") if environment == "production" else settings.get("csid_expiry")
    
    if not expiry_date_str:
        return {"success": False, "message": "No expiry date configured"}
    
    try:
        expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d")
        days_until_expiry = (expiry_date - datetime.now()).days
        alert_days = settings.get("expiry_alert_days", 30)
        
        if days_until_expiry <= alert_days:
            # Create alert notification
            await create_csid_expiry_notification(days_until_expiry, expiry_date_str, environment)
            
            return {
                "success": True,
                "needs_renewal": True,
                "days_until_expiry": days_until_expiry,
                "message": f"CSID expires in {days_until_expiry} days. Please renew!"
            }
        
        return {
            "success": True,
            "needs_renewal": False,
            "days_until_expiry": days_until_expiry,
            "message": f"CSID is valid for {days_until_expiry} more days"
        }
        
    except ValueError:
        return {"success": False, "message": "Invalid expiry date format"}


async def create_csid_expiry_notification(days_until_expiry: int, expiry_date: str, environment: str):
    """Create notification for CSID expiry"""
    # Check if we already sent a notification today
    today = datetime.now().strftime("%Y-%m-%d")
    existing = await db.notifications.find_one({
        "type": "zatca_csid_expiry",
        "created_at": {"$regex": f"^{today}"}
    })
    
    if existing:
        return  # Already notified today
    
    if days_until_expiry < 0:
        title = "⚠️ ZATCA CSID Expired!"
        message = f"Your {environment} CSID expired on {expiry_date}. Please renew immediately to continue e-invoicing."
        priority = "critical"
    elif days_until_expiry == 0:
        title = "🚨 ZATCA CSID Expires Today!"
        message = f"Your {environment} CSID expires TODAY ({expiry_date}). Renew now to avoid invoice submission failures."
        priority = "critical"
    elif days_until_expiry <= 7:
        title = "⚠️ ZATCA CSID Expiring Soon!"
        message = f"Your {environment} CSID expires in {days_until_expiry} days ({expiry_date}). Please renew soon."
        priority = "high"
    else:
        title = "📋 ZATCA CSID Renewal Reminder"
        message = f"Your {environment} CSID will expire in {days_until_expiry} days ({expiry_date}). Consider renewing."
        priority = "medium"
    
    notification = {
        "id": f"zatca_expiry_{datetime.now().strftime('%Y%m%d%H%M%S')}",
        "type": "zatca_csid_expiry",
        "title": title,
        "message": message,
        "priority": priority,
        "data": {
            "days_until_expiry": days_until_expiry,
            "expiry_date": expiry_date,
            "environment": environment,
            "action_url": "/settings"
        },
        "read": False,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.notifications.insert_one(notification)
    
    # Also try to send WhatsApp/Email if configured
    try:
        # WhatsApp
        wa_config = await db.whatsapp_config.find_one({}, {"_id": 0})
        if wa_config and wa_config.get("enabled") and wa_config.get("account_sid"):
            from twilio.rest import Client
            client = Client(wa_config["account_sid"], wa_config["auth_token"])
            recipients = [r.strip() for r in wa_config.get("recipient_number", "").split(",") if r.strip()]
            wa_message = f"🔔 *ZATCA CSID Alert*\n\n{message}\n\nRenew at: https://fatoora.zatca.gov.sa/onboard"
            for recipient in recipients[:3]:  # Max 3 recipients
                try:
                    client.messages.create(
                        from_=f'whatsapp:{wa_config["phone_number"]}',
                        body=wa_message,
                        to=f'whatsapp:{recipient}'
                    )
                except Exception:
                    pass
    except Exception:
        pass
    
    return notification
