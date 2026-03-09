from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone, timedelta
from io import BytesIO
from pydantic import BaseModel
from typing import Optional
import uuid

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import db, get_current_user
from models import User

router = APIRouter(prefix="/export-center", tags=["export-center"])


class ExportRequest(BaseModel):
    report_type: str  # sales, expenses, supplier_payments, profit_loss, daily_summary
    format: str = "excel"  # excel or pdf
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    branch_id: Optional[str] = None


def parse_date(date_str):
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str.replace("Z", "+00:00"))
    except Exception:
        return datetime.strptime(date_str, "%Y-%m-%d")


def filter_by_date(items, start_date, end_date):
    filtered = []
    for item in items:
        try:
            d = item.get("date", "")
            if isinstance(d, str):
                item_date = datetime.fromisoformat(d.replace("Z", "+00:00")).replace(tzinfo=None)
            else:
                item_date = d.replace(tzinfo=None) if d.tzinfo else d
            s = start_date.replace(tzinfo=None) if start_date and start_date.tzinfo else start_date
            e = end_date.replace(tzinfo=None) if end_date and end_date.tzinfo else end_date
            if s and item_date < s:
                continue
            if e and item_date > e + timedelta(days=1):
                continue
            filtered.append(item)
        except Exception:
            filtered.append(item)
    return filtered


def filter_by_branch(items, branch_id, branch_field="branch_id"):
    if not branch_id:
        return items
    return [i for i in items if i.get(branch_field) == branch_id]


@router.get("/report-types")
async def get_report_types(current_user: User = Depends(get_current_user)):
    return [
        {"id": "sales", "name": "Sales Report", "description": "All sales with payment breakdowns", "icon": "ShoppingCart", "color": "emerald"},
        {"id": "expenses", "name": "Expenses Report", "description": "All expenses by category and payment mode", "icon": "Receipt", "color": "red"},
        {"id": "supplier_payments", "name": "Supplier Payments", "description": "All supplier payments and purchase bills", "icon": "Truck", "color": "blue"},
        {"id": "profit_loss", "name": "Profit & Loss", "description": "Revenue vs expenses with net profit", "icon": "TrendingUp", "color": "amber"},
        {"id": "daily_summary", "name": "Daily Summary", "description": "Day-by-day sales, expenses, and profit", "icon": "CalendarDays", "color": "purple"},
        {"id": "customers", "name": "Customers", "description": "Customer list with contact details", "icon": "Users", "color": "teal"},
        {"id": "employees", "name": "Employees", "description": "Employee list with salary and document info", "icon": "UserCheck", "color": "indigo"},
        {"id": "stock", "name": "Inventory", "description": "Current stock levels and values", "icon": "Package", "color": "orange"},
    ]


@router.get("/history")
async def get_export_history(current_user: User = Depends(get_current_user)):
    history = await db.export_history.find(
        {}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    return history


@router.post("/generate")
async def generate_export(request: ExportRequest, current_user: User = Depends(get_current_user)):
    start = parse_date(request.start_date) if request.start_date else None
    end = parse_date(request.end_date) if request.end_date else None
    branch_id = request.branch_id
    report_type = request.report_type
    fmt = request.format

    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}
    branch_name = branch_map.get(branch_id, "All Branches") if branch_id else "All Branches"

    date_label = ""
    if start and end:
        date_label = f"{start.strftime('%Y-%m-%d')} to {end.strftime('%Y-%m-%d')}"
    elif start:
        date_label = f"From {start.strftime('%Y-%m-%d')}"
    elif end:
        date_label = f"Until {end.strftime('%Y-%m-%d')}"
    else:
        date_label = "All Time"

    rows = []
    headers = []
    title = ""

    if report_type == "sales":
        sales = await db.sales.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
        if start or end:
            sales = filter_by_date(sales, start, end)
        if branch_id:
            sales = filter_by_branch(sales, branch_id)
        headers = ["Date", "Type", "Branch", "Amount", "Discount", "Final", "Cash", "Bank", "Online", "Credit", "Notes"]
        title = "Sales Report"
        for s in sales:
            try:
                d = datetime.fromisoformat(s["date"]).strftime("%Y-%m-%d") if isinstance(s["date"], str) else s["date"].strftime("%Y-%m-%d")
            except Exception:
                d = str(s.get("date", "-"))
            cash = bank = online = credit = 0
            for p in s.get("payment_details", []):
                amt = p.get("amount", 0)
                if p["mode"] == "cash":
                    cash += amt
                elif p["mode"] == "bank":
                    bank += amt
                elif p["mode"] == "credit":
                    credit += amt
                else:
                    online += amt
            rows.append([d, s.get("sale_type", "branch"), branch_map.get(s.get("branch_id"), "-"),
                         s.get("amount", 0), s.get("discount", 0), s.get("final_amount", s.get("amount", 0) - s.get("discount", 0)),
                         cash, bank, online, credit, s.get("notes", "")])

    elif report_type == "expenses":
        expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
        if start or end:
            expenses = filter_by_date(expenses, start, end)
        if branch_id:
            expenses = filter_by_branch(expenses, branch_id)
        headers = ["Date", "Category", "Description", "Branch", "Amount", "Payment Mode", "Notes"]
        title = "Expenses Report"
        for e in expenses:
            try:
                d = datetime.fromisoformat(e["date"]).strftime("%Y-%m-%d") if isinstance(e["date"], str) else e["date"].strftime("%Y-%m-%d")
            except Exception:
                d = str(e.get("date", "-"))
            rows.append([d, e.get("category", "-"), e.get("description", "-"),
                         branch_map.get(e.get("branch_id"), "-"), e.get("amount", 0),
                         e.get("payment_mode", "-"), e.get("notes", "")])

    elif report_type == "supplier_payments":
        payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).sort("date", -1).to_list(10000)
        if start or end:
            payments = filter_by_date(payments, start, end)
        if branch_id:
            payments = filter_by_branch(payments, branch_id)
        headers = ["Date", "Supplier", "Branch", "Amount", "Payment Mode", "Type", "Notes"]
        title = "Supplier Payments Report"
        for p in payments:
            try:
                d = datetime.fromisoformat(p["date"]).strftime("%Y-%m-%d") if isinstance(p["date"], str) else p["date"].strftime("%Y-%m-%d")
            except Exception:
                d = str(p.get("date", "-"))
            rows.append([d, p.get("supplier_name", "-"), branch_map.get(p.get("branch_id"), "-"),
                         p.get("amount", 0), p.get("payment_mode", "-"),
                         p.get("type", "payment"), p.get("notes", "")])

    elif report_type == "profit_loss":
        sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
        if start or end:
            sales = filter_by_date(sales, start, end)
            expenses = filter_by_date(expenses, start, end)
        if branch_id:
            sales = filter_by_branch(sales, branch_id)
            expenses = filter_by_branch(expenses, branch_id)

        total_sales = sum(s.get("final_amount", s.get("amount", 0) - s.get("discount", 0)) for s in sales)
        total_expenses = sum(e.get("amount", 0) for e in expenses)
        net_profit = total_sales - total_expenses

        # Break down by category
        cat_totals = {}
        for e in expenses:
            cat = e.get("category", "Other")
            cat_totals[cat] = cat_totals.get(cat, 0) + e.get("amount", 0)

        headers = ["Item", "Amount (SAR)", "% of Total"]
        title = "Profit & Loss Statement"
        rows.append(["REVENUE", "", ""])
        rows.append(["Total Sales", f"{total_sales:,.2f}", "100%"])
        rows.append(["", "", ""])
        rows.append(["EXPENSES", "", ""])
        for cat, amt in sorted(cat_totals.items(), key=lambda x: -x[1]):
            pct = f"{(amt / total_expenses * 100):.1f}%" if total_expenses > 0 else "0%"
            rows.append([f"  {cat}", f"{amt:,.2f}", pct])
        rows.append(["Total Expenses", f"{total_expenses:,.2f}", f"{(total_expenses / total_sales * 100):.1f}%" if total_sales > 0 else "N/A"])
        rows.append(["", "", ""])
        rows.append(["NET PROFIT", f"{net_profit:,.2f}", f"{(net_profit / total_sales * 100):.1f}%" if total_sales > 0 else "N/A"])

    elif report_type == "daily_summary":
        sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
        if start or end:
            sales = filter_by_date(sales, start, end)
            expenses = filter_by_date(expenses, start, end)
        if branch_id:
            sales = filter_by_branch(sales, branch_id)
            expenses = filter_by_branch(expenses, branch_id)

        daily = {}
        for s in sales:
            try:
                d = datetime.fromisoformat(s["date"]).strftime("%Y-%m-%d") if isinstance(s["date"], str) else s["date"].strftime("%Y-%m-%d")
            except Exception:
                continue
            if d not in daily:
                daily[d] = {"sales": 0, "expenses": 0, "cash_in": 0, "bank_in": 0}
            daily[d]["sales"] += s.get("final_amount", s.get("amount", 0) - s.get("discount", 0))
            for p in s.get("payment_details", []):
                if p["mode"] == "cash":
                    daily[d]["cash_in"] += p.get("amount", 0)
                elif p["mode"] == "bank":
                    daily[d]["bank_in"] += p.get("amount", 0)
        for e in expenses:
            try:
                d = datetime.fromisoformat(e["date"]).strftime("%Y-%m-%d") if isinstance(e["date"], str) else e["date"].strftime("%Y-%m-%d")
            except Exception:
                continue
            if d not in daily:
                daily[d] = {"sales": 0, "expenses": 0, "cash_in": 0, "bank_in": 0}
            daily[d]["expenses"] += e.get("amount", 0)

        headers = ["Date", "Sales (SAR)", "Expenses (SAR)", "Net Profit (SAR)", "Cash In", "Bank In"]
        title = "Daily Summary Report"
        for d in sorted(daily.keys(), reverse=True):
            v = daily[d]
            rows.append([d, f"{v['sales']:,.2f}", f"{v['expenses']:,.2f}",
                         f"{v['sales'] - v['expenses']:,.2f}", f"{v['cash_in']:,.2f}", f"{v['bank_in']:,.2f}"])

    elif report_type == "customers":
        customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
        headers = ["Name", "Branch", "Phone", "Email", "Credit Balance"]
        title = "Customers Report"
        for c in customers:
            rows.append([c.get("name", "-"), branch_map.get(c.get("branch_id"), "All"),
                         c.get("phone", "-"), c.get("email", "-"), c.get("credit_balance", 0)])

    elif report_type == "employees":
        employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
        headers = ["Name", "Position", "Branch", "Salary", "Pay Frequency", "Document Expiry"]
        title = "Employees Report"
        for emp in employees:
            try:
                exp = datetime.fromisoformat(emp["document_expiry"]).strftime("%Y-%m-%d") if emp.get("document_expiry") else "-"
            except Exception:
                exp = "-"
            rows.append([emp.get("name", "-"), emp.get("position", "-"),
                         branch_map.get(emp.get("branch_id"), "-"), emp.get("salary", 0),
                         emp.get("pay_frequency", "monthly"), exp])

    elif report_type == "stock":
        items = await db.stock.find({}, {"_id": 0}).to_list(5000)
        headers = ["Item", "Category", "Branch", "Current Qty", "Unit", "Min Level", "Value (SAR)"]
        title = "Inventory Report"
        for item in items:
            if branch_id and item.get("branch_id") != branch_id:
                continue
            rows.append([item.get("name", "-"), item.get("category", "-"),
                         branch_map.get(item.get("branch_id"), "-"),
                         item.get("current_balance", 0), item.get("unit", "-"),
                         item.get("min_level", 0),
                         round(item.get("current_balance", 0) * item.get("unit_price", 0), 2)])
    else:
        raise HTTPException(status_code=400, detail="Invalid report type")

    # Log export history
    await db.export_history.insert_one({
        "id": str(uuid.uuid4()),
        "report_type": report_type,
        "title": title,
        "format": fmt,
        "date_range": date_label,
        "branch": branch_name,
        "row_count": len(rows),
        "user_name": current_user.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    subtitle = f"{date_label} | {branch_name}"
    buffer = BytesIO()

    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]
        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = Font(bold=True, size=14, color="E8501A")
        title_cell.alignment = Alignment(horizontal="center")
        # Subtitle row
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = Font(size=10, color="666666")
        sub_cell.alignment = Alignment(horizontal="center")
        # Headers
        header_fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        # Data rows
        alt_fill = PatternFill(start_color="FFF8F0", end_color="FFF8F0", fill_type="solid")
        for row_idx, row in enumerate(rows):
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx + 5, column=col_idx, value=val)
                cell.border = thin_border
                if row_idx % 2 == 1:
                    cell.fill = alt_fill
                # Right-align numeric columns
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
        # Auto-width - handle merged cells safely
        for col_idx, col_cells in enumerate(ws.columns, 1):
            # Skip merged cells by checking the first cell
            first_cell = col_cells[0]
            if hasattr(first_cell, 'column_letter'):
                col_letter = first_cell.column_letter
            else:
                # Use openpyxl utility to get column letter
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(col_idx)
            # Calculate max length for regular cells only
            max_len = 0
            for cell in col_cells:
                if hasattr(cell, 'value') and cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            if max_len > 0:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 35)
        # Footer
        footer_row = len(rows) + 6
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=len(headers))
        footer = ws.cell(row=footer_row, column=1, value=f"Generated by SSC Track on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | {len(rows)} records")
        footer.font = Font(size=8, color="999999", italic=True)

        wb.save(buffer)
        buffer.seek(0)
        filename = f"{report_type}_report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(buffer,
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})
    else:
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('ExportTitle', parent=styles['Heading1'], fontSize=18,
                                     textColor=colors.HexColor('#E8501A'), alignment=1, spaceAfter=4)
        sub_style = ParagraphStyle('ExportSub', parent=styles['Normal'], fontSize=9,
                                   textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=16)
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(subtitle, sub_style))

        if rows:
            table_data = [headers] + [[str(c) for c in row] for row in rows[:200]]
            col_count = len(headers)
            col_width = 7.2 * inch / col_count
            t = Table(table_data, colWidths=[col_width] * col_count)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DDDDDD')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF8F0')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(t)

        footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                                      textColor=colors.HexColor('#999999'), alignment=1, spaceBefore=20)
        elements.append(Paragraph(
            f"Generated by SSC Track on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | {len(rows)} records",
            footer_style))

        doc.build(elements)
        buffer.seek(0)
        filename = f"{report_type}_report_{datetime.now(timezone.utc).strftime('%Y%m%d')}.pdf"
        return StreamingResponse(buffer, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})
