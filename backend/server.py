from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import shutil
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from io import BytesIO
from twilio.rest import Client

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

SECRET_KEY = os.environ.get("JWT_SECRET_KEY", "your-secret-key-change-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 10080

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str = "operator"  # "admin", "manager", "operator"
    branch_id: Optional[str] = None  # For branch-specific access
    permissions: List[str] = []  # ["sales", "expenses", "reports", "branches", "customers", "suppliers", "users"]
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Optional[str] = "operator"
    branch_id: Optional[str] = None
    permissions: Optional[List[str]] = []

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    branch_id: Optional[str] = None
    permissions: Optional[List[str]] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Branch(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    location: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class BranchCreate(BaseModel):
    name: str
    location: Optional[str] = None

class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    branch_id: Optional[str] = None  # None means available to all branches
    phone: Optional[str] = None
    email: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomerCreate(BaseModel):
    name: str
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None

class Sale(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    sale_type: str  # "branch" or "online"
    branch_id: Optional[str] = None
    customer_id: Optional[str] = None
    amount: float  # Original amount before discount
    discount: float = 0  # Discount amount
    final_amount: float = 0  # Amount after discount
    payment_details: Optional[List[dict]] = []  # [{"mode": "cash", "amount": 100}, {"mode": "bank", "amount": 50}]
    credit_amount: float = 0  # Amount still pending for credit sales
    credit_received: float = 0  # Amount received against credit
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    
    # Legacy fields for backward compatibility
    payment_mode: Optional[str] = None
    payment_status: Optional[str] = None
    received_mode: Optional[str] = None

class SaleCreate(BaseModel):
    sale_type: str
    branch_id: Optional[str] = None
    customer_id: Optional[str] = None
    amount: float
    discount: Optional[float] = 0
    payment_details: List[dict]
    date: datetime
    notes: Optional[str] = None

class SalePayment(BaseModel):
    payment_mode: str  # "cash" or "bank"
    amount: float
    discount: Optional[float] = 0  # Discount when receiving credit

class Supplier(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: Optional[str] = None
    sub_category: Optional[str] = None
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    credit_limit: Optional[float] = 0
    current_credit: Optional[float] = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SupplierCreate(BaseModel):
    name: str
    category: Optional[str] = None
    sub_category: Optional[str] = None
    branch_id: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    credit_limit: Optional[float] = 0

class SupplierPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    supplier_id: str
    supplier_name: str
    amount: float
    payment_mode: str  # "cash", "bank", "credit"
    branch_id: Optional[str] = None  # Which branch cash/bank is used
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class SupplierPaymentCreate(BaseModel):
    supplier_id: str
    amount: float
    payment_mode: str
    branch_id: Optional[str] = None
    date: datetime
    notes: Optional[str] = None

class SupplierCreditPayment(BaseModel):
    payment_mode: str  # "cash" or "bank"
    amount: float
    branch_id: Optional[str] = None

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    category: str
    sub_category: Optional[str] = None
    description: str
    amount: float
    payment_mode: str
    branch_id: Optional[str] = None
    supplier_id: Optional[str] = None
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class ExpenseCreate(BaseModel):
    category: str
    sub_category: Optional[str] = None
    description: str
    amount: float
    payment_mode: str
    branch_id: Optional[str] = None
    supplier_id: Optional[str] = None
    date: datetime
    notes: Optional[str] = None

class DashboardStats(BaseModel):
    total_sales: float
    total_expenses: float
    total_supplier_payments: float
    net_profit: float
    pending_credits: float
    cash_sales: float
    bank_sales: float
    credit_sales: float

class WhatsAppSettings(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    phone_number: str
    enabled: bool = True
    notification_time: str = "09:00"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class WhatsAppSettingsCreate(BaseModel):
    phone_number: str
    enabled: bool = True
    notification_time: str = "09:00"

class ExportRequest(BaseModel):
    format: str  # "pdf" or "excel"
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    branch_id: Optional[str] = None

class Category(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    type: str  # "supplier" or "expense"
    parent_id: Optional[str] = None  # For sub-categories
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CategoryCreate(BaseModel):
    name: str
    type: str
    parent_id: Optional[str] = None
    description: Optional[str] = None

# Employee / Payroll Models
class Employee(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    document_id: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    position: Optional[str] = None
    branch_id: Optional[str] = None
    user_id: Optional[str] = None  # Linked user account for portal access
    salary: float = 0
    pay_frequency: str = "monthly"
    join_date: Optional[datetime] = None
    document_expiry: Optional[datetime] = None
    loan_balance: float = 0
    old_salary_balance: float = 0  # Old unpaid salary balance
    annual_leave_entitled: int = 30
    sick_leave_entitled: int = 15
    ticket_entitled: int = 1  # Tickets per 2 years
    ticket_years: int = 2  # Every X years
    ticket_used: int = 0
    notes: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeCreate(BaseModel):
    name: str
    document_id: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    position: Optional[str] = None
    branch_id: Optional[str] = None
    salary: float = 0
    pay_frequency: Optional[str] = "monthly"
    join_date: Optional[datetime] = None
    document_expiry: Optional[datetime] = None
    annual_leave_entitled: Optional[int] = 30
    sick_leave_entitled: Optional[int] = 15
    ticket_entitled: Optional[int] = 1
    ticket_years: Optional[int] = 2
    notes: Optional[str] = None

# Document Expiry Tracking
class Document(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    document_type: str
    document_number: Optional[str] = None
    related_to: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: datetime
    alert_days: int = 30
    file_path: Optional[str] = None  # Stored file path
    file_name: Optional[str] = None  # Original file name
    notes: Optional[str] = None
    status: str = "active"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DocumentCreate(BaseModel):
    name: str
    document_type: str
    document_number: Optional[str] = None
    related_to: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: datetime
    alert_days: Optional[int] = 30
    notes: Optional[str] = None

# Salary Payment Tracking
class SalaryPayment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_name: str
    payment_type: str = "salary"
    amount: float
    payment_mode: str
    branch_id: Optional[str] = None
    period: str
    date: datetime
    notes: Optional[str] = None
    acknowledged: bool = False
    acknowledged_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class SalaryPaymentCreate(BaseModel):
    employee_id: str
    payment_type: Optional[str] = "salary"
    amount: float
    payment_mode: str
    branch_id: Optional[str] = None
    period: str
    date: datetime
    notes: Optional[str] = None

# Leave Management
class Leave(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_name: str
    leave_type: str
    start_date: datetime
    end_date: datetime
    days: int
    with_ticket: bool = False  # Request ticket with leave
    reason: Optional[str] = None
    status: str = "pending"  # "pending", "branch_approved", "manager_approved", "approved", "rejected"
    approved_by: Optional[str] = None
    approved_at: Optional[datetime] = None
    rejection_reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class LeaveCreate(BaseModel):
    employee_id: str
    leave_type: str
    start_date: datetime
    end_date: datetime
    days: int
    with_ticket: Optional[bool] = False
    reason: Optional[str] = None
    status: Optional[str] = "pending"

# Employee Request (letter, loan, etc.)
class EmployeeRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_name: str
    request_type: str  # "letter", "loan", "salary_advance", "other"
    subject: str
    details: Optional[str] = None
    amount: Optional[float] = None  # For loan/advance requests
    status: str = "pending"  # "pending", "approved", "rejected"
    response: Optional[str] = None
    processed_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeRequestCreate(BaseModel):
    request_type: str
    subject: str
    details: Optional[str] = None
    amount: Optional[float] = None

# Notification Model
class Notification(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str  # Target user
    title: str
    message: str
    type: str  # "leave_approved", "leave_rejected", "salary_paid", "document_expiry"
    read: bool = False
    related_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Cash Transfer Model
class CashTransfer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    from_branch_id: Optional[str] = None
    to_branch_id: Optional[str] = None
    from_branch_name: Optional[str] = None
    to_branch_name: Optional[str] = None
    amount: float
    transfer_mode: str = "cash"  # "cash", "bank"
    sender_name: str
    receiver_name: str
    date: datetime
    notes: Optional[str] = None
    status: str = "completed"  # "pending", "completed"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class CashTransferCreate(BaseModel):
    from_branch_id: Optional[str] = None
    to_branch_id: Optional[str] = None
    amount: float
    transfer_mode: Optional[str] = "cash"
    sender_name: str
    receiver_name: str
    date: datetime
    notes: Optional[str] = None

# Invoice Model
class InvoiceItem(BaseModel):
    description: str
    quantity: float = 1
    unit_price: float
    total: float = 0

class Invoice(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: str
    branch_id: Optional[str] = None
    customer_id: Optional[str] = None
    customer_name: Optional[str] = None
    items: List[dict] = []
    subtotal: float = 0
    discount: float = 0
    total: float = 0
    payment_mode: str = "cash"  # "cash", "bank", "credit"
    payment_details: List[dict] = []
    sale_id: Optional[str] = None  # Linked sale entry
    date: datetime
    notes: Optional[str] = None
    status: str = "paid"  # "draft", "paid", "cancelled"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class InvoiceCreate(BaseModel):
    branch_id: Optional[str] = None
    customer_id: Optional[str] = None
    items: List[dict]
    discount: Optional[float] = 0
    payment_mode: str = "cash"
    payment_details: Optional[List[dict]] = None
    date: datetime
    notes: Optional[str] = None

# Product/Service Items Master
class Item(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    unit_price: float = 0
    category: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ItemCreate(BaseModel):
    name: str
    description: Optional[str] = None
    unit_price: float = 0
    category: Optional[str] = None

# Recurring Expense
class RecurringExpense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: str
    amount: float
    frequency: str = "monthly"  # monthly, quarterly, yearly
    branch_id: Optional[str] = None
    next_due_date: datetime
    alert_days: int = 7
    notes: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class RecurringExpenseCreate(BaseModel):
    name: str
    category: str
    amount: float
    frequency: Optional[str] = "monthly"
    branch_id: Optional[str] = None
    next_due_date: datetime
    alert_days: Optional[int] = 7
    notes: Optional[str] = None

# Attendance Model
class Attendance(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_name: str
    date: str  # "2026-02-11"
    time_in: Optional[datetime] = None
    time_out: Optional[datetime] = None
    status: str = "present"  # "present", "late", "absent"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Employee Document (multiple per employee)
class EmployeeDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    document_type: str  # "passport", "visa", "labor_card", "emirates_id", "health_card", "other"
    document_number: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class EmployeeDocumentCreate(BaseModel):
    employee_id: str
    document_type: str
    document_number: Optional[str] = None
    issue_date: Optional[datetime] = None
    expiry_date: Optional[datetime] = None
    notes: Optional[str] = None

# Government Fine / Penalty
class Fine(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    fine_type: str  # "government", "traffic", "labor", "municipality", "other"
    department: str
    description: str
    amount: float
    branch_id: Optional[str] = None
    employee_id: Optional[str] = None  # If fine is charged to employee
    payment_status: str = "unpaid"  # "unpaid", "paid", "partial"
    paid_amount: float = 0
    payment_mode: Optional[str] = None
    fine_date: datetime
    due_date: Optional[datetime] = None
    paid_date: Optional[datetime] = None
    deduct_from_salary: bool = False
    monthly_deduction: float = 0  # Monthly amount to deduct
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FineCreate(BaseModel):
    fine_type: str
    department: str
    description: str
    amount: float
    branch_id: Optional[str] = None
    employee_id: Optional[str] = None
    fine_date: datetime
    due_date: Optional[datetime] = None
    deduct_from_salary: Optional[bool] = False
    monthly_deduction: Optional[float] = 0
    notes: Optional[str] = None

# Salary Deduction
class SalaryDeduction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    employee_name: str
    deduction_type: str  # "fine", "late", "absence", "misbehavior", "damage", "other"
    amount: float
    period: str
    reason: str
    fine_id: Optional[str] = None  # Linked fine if applicable
    branch_id: Optional[str] = None
    date: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class SalaryDeductionCreate(BaseModel):
    employee_id: str
    deduction_type: str
    amount: float
    period: str
    reason: str
    fine_id: Optional[str] = None
    branch_id: Optional[str] = None
    date: datetime

# Salary History (increments tracking)
class SalaryHistory(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    old_salary: float
    new_salary: float
    effective_date: datetime
    reason: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

# Branch Payback (settle cross-branch dues)
class BranchPayback(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    from_branch_id: str
    to_branch_id: str
    from_branch_name: str
    to_branch_name: str
    amount: float
    payment_mode: str = "cash"
    date: datetime
    notes: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

# Partner Model
class Partner(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    share_percentage: float = 0
    notes: Optional[str] = None
    active: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PartnerCreate(BaseModel):
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    share_percentage: Optional[float] = 0
    notes: Optional[str] = None

class PartnerTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    partner_id: str
    partner_name: str
    transaction_type: str  # "investment", "withdrawal", "profit_share", "expense", "other"
    amount: float
    payment_mode: str = "cash"
    branch_id: Optional[str] = None
    description: Optional[str] = None
    date: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class PartnerTransactionCreate(BaseModel):
    partner_id: str
    transaction_type: str
    amount: float
    payment_mode: Optional[str] = "cash"
    branch_id: Optional[str] = None
    description: Optional[str] = None
    date: datetime

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    return User(**user)

# Health Check
@api_router.get("/")
async def root():
    return {"message": "SSC Track API"}

# Auth Routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # First user becomes admin
    user_count = await db.users.count_documents({})
    role = "admin" if user_count == 0 else user_data.role or "operator"
    
    # Set default permissions based on role
    if role == "admin":
        permissions = ["sales", "expenses", "reports", "branches", "customers", "suppliers", "users"]
    elif role == "manager":
        permissions = ["sales", "expenses", "reports", "branches", "customers", "suppliers"]
    else:
        permissions = ["sales", "expenses"]
    
    user = User(
        email=user_data.email, 
        name=user_data.name,
        role=role,
        branch_id=user_data.branch_id,
        permissions=user_data.permissions or permissions
    )
    user_dict = user.model_dump()
    user_dict["password"] = hash_password(user_data.password)
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    
    await db.users.insert_one(user_dict)
    
    access_token = create_access_token(data={"sub": user.id})
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user_doc = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(credentials.password, user_doc["password"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    user = User(**{k: v for k, v in user_doc.items() if k != "password"})
    access_token = create_access_token(data={"sub": user.id})
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# User Management Routes (Admin only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    return users

# Category Management Routes
@api_router.get("/categories")
async def get_categories(category_type: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if category_type:
        query["type"] = category_type
    
    categories = await db.categories.find(query, {"_id": 0}).to_list(1000)
    for category in categories:
        if isinstance(category.get('created_at'), str):
            category['created_at'] = datetime.fromisoformat(category['created_at'])
    return categories

@api_router.post("/categories", response_model=Category)
async def create_category(category_data: CategoryCreate, current_user: User = Depends(get_current_user)):
    # Check if category already exists
    existing = await db.categories.find_one({"name": category_data.name, "type": category_data.type}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    
    category = Category(**category_data.model_dump())
    category_dict = category.model_dump()
    category_dict["created_at"] = category_dict["created_at"].isoformat()
    await db.categories.insert_one(category_dict)
    return category

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: User = Depends(get_current_user)):
    result = await db.categories.delete_one({"id": category_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted successfully"}

# User Management Routes (Admin only) - Continued
@api_router.post("/users", response_model=User)
async def create_user(user_data: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing_user = await db.users.find_one({"email": user_data.email}, {"_id": 0})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Set permissions based on role
    if user_data.role == "admin":
        permissions = ["sales", "expenses", "reports", "branches", "customers", "suppliers", "users"]
    elif user_data.role == "manager":
        permissions = ["sales", "expenses", "reports", "branches", "customers", "suppliers"]
    else:
        permissions = ["sales", "expenses"]
    
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role or "operator",
        branch_id=user_data.branch_id,
        permissions=user_data.permissions or permissions
    )
    user_dict = user.model_dump()
    user_dict["password"] = hash_password(user_data.password)
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    
    await db.users.insert_one(user_dict)
    return user

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_update: UserUpdate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    result = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {k: v for k, v in user_update.model_dump().items() if v is not None}
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return User(**updated)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"message": "User deleted successfully"}

# Branch Routes
@api_router.get("/branches", response_model=List[Branch])
async def get_branches(current_user: User = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).to_list(1000)
    for branch in branches:
        if isinstance(branch.get('created_at'), str):
            branch['created_at'] = datetime.fromisoformat(branch['created_at'])
    return branches

@api_router.post("/branches", response_model=Branch)
async def create_branch(branch_data: BranchCreate, current_user: User = Depends(get_current_user)):
    branch = Branch(**branch_data.model_dump())
    branch_dict = branch.model_dump()
    branch_dict["created_at"] = branch_dict["created_at"].isoformat()
    await db.branches.insert_one(branch_dict)
    return branch

@api_router.put("/branches/{branch_id}", response_model=Branch)
async def update_branch(branch_id: str, branch_data: BranchCreate, current_user: User = Depends(get_current_user)):
    result = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    await db.branches.update_one({"id": branch_id}, {"$set": branch_data.model_dump()})
    updated = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Branch(**updated)

@api_router.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, current_user: User = Depends(get_current_user)):
    result = await db.branches.delete_one({"id": branch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"message": "Branch deleted successfully"}

# Supplier Routes
@api_router.get("/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: User = Depends(get_current_user)):
    # If user has branch_id, filter suppliers by branch
    query = {}
    if current_user.branch_id and current_user.role != "admin":
        query["branch_id"] = current_user.branch_id
    
    suppliers = await db.suppliers.find(query, {"_id": 0}).to_list(1000)
    for supplier in suppliers:
        if isinstance(supplier.get('created_at'), str):
            supplier['created_at'] = datetime.fromisoformat(supplier['created_at'])
    return suppliers

@api_router.post("/suppliers", response_model=Supplier)
async def create_supplier(supplier_data: SupplierCreate, current_user: User = Depends(get_current_user)):
    data = supplier_data.model_dump()
    for f in ['branch_id', 'category', 'sub_category', 'phone', 'email']:
        if data.get(f) == '':
            data[f] = None
    supplier = Supplier(**data)
    supplier_dict = supplier.model_dump()
    supplier_dict["created_at"] = supplier_dict["created_at"].isoformat()
    await db.suppliers.insert_one(supplier_dict)
    return supplier

@api_router.put("/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(supplier_id: str, supplier_data: SupplierCreate, current_user: User = Depends(get_current_user)):
    result = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    update_data = supplier_data.model_dump()
    for f in ['branch_id', 'category', 'sub_category', 'phone', 'email']:
        if update_data.get(f) == '':
            update_data[f] = None
    await db.suppliers.update_one({"id": supplier_id}, {"$set": update_data})
    updated = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Supplier(**updated)

@api_router.delete("/suppliers/{supplier_id}")
async def delete_supplier(supplier_id: str, current_user: User = Depends(get_current_user)):
    result = await db.suppliers.delete_one({"id": supplier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}

@api_router.post("/suppliers/{supplier_id}/pay-credit")
async def pay_supplier_credit(supplier_id: str, payment: SupplierCreditPayment, current_user: User = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    if payment.amount > supplier["current_credit"]:
        raise HTTPException(status_code=400, detail="Payment amount exceeds current credit")
    
    # Update supplier credit
    new_credit = supplier["current_credit"] - payment.amount
    await db.suppliers.update_one({"id": supplier_id}, {"$set": {"current_credit": new_credit}})
    
    # Record payment
    payment_record = SupplierPayment(
        supplier_id=supplier_id,
        supplier_name=supplier["name"],
        amount=payment.amount,
        payment_mode=payment.payment_mode,
        branch_id=payment.branch_id,
        date=datetime.now(timezone.utc),
        notes=f"Credit payment - Remaining: SAR {new_credit:.2f}",
        created_by=current_user.id
    )
    payment_dict = payment_record.model_dump()
    payment_dict["date"] = payment_dict["date"].isoformat()
    payment_dict["created_at"] = payment_dict["created_at"].isoformat()
    await db.supplier_payments.insert_one(payment_dict)
    
    return {"message": "Credit payment recorded", "remaining_credit": new_credit}

# Customer Routes
@api_router.get("/customers", response_model=List[Customer])
async def get_customers(current_user: User = Depends(get_current_user)):
    query = {}
    # Show customers that belong to user's branch or all branches (branch_id = None)
    if current_user.branch_id and current_user.role != "admin":
        query["$or"] = [{"branch_id": current_user.branch_id}, {"branch_id": None}]
    
    customers = await db.customers.find(query, {"_id": 0}).to_list(1000)
    for customer in customers:
        if isinstance(customer.get('created_at'), str):
            customer['created_at'] = datetime.fromisoformat(customer['created_at'])
    return customers

@api_router.post("/customers", response_model=Customer)
async def create_customer(customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    customer = Customer(**customer_data.model_dump())
    customer_dict = customer.model_dump()
    customer_dict["created_at"] = customer_dict["created_at"].isoformat()
    await db.customers.insert_one(customer_dict)
    return customer

@api_router.put("/customers/{customer_id}", response_model=Customer)
async def update_customer(customer_id: str, customer_data: CustomerCreate, current_user: User = Depends(get_current_user)):
    result = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not result:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await db.customers.update_one({"id": customer_id}, {"$set": customer_data.model_dump()})
    updated = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if isinstance(updated.get('created_at'), str):
        updated['created_at'] = datetime.fromisoformat(updated['created_at'])
    return Customer(**updated)

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: User = Depends(get_current_user)):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {"message": "Customer deleted successfully"}

# Customer Balance
@api_router.get("/customers/{customer_id}/balance")
async def get_customer_balance(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    sales = await db.sales.find({"customer_id": customer_id}, {"_id": 0}).to_list(10000)
    
    total_sales = 0
    total_cash = 0
    total_bank = 0
    total_credit_given = 0
    total_credit_received = 0
    
    for sale in sales:
        total_sales += sale.get("final_amount", sale.get("amount", 0) - sale.get("discount", 0))
        for p in sale.get("payment_details", []):
            if p.get("mode") == "cash":
                total_cash += p["amount"]
            elif p.get("mode") == "bank":
                total_bank += p["amount"]
        total_credit_given += sale.get("credit_amount", 0)
        total_credit_received += sale.get("credit_received", 0)
    
    return {
        "customer_id": customer_id,
        "customer_name": customer["name"],
        "total_sales": total_sales,
        "total_cash": total_cash,
        "total_bank": total_bank,
        "total_credit_given": total_credit_given,
        "total_credit_received": total_credit_received,
        "credit_balance": total_credit_given - total_credit_received,
        "sales_count": len(sales)
    }

# Customer Purchase Report
@api_router.get("/customers/{customer_id}/report")
async def get_customer_report(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    sales = await db.sales.find({"customer_id": customer_id}, {"_id": 0}).sort("date", -1).to_list(10000)
    invoices_list = await db.invoices.find({"customer_id": customer_id}, {"_id": 0}).to_list(10000)
    inv_by_sale = {inv.get("sale_id"): inv for inv in invoices_list if inv.get("sale_id")}
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    
    purchases = []
    for s in sales:
        payments_detail = [{"mode": p.get("mode",""), "amount": p.get("amount", 0)} for p in s.get("payment_details", [])]
        credit_given = s.get("credit_amount", 0)
        credit_received = s.get("credit_received", 0)
        
        # Link invoice items
        inv = inv_by_sale.get(s["id"])
        items = inv.get("items", []) if inv else []
        inv_number = inv.get("invoice_number", "") if inv else ""
        
        purchases.append({
            "date": s["date"], "type": "Sale",
            "branch": branches.get(s.get("branch_id"), "-"),
            "amount": s.get("final_amount", s["amount"]),
            "discount": s.get("discount", 0),
            "payments": payments_detail,
            "payment": ", ".join(f'{p["mode"]}' for p in payments_detail),
            "credit_given": credit_given,
            "credit_received": credit_received,
            "credit": credit_given - credit_received,
            "invoice_number": inv_number,
            "items": items
        })
    
    total = sum(p["amount"] for p in purchases)
    total_disc = sum(p["discount"] for p in purchases)
    total_credit = sum(p["credit"] for p in purchases if p["credit"] > 0)
    
    return {"customer": customer, "purchases": purchases, "total": total, "total_discount": total_disc, "credit_balance": total_credit, "count": len(purchases)}

@api_router.get("/customers/{customer_id}/report/pdf")
async def export_customer_report_pdf(customer_id: str, current_user: User = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    sales = await db.sales.find({"customer_id": customer_id}, {"_id": 0}).sort("date", -1).to_list(10000)
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    co_name = company.get("company_name", "Smart Standard Company")
    co_addr = ", ".join([p for p in [company.get("address_line1",""), company.get("city",""), company.get("country","")] if p])
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#F5841F'), alignment=1, spaceAfter=3)
    
    logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.jpg"
    if not logo_path.exists(): logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.png"
    if logo_path.exists():
        from reportlab.platypus import Image as RLImage
        try:
            logo = RLImage(str(logo_path), width=1.5*inch, height=0.7*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
        except: pass
    
    elements.append(Paragraph(co_name.upper(), title_s))
    if co_addr:
        elements.append(Paragraph(co_addr, ParagraphStyle('A', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#F5841F')))
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(f"<b>Customer Statement - {customer['name']}</b>", ParagraphStyle('H', parent=styles['Heading2'], fontSize=13, alignment=1)))
    elements.append(Paragraph(f"Phone: {customer.get('phone', '-')} | Date: {datetime.now().strftime('%d %b %Y')}", ParagraphStyle('S', parent=styles['Normal'], fontSize=9, textColor=colors.grey, alignment=1)))
    elements.append(Spacer(1, 0.15*inch))
    
    rows = [["Date", "Branch", "Amount", "Discount", "Payment", "Credit Due"]]
    total_amt = 0
    total_disc = 0
    total_credit = 0
    for s in sales:
        dt = datetime.fromisoformat(s["date"]).strftime("%d %b %Y") if isinstance(s["date"], str) else s["date"].strftime("%d %b %Y")
        amt = s.get("final_amount", s["amount"])
        disc = s.get("discount", 0)
        modes = ", ".join(p["mode"] for p in s.get("payment_details", []))
        credit = s.get("credit_amount", 0) - s.get("credit_received", 0)
        rows.append([dt, branches.get(s.get("branch_id"), "-"), f"SAR {amt:.2f}", f"SAR {disc:.2f}" if disc > 0 else "-", modes, f"SAR {credit:.2f}" if credit > 0 else "-"])
        total_amt += amt
        total_disc += disc
        if credit > 0: total_credit += credit
    
    rows.append(["", "", "", "", "", ""])
    rows.append(["TOTAL", "", f"SAR {total_amt:.2f}", f"SAR {total_disc:.2f}", "", f"SAR {total_credit:.2f}"])
    
    t = Table(rows, colWidths=[1*inch, 1*inch, 1.1*inch, 0.9*inch, 1*inch, 1*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFF3E0')),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
    ]))
    elements.append(t)
    
    doc.build(elements)
    buffer.seek(0)
    fname = f"statement_{customer['name'].replace(' ','_')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})

# All Customers Balance Summary
@api_router.get("/customers-balance")
async def get_all_customers_balance(current_user: User = Depends(get_current_user)):
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    
    result = []
    for customer in customers:
        cid = customer["id"]
        cust_sales = [s for s in sales if s.get("customer_id") == cid]
        total_sales = sum(s.get("final_amount", s.get("amount", 0)) for s in cust_sales)
        total_cash = sum(p["amount"] for s in cust_sales for p in s.get("payment_details", []) if p.get("mode") == "cash")
        total_bank = sum(p["amount"] for s in cust_sales for p in s.get("payment_details", []) if p.get("mode") == "bank")
        total_credit = sum(s.get("credit_amount", 0) for s in cust_sales)
        total_received = sum(s.get("credit_received", 0) for s in cust_sales)
        
        result.append({
            "id": cid,
            "name": customer["name"],
            "phone": customer.get("phone", ""),
            "branch_id": customer.get("branch_id"),
            "total_sales": total_sales,
            "cash": total_cash,
            "bank": total_bank,
            "credit_given": total_credit,
            "credit_received": total_received,
            "credit_balance": total_credit - total_received,
            "sales_count": len(cust_sales)
        })
    
    return result

# Sales Routes
@api_router.get("/sales", response_model=List[Sale])
async def get_sales(current_user: User = Depends(get_current_user)):
    query = {}
    # Filter by branch for non-admin users
    if current_user.branch_id and current_user.role != "admin":
        query["branch_id"] = current_user.branch_id
    
    sales = await db.sales.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    
    # Migrate old sales to new format
    for sale in sales:
        if isinstance(sale.get('date'), str):
            sale['date'] = datetime.fromisoformat(sale['date'])
        if isinstance(sale.get('created_at'), str):
            sale['created_at'] = datetime.fromisoformat(sale['created_at'])
        
        # Add discount and final_amount if missing
        if 'discount' not in sale:
            sale['discount'] = 0
        if 'final_amount' not in sale:
            sale['final_amount'] = sale.get('amount', 0) - sale.get('discount', 0)
        
        # Convert old format to new format
        if 'payment_details' not in sale or sale['payment_details'] is None:
            payment_mode = sale.get('payment_mode', 'cash')
            payment_status = sale.get('payment_status', 'received')
            amount = sale.get('amount', 0)
            
            if payment_mode == 'credit':
                if payment_status == 'pending':
                    sale['payment_details'] = []
                    sale['credit_amount'] = sale['final_amount']
                    sale['credit_received'] = 0
                else:
                    received_mode = sale.get('received_mode', 'cash')
                    sale['payment_details'] = [{"mode": received_mode, "amount": sale['final_amount']}]
                    sale['credit_amount'] = 0
                    sale['credit_received'] = 0
            else:
                sale['payment_details'] = [{"mode": payment_mode, "amount": sale['final_amount']}]
                sale['credit_amount'] = 0
                sale['credit_received'] = 0
            
            # Update the database with new format
            await db.sales.update_one(
                {"id": sale['id']},
                {"$set": {
                    "payment_details": sale['payment_details'],
                    "credit_amount": sale['credit_amount'],
                    "credit_received": sale['credit_received'],
                    "discount": sale['discount'],
                    "final_amount": sale['final_amount']
                }}
            )
    
    return sales

@api_router.post("/sales", response_model=Sale)
async def create_sale(sale_data: SaleCreate, current_user: User = Depends(get_current_user)):
    discount = sale_data.discount or 0
    final_amount = sale_data.amount - discount
    
    # Calculate credit from payment details
    total_cash_bank = sum(p["amount"] for p in sale_data.payment_details if p["mode"] in ["cash", "bank"])
    total_credit = sum(p["amount"] for p in sale_data.payment_details if p["mode"] == "credit")
    credit_amount = max(0, total_credit if total_credit > 0 else final_amount - total_cash_bank)
    
    # Clean empty strings
    sale_data_dict = sale_data.model_dump()
    sale_data_dict.pop('discount', None)
    for f in ['branch_id', 'customer_id']:
        if not sale_data_dict.get(f):
            sale_data_dict[f] = None
    
    sale = Sale(
        **sale_data_dict,
        discount=discount,
        final_amount=final_amount,
        credit_amount=credit_amount,
        credit_received=0,
        created_by=current_user.id
    )
    sale_dict = sale.model_dump()
    sale_dict["date"] = sale_dict["date"].isoformat()
    sale_dict["created_at"] = sale_dict["created_at"].isoformat()
    await db.sales.insert_one(sale_dict)
    return sale

@api_router.post("/sales/{sale_id}/receive-credit")
async def receive_credit_payment(sale_id: str, payment: SalePayment, current_user: User = Depends(get_current_user)):
    sale = await db.sales.find_one({"id": sale_id}, {"_id": 0})
    if not sale:
        raise HTTPException(status_code=404, detail="Sale not found")
    
    remaining_credit = sale["credit_amount"] - sale["credit_received"]
    discount = payment.discount or 0
    total_settle = payment.amount + discount
    
    if total_settle > remaining_credit + 0.01:
        raise HTTPException(status_code=400, detail=f"Payment + discount exceeds remaining credit of ${remaining_credit:.2f}")
    
    new_credit_received = sale["credit_received"] + total_settle
    new_payment_details = list(sale.get("payment_details", []))
    if payment.amount > 0:
        new_payment_details.append({"mode": payment.payment_mode, "amount": payment.amount})
    if discount > 0:
        new_payment_details.append({"mode": "discount", "amount": discount})
    
    # Also update credit_amount if discount reduces it
    new_credit_amount = sale["credit_amount"]
    
    await db.sales.update_one(
        {"id": sale_id},
        {"$set": {
            "credit_received": new_credit_received,
            "payment_details": new_payment_details
        }}
    )
    
    return {
        "message": "Credit payment received",
        "received": payment.amount,
        "discount": discount,
        "remaining_credit": remaining_credit - total_settle
    }

@api_router.delete("/sales/{sale_id}")
async def delete_sale(sale_id: str, current_user: User = Depends(get_current_user)):
    result = await db.sales.delete_one({"id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Sale not found")
    return {"message": "Sale deleted successfully"}

# Supplier Payment Routes
@api_router.get("/supplier-payments", response_model=List[SupplierPayment])
async def get_supplier_payments(current_user: User = Depends(get_current_user)):
    payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).sort("date", -1).to_list(1000)
    for payment in payments:
        if isinstance(payment.get('date'), str):
            payment['date'] = datetime.fromisoformat(payment['date'])
        if isinstance(payment.get('created_at'), str):
            payment['created_at'] = datetime.fromisoformat(payment['created_at'])
    return payments

@api_router.post("/supplier-payments", response_model=SupplierPayment)
async def create_supplier_payment(payment_data: SupplierPaymentCreate, current_user: User = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": payment_data.supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    # Validate credit limit BEFORE inserting payment
    if payment_data.payment_mode == "credit":
        credit_limit = supplier.get("credit_limit", 0)
        current_credit = supplier.get("current_credit", 0)
        new_credit = current_credit + payment_data.amount
        # Only enforce limit if credit_limit > 0 (0 means no limit set)
        if credit_limit > 0 and new_credit > credit_limit:
            available = credit_limit - current_credit
            raise HTTPException(status_code=400, detail=f"Payment exceeds credit limit. Available: SAR {available:.2f}")
    
    payment = SupplierPayment(
        **payment_data.model_dump(),
        supplier_name=supplier["name"],
        created_by=current_user.id
    )
    payment_dict = payment.model_dump()
    payment_dict["date"] = payment_dict["date"].isoformat()
    payment_dict["created_at"] = payment_dict["created_at"].isoformat()
    await db.supplier_payments.insert_one(payment_dict)
    
    # Update supplier's current credit after successful insert
    if payment_data.payment_mode == "credit":
        new_credit = supplier.get("current_credit", 0) + payment_data.amount
        await db.suppliers.update_one({"id": payment_data.supplier_id}, {"$set": {"current_credit": new_credit}})
    
    return payment

@api_router.delete("/supplier-payments/{payment_id}")
async def delete_supplier_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    result = await db.supplier_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Payment deleted successfully"}

# Expense Routes
@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(current_user: User = Depends(get_current_user)):
    expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for expense in expenses:
        if isinstance(expense.get('date'), str):
            expense['date'] = datetime.fromisoformat(expense['date'])
        if isinstance(expense.get('created_at'), str):
            expense['created_at'] = datetime.fromisoformat(expense['created_at'])
    return expenses

@api_router.post("/expenses", response_model=Expense)
async def create_expense(expense_data: ExpenseCreate, current_user: User = Depends(get_current_user)):
    # Clean empty strings
    data = expense_data.model_dump()
    for f in ['branch_id', 'supplier_id', 'sub_category']:
        if data.get(f) == '':
            data[f] = None
    
    # Validate credit limit BEFORE inserting expense
    if data.get('supplier_id') and expense_data.payment_mode == "credit":
        supplier = await db.suppliers.find_one({"id": data['supplier_id']}, {"_id": 0})
        if supplier:
            credit_limit = supplier.get("credit_limit", 0)
            current_credit = supplier.get("current_credit", 0)
            new_credit = current_credit + expense_data.amount
            if credit_limit > 0 and new_credit > credit_limit:
                available = credit_limit - current_credit
                raise HTTPException(status_code=400, detail=f"Expense exceeds supplier credit limit. Available: SAR {available:.2f}")
    
    expense = Expense(**data, created_by=current_user.id)
    expense_dict = expense.model_dump()
    expense_dict["date"] = expense_dict["date"].isoformat()
    expense_dict["created_at"] = expense_dict["created_at"].isoformat()
    await db.expenses.insert_one(expense_dict)
    
    # Update supplier credit after successful insert
    if data.get('supplier_id') and expense_data.payment_mode == "credit":
        supplier = await db.suppliers.find_one({"id": data['supplier_id']}, {"_id": 0})
        if supplier:
            new_credit = supplier.get("current_credit", 0) + expense_data.amount
            await db.suppliers.update_one({"id": data['supplier_id']}, {"$set": {"current_credit": new_credit}})
    
    return expense

@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: User = Depends(get_current_user)):
    result = await db.expenses.delete_one({"id": expense_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Expense not found")
    return {"message": "Expense deleted successfully"}

# Branch Summary Stats
@api_router.get("/branches/{branch_id}/summary")
async def get_branch_summary(branch_id: str, current_user: User = Depends(get_current_user)):
    branch = await db.branches.find_one({"id": branch_id}, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="Branch not found")
    
    sales = await db.sales.find({"branch_id": branch_id}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({"branch_id": branch_id}, {"_id": 0}).to_list(10000)
    sp = await db.supplier_payments.find({"branch_id": branch_id, "supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    
    total_sales = sum(s.get("final_amount", s.get("amount", 0)) for s in sales)
    cash_sales = sum(p["amount"] for s in sales for p in s.get("payment_details", []) if p.get("mode") == "cash")
    bank_sales = sum(p["amount"] for s in sales for p in s.get("payment_details", []) if p.get("mode") == "bank")
    credit_sales = sum(s.get("credit_amount", 0) - s.get("credit_received", 0) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    exp_cash = sum(e["amount"] for e in expenses if e.get("payment_mode") == "cash")
    exp_bank = sum(e["amount"] for e in expenses if e.get("payment_mode") == "bank")
    total_sp = sum(p["amount"] for p in sp)
    sp_cash = sum(p["amount"] for p in sp if p.get("payment_mode") == "cash")
    sp_bank = sum(p["amount"] for p in sp if p.get("payment_mode") == "bank")
    
    return {
        "branch_id": branch_id, "branch_name": branch["name"],
        "total_sales": total_sales, "sales_cash": cash_sales, "sales_bank": bank_sales, "sales_credit": credit_sales, "sales_count": len(sales),
        "total_expenses": total_expenses, "expenses_cash": exp_cash, "expenses_bank": exp_bank, "expenses_count": len(expenses),
        "total_supplier_payments": total_sp, "sp_cash": sp_cash, "sp_bank": sp_bank, "sp_count": len(sp),
        "net_profit": total_sales - total_expenses - total_sp,
        "cash_in_hand": cash_sales - exp_cash - sp_cash,
        "bank_in_hand": bank_sales - exp_bank - sp_bank
    }

# Dashboard Stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(branch_ids: Optional[str] = None, start_date: Optional[str] = None, end_date: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    exp_query = {}
    sp_query = {"supplier_id": {"$exists": True, "$ne": None}}
    
    if branch_ids:
        bid_list = [b.strip() for b in branch_ids.split(",") if b.strip()]
        if bid_list:
            query["branch_id"] = {"$in": bid_list}
            exp_query["branch_id"] = {"$in": bid_list}
            sp_query["branch_id"] = {"$in": bid_list}
    elif current_user.branch_id and current_user.role != "admin":
        query["branch_id"] = current_user.branch_id
    
    if start_date and end_date:
        date_filter = {"$gte": start_date, "$lte": end_date}
        query["date"] = date_filter
        exp_query["date"] = date_filter
        sp_query["date"] = date_filter
    
    sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find(exp_query, {"_id": 0}).to_list(10000)
    supplier_payments = await db.supplier_payments.find(sp_query, {"_id": 0}).to_list(10000)
    
    # Calculate total sales (amount minus remaining credit)
    total_sales = sum(sale["amount"] - (sale.get("credit_amount", 0) - sale.get("credit_received", 0)) for sale in sales)
    total_expenses = sum(expense["amount"] for expense in expenses)
    total_supplier_payments = sum(payment["amount"] for payment in supplier_payments if payment.get("payment_mode") != "credit")
    
    # Calculate pending credits
    pending_credits = sum(sale.get("credit_amount", 0) - sale.get("credit_received", 0) for sale in sales)
    
    # Calculate payment mode breakdown
    cash_sales = 0
    bank_sales = 0
    credit_sales = pending_credits
    
    for sale in sales:
        for payment in sale.get("payment_details", []):
            if payment["mode"] == "cash":
                cash_sales += payment["amount"]
            elif payment["mode"] == "bank":
                bank_sales += payment["amount"]
    
    net_profit = total_sales - total_expenses - total_supplier_payments
    
    # Cash & Bank in hand (income - outgoing per mode)
    exp_cash = sum(e["amount"] for e in expenses if e.get("payment_mode") == "cash")
    exp_bank = sum(e["amount"] for e in expenses if e.get("payment_mode") == "bank")
    sp_cash = sum(p["amount"] for p in supplier_payments if p.get("payment_mode") == "cash")
    sp_bank = sum(p["amount"] for p in supplier_payments if p.get("payment_mode") == "bank")
    cash_in_hand = cash_sales - exp_cash - sp_cash
    bank_in_hand = bank_sales - exp_bank - sp_bank
    
    # Expense breakdown by category
    expense_by_category = {}
    for e in expenses:
        cat = e.get("category", "other")
        expense_by_category[cat] = expense_by_category.get(cat, 0) + e["amount"]
    
    # Supplier dues (all suppliers credit)
    all_suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    supplier_dues = sum(s.get("current_credit", 0) for s in all_suppliers)
    
    # Recurring expense alerts
    recurring = await db.recurring_expenses.find({"active": True}, {"_id": 0}).to_list(100)
    now = datetime.now(timezone.utc)
    upcoming_expenses = []
    for r in recurring:
        due = r.get("next_due_date")
        if isinstance(due, str):
            due = datetime.fromisoformat(due)
        if due and due.tzinfo is None:
            due = due.replace(tzinfo=timezone.utc)
        if due:
            days_left = (due - now).days
            if days_left <= r.get("alert_days", 7):
                upcoming_expenses.append({"name": r["name"], "category": r.get("category", ""), "amount": r["amount"], "due_date": due.isoformat(), "days_left": days_left})
    
    # Branch-to-branch dues from cash transfers
    transfers = await db.cash_transfers.find({}, {"_id": 0}).to_list(10000)
    branch_dues = {}
    for t in transfers:
        from_b = t.get("from_branch_name", "Office")
        to_b = t.get("to_branch_name", "Office")
        if from_b != to_b:
            key = f"{to_b} → {from_b}"
            branch_dues[key] = branch_dues.get(key, 0) + t["amount"]
    
    # Previous month comparison
    prev_month_start = (now.replace(day=1) - timedelta(days=1)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_end = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(seconds=1)
    pms = prev_month_start.isoformat()
    pme = prev_month_end.isoformat()
    
    prev_sales_q = {"date": {"$gte": pms, "$lte": pme}}
    prev_exp_q = {"date": {"$gte": pms, "$lte": pme}}
    if branch_ids:
        bid_list = [b.strip() for b in branch_ids.split(",") if b.strip()]
        if bid_list:
            prev_sales_q["branch_id"] = {"$in": bid_list}
            prev_exp_q["branch_id"] = {"$in": bid_list}
    
    prev_sales = await db.sales.find(prev_sales_q, {"_id": 0}).to_list(10000)
    prev_expenses = await db.expenses.find(prev_exp_q, {"_id": 0}).to_list(10000)
    
    prev_total_sales = sum(s.get("final_amount", s["amount"]) for s in prev_sales)
    prev_total_expenses = sum(e["amount"] for e in prev_expenses)
    prev_net = prev_total_sales - prev_total_expenses
    
    # Due fines
    all_fines = await db.fines.find({"payment_status": {"$ne": "paid"}}, {"_id": 0}).to_list(1000)
    due_fines = sum(f["amount"] - f.get("paid_amount", 0) for f in all_fines)
    due_fines_list = [{"department": f.get("department",""), "amount": f["amount"] - f.get("paid_amount",0), "type": f.get("fine_type","")} for f in all_fines[:5]]
    
    return {
        "total_sales": total_sales,
        "total_expenses": total_expenses,
        "total_supplier_payments": total_supplier_payments,
        "net_profit": net_profit,
        "pending_credits": pending_credits,
        "cash_sales": cash_sales,
        "bank_sales": bank_sales,
        "credit_sales": credit_sales,
        "cash_in_hand": cash_in_hand,
        "bank_in_hand": bank_in_hand,
        "expenses_cash": exp_cash,
        "expenses_bank": exp_bank,
        "sp_cash": sp_cash,
        "sp_bank": sp_bank,
        "expense_by_category": expense_by_category,
        "supplier_dues": supplier_dues,
        "upcoming_expenses": upcoming_expenses,
        "branch_dues": branch_dues,
        "prev_sales": prev_total_sales,
        "prev_expenses": prev_total_expenses,
        "prev_net": prev_net,
        "expenses_pct_of_sales": round(total_expenses / total_sales * 100, 1) if total_sales > 0 else 0,
        "sp_pct_of_sales": round(total_supplier_payments / total_sales * 100, 1) if total_sales > 0 else 0,
        "profit_pct_of_sales": round(net_profit / total_sales * 100, 1) if total_sales > 0 else 0,
        "due_fines": due_fines,
        "due_fines_list": due_fines_list,
        "vat_on_sales": round(total_sales * 0.15, 2),
        "vat_on_purchases": round(total_supplier_payments * 0.15, 2),
        "vat_payable": round((total_sales - total_supplier_payments) * 0.15, 2)
    }

# Credit Sales Report
@api_router.get("/reports/credit-sales")
async def get_credit_sales_report(current_user: User = Depends(get_current_user)):
    query = {}
    if current_user.branch_id and current_user.role != "admin":
        query["branch_id"] = current_user.branch_id
    
    sales = await db.sales.find(query, {"_id": 0}).to_list(10000)
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
    
    credit_sales = []
    total_credit_given = 0
    total_credit_received = 0
    total_credit_remaining = 0
    
    for sale in sales:
        credit_amount = sale.get('credit_amount', 0)
        credit_received = sale.get('credit_received', 0)
        remaining = credit_amount - credit_received
        
        if credit_amount > 0:
            branch_name = next((b["name"] for b in branches if b["id"] == sale.get("branch_id")), "-")
            customer_name = next((c["name"] for c in customers if c["id"] == sale.get("customer_id")), "-")
            
            credit_sales.append({
                "id": sale["id"],
                "date": sale["date"],
                "sale_type": sale["sale_type"],
                "reference": branch_name if sale["sale_type"] == "branch" else customer_name,
                "branch": branch_name,
                "total_amount": sale["amount"],
                "discount": sale.get("discount", 0),
                "final_amount": sale.get("final_amount", sale["amount"] - sale.get("discount", 0)),
                "credit_given": credit_amount,
                "credit_received": credit_received,
                "remaining": remaining,
                "status": "paid" if remaining == 0 else "partial" if credit_received > 0 else "pending"
            })
            
            total_credit_given += credit_amount
            total_credit_received += credit_received
            total_credit_remaining += remaining
    
    return {
        "credit_sales": credit_sales,
        "summary": {
            "total_credit_given": total_credit_given,
            "total_credit_received": total_credit_received,
            "total_credit_remaining": total_credit_remaining
        }
    }

# Supplier Reports
@api_router.get("/reports/suppliers")
async def get_supplier_report(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    supplier_payments = await db.supplier_payments.find({}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    
    supplier_report = []
    
    for supplier in suppliers:
        # Calculate total payments to this supplier
        payments = [p for p in supplier_payments if p.get("supplier_id") == supplier["id"]]
        total_paid = sum(p["amount"] for p in payments if p.get("payment_mode") != "credit")
        
        # Calculate supplier expenses
        supplier_expenses = [e for e in expenses if e.get("supplier_id") == supplier["id"]]
        total_expenses = sum(e["amount"] for e in supplier_expenses)
        
        supplier_report.append({
            "id": supplier["id"],
            "name": supplier["name"],
            "category": supplier.get("category", "-"),
            "total_expenses": total_expenses,
            "total_paid": total_paid,
            "current_credit": supplier.get("current_credit", 0),
            "credit_limit": supplier.get("credit_limit", 0),
            "transaction_count": len(payments) + len(supplier_expenses)
        })
    
    return supplier_report

# Category-wise Supplier Report
@api_router.get("/reports/supplier-categories")
async def get_supplier_category_report(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    supplier_payments = await db.supplier_payments.find({}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    
    category_report = {}
    
    for supplier in suppliers:
        category = supplier.get("category", "Uncategorized")
        
        if category not in category_report:
            category_report[category] = {
                "category": category,
                "supplier_count": 0,
                "total_expenses": 0,
                "total_paid": 0,
                "total_credit": 0
            }
        
        category_report[category]["supplier_count"] += 1
        
        # Calculate expenses for this supplier
        supplier_expenses = [e for e in expenses if e.get("supplier_id") == supplier["id"]]
        category_report[category]["total_expenses"] += sum(e["amount"] for e in supplier_expenses)
        
        # Calculate payments for this supplier
        payments = [p for p in supplier_payments if p.get("supplier_id") == supplier["id"]]
        category_report[category]["total_paid"] += sum(p["amount"] for p in payments if p.get("payment_mode") != "credit")
        
        # Add current credit
        category_report[category]["total_credit"] += supplier.get("current_credit", 0)
    
    return list(category_report.values())

# Branch-wise Cash/Bank Report
@api_router.get("/reports/branch-cashbank")
async def get_branch_cashbank_report(current_user: User = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
    expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
    supplier_payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    
    branch_data = []
    for branch in branches:
        bid = branch["id"]
        # Sales cash/bank by branch
        branch_sales = [s for s in sales if s.get("branch_id") == bid]
        sales_cash = sum(p["amount"] for s in branch_sales for p in s.get("payment_details", []) if p.get("mode") == "cash")
        sales_bank = sum(p["amount"] for s in branch_sales for p in s.get("payment_details", []) if p.get("mode") == "bank")
        sales_credit = sum(s.get("credit_amount", 0) - s.get("credit_received", 0) for s in branch_sales)
        
        # Expenses by branch
        branch_expenses = [e for e in expenses if e.get("branch_id") == bid]
        exp_cash = sum(e["amount"] for e in branch_expenses if e.get("payment_mode") == "cash")
        exp_bank = sum(e["amount"] for e in branch_expenses if e.get("payment_mode") == "bank")
        
        # Supplier payments by branch
        branch_sp = [p for p in supplier_payments if p.get("branch_id") == bid]
        sp_cash = sum(p["amount"] for p in branch_sp if p.get("payment_mode") == "cash")
        sp_bank = sum(p["amount"] for p in branch_sp if p.get("payment_mode") == "bank")
        
        branch_data.append({
            "branch_id": bid,
            "branch_name": branch["name"],
            "sales_cash": sales_cash,
            "sales_bank": sales_bank,
            "sales_credit": sales_credit,
            "sales_total": sales_cash + sales_bank + sales_credit,
            "expenses_cash": exp_cash,
            "expenses_bank": exp_bank,
            "expenses_total": exp_cash + exp_bank,
            "supplier_cash": sp_cash,
            "supplier_bank": sp_bank,
            "supplier_total": sp_cash + sp_bank,
        })
    
    return branch_data

# Supplier Balance Report with date filtering
@api_router.get("/reports/supplier-balance")
async def get_supplier_balance_report(
    period: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    
    date_query = {}
    if period == "today":
        today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
        date_query = {"date": {"$gte": today.isoformat()}}
    elif period == "month":
        month_start = datetime.now(timezone.utc).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        date_query = {"date": {"$gte": month_start.isoformat()}}
    elif period == "year":
        year_start = datetime.now(timezone.utc).replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        date_query = {"date": {"$gte": year_start.isoformat()}}
    elif start_date and end_date:
        date_query = {"date": {"$gte": start_date, "$lte": end_date}}
    
    sp_query = {"supplier_id": {"$exists": True, "$ne": None}}
    if date_query:
        sp_query.update(date_query)
    
    supplier_payments = await db.supplier_payments.find(sp_query, {"_id": 0}).to_list(10000)
    
    exp_query = {}
    if date_query:
        exp_query.update(date_query)
    expenses = await db.expenses.find(exp_query, {"_id": 0}).to_list(10000)
    
    result = []
    for supplier in suppliers:
        sid = supplier["id"]
        payments = [p for p in supplier_payments if p.get("supplier_id") == sid]
        cash_paid = sum(p["amount"] for p in payments if p.get("payment_mode") == "cash")
        bank_paid = sum(p["amount"] for p in payments if p.get("payment_mode") == "bank")
        credit_added = sum(p["amount"] for p in payments if p.get("payment_mode") == "credit")
        
        sup_expenses = [e for e in expenses if e.get("supplier_id") == sid]
        total_expenses = sum(e["amount"] for e in sup_expenses)
        
        result.append({
            "id": sid,
            "name": supplier["name"],
            "category": supplier.get("category", "-"),
            "branch_id": supplier.get("branch_id"),
            "cash_paid": cash_paid,
            "bank_paid": bank_paid,
            "credit_added": credit_added,
            "total_paid": cash_paid + bank_paid,
            "total_expenses": total_expenses,
            "current_credit": supplier.get("current_credit", 0),
            "credit_limit": supplier.get("credit_limit", 0),
            "transaction_count": len(payments) + len(sup_expenses)
        })
    
    return result

# Generic Data Export
@api_router.post("/export/data")
async def export_data(request: dict, current_user: User = Depends(get_current_user)):
    data_type = request.get("type", "sales")
    fmt = request.get("format", "excel")
    
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}
    
    buffer = BytesIO()
    
    if data_type == "sales":
        sales = await db.sales.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
        customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
        cust_map = {c["id"]: c["name"] for c in customers}
        rows = []
        for s in sales:
            modes = ", ".join(f'{p["mode"]}:${p["amount"]:.2f}' for p in s.get("payment_details", []))
            rows.append([
                datetime.fromisoformat(s["date"]).strftime("%Y-%m-%d") if isinstance(s["date"], str) else s["date"].strftime("%Y-%m-%d"),
                s["sale_type"].capitalize(),
                branch_map.get(s.get("branch_id"), "-"),
                cust_map.get(s.get("customer_id"), "-"),
                s["amount"],
                s.get("discount", 0),
                s.get("final_amount", s["amount"] - s.get("discount", 0)),
                modes,
                s.get("credit_amount", 0) - s.get("credit_received", 0)
            ])
        headers = ["Date", "Type", "Branch", "Customer", "Amount", "Discount", "Final", "Payments", "Credit Remaining"]
        title = "Sales Report"
    elif data_type == "expenses":
        expenses = await db.expenses.find({}, {"_id": 0}).sort("date", -1).to_list(10000)
        suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
        sup_map = {s["id"]: s["name"] for s in suppliers}
        rows = []
        for e in expenses:
            rows.append([
                datetime.fromisoformat(e["date"]).strftime("%Y-%m-%d") if isinstance(e["date"], str) else e["date"].strftime("%Y-%m-%d"),
                e["category"].capitalize(),
                e["description"],
                sup_map.get(e.get("supplier_id"), "-"),
                branch_map.get(e.get("branch_id"), "-"),
                e["amount"],
                e["payment_mode"].capitalize()
            ])
        headers = ["Date", "Category", "Description", "Supplier", "Branch", "Amount", "Payment Mode"]
        title = "Expenses Report"
    elif data_type == "supplier-payments":
        payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).sort("date", -1).to_list(10000)
        rows = []
        for p in payments:
            rows.append([
                datetime.fromisoformat(p["date"]).strftime("%Y-%m-%d") if isinstance(p["date"], str) else p["date"].strftime("%Y-%m-%d"),
                p["supplier_name"],
                branch_map.get(p.get("branch_id"), "-"),
                p["amount"],
                p["payment_mode"].capitalize(),
                p.get("notes", "")
            ])
        headers = ["Date", "Supplier", "Branch", "Amount", "Payment Mode", "Notes"]
        title = "Supplier Payments Report"
    elif data_type == "customers":
        customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
        rows = [[c["name"], branch_map.get(c.get("branch_id"), "All Branches"), c.get("phone", "-"), c.get("email", "-")] for c in customers]
        headers = ["Name", "Branch", "Phone", "Email"]
        title = "Customers Report"
    elif data_type == "suppliers":
        suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
        rows = [[s["name"], s.get("category", "-"), branch_map.get(s.get("branch_id"), "All"), s.get("phone", "-"), s.get("current_credit", 0), s.get("credit_limit", 0)] for s in suppliers]
        headers = ["Name", "Category", "Branch", "Phone", "Current Credit", "Credit Limit"]
        title = "Suppliers Report"
    elif data_type == "employees":
        employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
        rows = []
        for emp in employees:
            rows.append([
                emp["name"],
                emp.get("position", "-"),
                emp.get("document_id", "-"),
                branch_map.get(emp.get("branch_id"), "-"),
                emp.get("salary", 0),
                emp.get("pay_frequency", "monthly"),
                datetime.fromisoformat(emp["document_expiry"]).strftime("%Y-%m-%d") if emp.get("document_expiry") else "-"
            ])
        headers = ["Name", "Position", "Document ID", "Branch", "Salary", "Pay Frequency", "Doc Expiry"]
        title = "Employees Report"
    else:
        raise HTTPException(status_code=400, detail="Invalid data type")
    
    if fmt == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(headers)
        for col in ws[1]:
            col.font = Font(bold=True)
            col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
            col.font = Font(bold=True, color="FFFFFF")
        for row in rows:
            ws.append(row)
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)
        wb.save(buffer)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={data_type}_report.xlsx"})
    else:
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('T', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#F5841F'), alignment=1)
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        table_data = [headers] + [[str(c) for c in row] for row in rows[:50]]
        col_count = len(headers)
        col_width = 7.5 * inch / col_count
        t = Table(table_data, colWidths=[col_width] * col_count)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F3FF')])
        ]))
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return StreamingResponse(buffer, media_type="application/pdf",
                                 headers={"Content-Disposition": f"attachment; filename={data_type}_report.pdf"})

# Employee Routes
@api_router.get("/employees")
async def get_employees(current_user: User = Depends(get_current_user)):
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    for emp in employees:
        for f in ['created_at', 'join_date', 'document_expiry']:
            if isinstance(emp.get(f), str):
                emp[f] = datetime.fromisoformat(emp[f])
    return employees

@api_router.post("/employees")
async def create_employee(data: EmployeeCreate, current_user: User = Depends(get_current_user)):
    emp = Employee(**data.model_dump())
    emp_dict = emp.model_dump()
    
    # Auto-create user account for employee if email provided
    if data.email:
        existing_user = await db.users.find_one({"email": data.email}, {"_id": 0})
        if not existing_user:
            user = User(email=data.email, name=data.name, role="employee", permissions=["self_service"])
            user_dict = user.model_dump()
            user_dict["password"] = hash_password("emp@123")  # Default password
            user_dict["created_at"] = user_dict["created_at"].isoformat()
            await db.users.insert_one(user_dict)
            emp_dict["user_id"] = user.id
        else:
            emp_dict["user_id"] = existing_user["id"]
    
    for f in ['created_at', 'join_date', 'document_expiry']:
        if emp_dict.get(f):
            emp_dict[f] = emp_dict[f].isoformat()
    await db.employees.insert_one(emp_dict)
    return {k: v for k, v in emp_dict.items() if k != '_id'}

# Link existing employee to user account
@api_router.post("/employees/{emp_id}/link-user")
async def link_employee_user(emp_id: str, body: dict, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    email = body.get("email")
    if not email:
        raise HTTPException(status_code=400, detail="Email required")
    
    existing = await db.users.find_one({"email": email}, {"_id": 0})
    if existing:
        await db.employees.update_one({"id": emp_id}, {"$set": {"user_id": existing["id"]}})
        return {"message": f"Linked to existing user {email}", "user_id": existing["id"]}
    
    user = User(email=email, name=emp["name"], role="employee", permissions=["self_service"])
    user_dict = user.model_dump()
    user_dict["password"] = hash_password("emp@123")
    user_dict["created_at"] = user_dict["created_at"].isoformat()
    await db.users.insert_one(user_dict)
    await db.employees.update_one({"id": emp_id}, {"$set": {"user_id": user.id}})
    return {"message": f"Created account for {email} (password: emp@123)", "user_id": user.id}

@api_router.put("/employees/{emp_id}")
async def update_employee(emp_id: str, data: EmployeeCreate, current_user: User = Depends(get_current_user)):
    result = await db.employees.find_one({"id": emp_id})
    if not result:
        raise HTTPException(status_code=404, detail="Employee not found")
    update = data.model_dump()
    for f in ['join_date', 'document_expiry']:
        if update.get(f):
            update[f] = update[f].isoformat()
    await db.employees.update_one({"id": emp_id}, {"$set": update})
    updated = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    return updated

@api_router.delete("/employees/{emp_id}")
async def delete_employee(emp_id: str, current_user: User = Depends(get_current_user)):
    result = await db.employees.delete_one({"id": emp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Employee not found")
    return {"message": "Employee deleted"}

# Salary Payment Routes
@api_router.get("/salary-payments")
async def get_salary_payments(current_user: User = Depends(get_current_user)):
    payments = await db.salary_payments.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for p in payments:
        for f in ['date', 'created_at']:
            if isinstance(p.get(f), str):
                p[f] = datetime.fromisoformat(p[f])
    return payments

@api_router.post("/salary-payments")
async def create_salary_payment(data: SalaryPaymentCreate, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": data.employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Prevent duplicate salary payment for same month
    if data.payment_type == "salary":
        existing = await db.salary_payments.find_one({"employee_id": data.employee_id, "period": data.period, "payment_type": "salary"}, {"_id": 0})
        if existing:
            raise HTTPException(status_code=400, detail=f"Salary already paid for {data.period}. Use overtime/bonus for extra payments.")
    
    payment = SalaryPayment(**data.model_dump(), employee_name=emp["name"], created_by=current_user.id)
    p_dict = payment.model_dump()
    p_dict["date"] = p_dict["date"].isoformat()
    p_dict["created_at"] = p_dict["created_at"].isoformat()
    await db.salary_payments.insert_one(p_dict)
    
    # Update loan balance
    loan_balance = emp.get("loan_balance", 0)
    if data.payment_type == "advance":
        await db.employees.update_one({"id": data.employee_id}, {"$set": {"loan_balance": loan_balance + data.amount}})
    elif data.payment_type == "loan_repayment":
        await db.employees.update_one({"id": data.employee_id}, {"$set": {"loan_balance": max(0, loan_balance - data.amount)}})
    
    # Reduce old salary balance if paying old_balance type
    if data.payment_type == "old_balance":
        old_bal = emp.get("old_salary_balance", 0)
        await db.employees.update_one({"id": data.employee_id}, {"$set": {"old_salary_balance": max(0, old_bal - data.amount)}})
    
    # Create expense for ALL salary payments (except loan_repayment)
    if data.payment_type != "loan_repayment":
        cat_map = {"salary": "salary", "advance": "salary", "overtime": "salary", "bonus": "salary", "old_balance": "salary", "tickets": "tickets", "id_card": "id_card"}
        type_label = data.payment_type.replace("_", " ").title()
        expense = Expense(
            category=cat_map.get(data.payment_type, "salary"),
            description=f"{type_label} - {emp['name']} - {data.period}",
            amount=data.amount,
            payment_mode=data.payment_mode,
            branch_id=data.branch_id,
            date=data.date,
            notes=data.notes or f"Employee: {emp['name']}",
            created_by=current_user.id
        )
        e_dict = expense.model_dump()
        e_dict["date"] = e_dict["date"].isoformat()
        e_dict["created_at"] = e_dict["created_at"].isoformat()
        await db.expenses.insert_one(e_dict)
    
    # Send notification to employee if they have a linked user account
    if emp.get("user_id"):
        type_label = data.payment_type.replace("_", " ").title()
        notif = Notification(
            user_id=emp["user_id"],
            title=f"{type_label} Payment Received",
            message=f"SAR {data.amount:.2f} {type_label} for {data.period} via {data.payment_mode}. Please acknowledge receipt.",
            type="salary_paid",
            related_id=payment.id
        )
        n_dict = notif.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    
    return {k: v for k, v in p_dict.items() if k != '_id'}

# Employee payment summary
@api_router.get("/employees/{emp_id}/summary")
async def get_employee_summary(emp_id: str, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    payments = await db.salary_payments.find({"employee_id": emp_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    leaves = await db.leaves.find({"employee_id": emp_id}, {"_id": 0}).to_list(1000)
    deductions = await db.salary_deductions.find({"employee_id": emp_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    fines = await db.fines.find({"employee_id": emp_id}, {"_id": 0}).sort("fine_date", -1).to_list(1000)
    
    # Group payments by period
    monthly = {}
    total_advance = 0
    total_repaid = 0
    for p in payments:
        period = p.get("period", "Unknown")
        if period not in monthly:
            monthly[period] = {"salary_paid": 0, "advance": 0, "overtime": 0, "tickets": 0, "id_card": 0, "loan_repayment": 0, "total": 0, "payments": []}
        pt = p.get("payment_type", "salary")
        amount = p.get("amount", 0)
        if pt in monthly[period]:
            monthly[period][pt] += amount
        else:
            monthly[period]["salary_paid"] += amount
        monthly[period]["total"] += amount
        if pt == "advance":
            total_advance += amount
        if pt == "loan_repayment":
            total_repaid += amount
        monthly[period]["payments"].append({
            "id": p["id"], "payment_type": pt, "amount": amount,
            "payment_mode": p.get("payment_mode", "cash"), "date": p["date"], "notes": p.get("notes", "")
        })
    
    salary = emp.get("salary", 0)
    summary = []
    for period, data in monthly.items():
        balance = salary - data["salary_paid"]
        summary.append({
            "period": period, "monthly_salary": salary,
            "salary_paid": data["salary_paid"], "advance": data["advance"],
            "overtime": data["overtime"], "tickets": data["tickets"],
            "id_card": data["id_card"], "loan_repayment": data["loan_repayment"],
            "total_paid": data["total"], "balance": balance,
            "payments": data["payments"]
        })
    
    # Leave summary
    annual_used = sum(l.get("days", 0) for l in leaves if l.get("leave_type") == "annual")
    sick_used = sum(l.get("days", 0) for l in leaves if l.get("leave_type") == "sick")
    unpaid_used = sum(l.get("days", 0) for l in leaves if l.get("leave_type") == "unpaid")
    
    return {
        "employee": {
            "id": emp["id"], "name": emp["name"], "salary": salary,
            "position": emp.get("position", ""),
            "loan_balance": emp.get("loan_balance", 0),
            "annual_leave_entitled": emp.get("annual_leave_entitled", 30),
            "sick_leave_entitled": emp.get("sick_leave_entitled", 15),
        },
        "monthly_summary": summary,
        "total_all_time": sum(p.get("amount", 0) for p in payments),
        "loan": {"total_advance": total_advance, "total_repaid": total_repaid, "balance": emp.get("loan_balance", 0)},
        "old_salary_balance": emp.get("old_salary_balance", 0),
        "leave": {
            "annual_used": annual_used, "annual_remaining": emp.get("annual_leave_entitled", 30) - annual_used,
            "sick_used": sick_used, "sick_remaining": emp.get("sick_leave_entitled", 15) - sick_used,
            "unpaid_used": unpaid_used
        },
        "deductions": [{"id": d["id"], "type": d.get("deduction_type",""), "amount": d["amount"], "period": d.get("period",""), "reason": d.get("reason",""), "date": d.get("date","")} for d in deductions],
        "total_deductions": sum(d["amount"] for d in deductions),
        "fines": [{"id": f["id"], "type": f.get("fine_type",""), "department": f.get("department",""), "amount": f["amount"], "paid": f.get("paid_amount",0), "status": f.get("payment_status","unpaid"), "description": f.get("description","")} for f in fines],
        "total_fines": sum(f["amount"] for f in fines),
        "unpaid_fines": sum(f["amount"] - f.get("paid_amount",0) for f in fines if f.get("payment_status") != "paid")
    }

@api_router.delete("/salary-payments/{payment_id}")
async def delete_salary_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    result = await db.salary_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Payment not found")
    return {"message": "Salary payment deleted"}

# Leave Routes
@api_router.get("/leaves")
async def get_leaves(employee_id: Optional[str] = None, status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    if status:
        query["status"] = status
    leaves = await db.leaves.find(query, {"_id": 0}).sort("start_date", -1).to_list(1000)
    for l in leaves:
        for f in ['start_date', 'end_date', 'created_at', 'approved_at']:
            if isinstance(l.get(f), str):
                l[f] = datetime.fromisoformat(l[f])
    return leaves

@api_router.post("/leaves")
async def create_leave(data: LeaveCreate, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": data.employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    # Validate ticket if requested
    if data.with_ticket:
        ticket_balance = emp.get("ticket_entitled", 1) - emp.get("ticket_used", 0)
        if ticket_balance <= 0:
            raise HTTPException(status_code=400, detail="No ticket balance available")
    
    leave = Leave(**data.model_dump(), employee_name=emp["name"])
    l_dict = leave.model_dump()
    for f in ['start_date', 'end_date', 'created_at', 'approved_at']:
        if l_dict.get(f):
            l_dict[f] = l_dict[f].isoformat()
    await db.leaves.insert_one(l_dict)
    
    # Notify admins about pending leave request
    if data.status == "pending":
        admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(100)
        for admin in admins:
            notif = Notification(
                user_id=admin["id"],
                title="New Leave Request",
                message=f"{emp['name']} has requested {data.days} days {data.leave_type} leave",
                type="leave_request",
                related_id=leave.id
            )
            n_dict = notif.model_dump()
            n_dict["created_at"] = n_dict["created_at"].isoformat()
            await db.notifications.insert_one(n_dict)
    
    return {k: v for k, v in l_dict.items() if k != '_id'}

@api_router.put("/leaves/{leave_id}/approve")
async def approve_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    
    await db.leaves.update_one({"id": leave_id}, {"$set": {
        "status": "approved",
        "approved_by": current_user.id,
        "approved_at": datetime.now(timezone.utc).isoformat()
    }})
    
    # If leave includes ticket, update ticket_used
    if leave.get("with_ticket"):
        emp_doc = await db.employees.find_one({"id": leave["employee_id"]}, {"_id": 0})
        if emp_doc:
            await db.employees.update_one({"id": leave["employee_id"]}, {"$set": {"ticket_used": emp_doc.get("ticket_used", 0) + 1}})
    
    # Notify employee
    emp = await db.employees.find_one({"id": leave["employee_id"]}, {"_id": 0})
    if emp and emp.get("user_id"):
        notif = Notification(
            user_id=emp["user_id"],
            title="Leave Approved",
            message=f"Your {leave['days']} days {leave['leave_type']} leave has been approved",
            type="leave_approved",
            related_id=leave_id
        )
        n_dict = notif.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    
    return {"message": "Leave approved"}

@api_router.put("/leaves/{leave_id}/reject")
async def reject_leave(leave_id: str, reason: Optional[dict] = None, current_user: User = Depends(get_current_user)):
    leave = await db.leaves.find_one({"id": leave_id}, {"_id": 0})
    if not leave:
        raise HTTPException(status_code=404, detail="Leave not found")
    
    rej_reason = reason.get("reason", "") if reason else ""
    await db.leaves.update_one({"id": leave_id}, {"$set": {
        "status": "rejected",
        "approved_by": current_user.id,
        "approved_at": datetime.now(timezone.utc).isoformat(),
        "rejection_reason": rej_reason
    }})
    
    # Notify employee
    emp = await db.employees.find_one({"id": leave["employee_id"]}, {"_id": 0})
    if emp and emp.get("user_id"):
        notif = Notification(
            user_id=emp["user_id"],
            title="Leave Rejected",
            message=f"Your {leave['leave_type']} leave request was rejected. {rej_reason}",
            type="leave_rejected",
            related_id=leave_id
        )
        n_dict = notif.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    
    return {"message": "Leave rejected"}

@api_router.delete("/leaves/{leave_id}")
async def delete_leave(leave_id: str, current_user: User = Depends(get_current_user)):
    result = await db.leaves.delete_one({"id": leave_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Leave not found")
    return {"message": "Leave deleted"}

# Employee Report PDF
@api_router.get("/employees/{emp_id}/report/pdf")
async def employee_report_pdf(emp_id: str, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not emp: raise HTTPException(status_code=404, detail="Employee not found")
    
    payments = await db.salary_payments.find({"employee_id": emp_id}, {"_id": 0}).sort("date", -1).to_list(1000)
    leaves = await db.leaves.find({"employee_id": emp_id}, {"_id": 0}).to_list(1000)
    deductions = await db.salary_deductions.find({"employee_id": emp_id}, {"_id": 0}).to_list(1000)
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    co_name = company.get("company_name", "Smart Standard Company")
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    title_s = ParagraphStyle('T', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#F5841F'), alignment=1, spaceAfter=5)
    body_s = ParagraphStyle('B', parent=styles['Normal'], fontSize=9, leading=14)
    
    logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.jpg"
    if not logo_path.exists(): logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.png"
    if logo_path.exists():
        from reportlab.platypus import Image as RLImage
        try:
            logo = RLImage(str(logo_path), width=1.3*inch, height=0.6*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
        except: pass
    
    elements.append(Paragraph(co_name.upper(), title_s))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#F5841F')))
    elements.append(Spacer(1, 0.15*inch))
    elements.append(Paragraph(f"<b>Employee Report - {emp['name']}</b>", ParagraphStyle('H', parent=styles['Heading2'], fontSize=12, alignment=1)))
    elements.append(Spacer(1, 0.1*inch))
    
    # Employee info
    info = [["Name:", emp["name"], "Position:", emp.get("position", "-")],
            ["Doc ID:", emp.get("document_id", "-"), "Salary:", f"SAR {emp.get('salary', 0):,.2f}"],
            ["Loan:", f"SAR {emp.get('loan_balance', 0):,.2f}", "Old Balance:", f"SAR {emp.get('old_salary_balance', 0):,.2f}"]]
    it = Table(info, colWidths=[1*inch, 2.2*inch, 1*inch, 2.2*inch])
    it.setStyle(TableStyle([('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'), ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 8), ('BOTTOMPADDING', (0,0), (-1,-1), 6)]))
    elements.append(it)
    elements.append(Spacer(1, 0.15*inch))
    
    # Payments table
    elements.append(Paragraph("<b>Salary Payments</b>", body_s))
    pay_rows = [["Date", "Type", "Period", "Mode", "Amount"]]
    for p in payments[:30]:
        dt = datetime.fromisoformat(p["date"]).strftime("%d %b %Y") if isinstance(p["date"], str) else p["date"].strftime("%d %b %Y")
        pay_rows.append([dt, p.get("payment_type","salary").replace("_"," ").title(), p.get("period",""), p.get("payment_mode",""), f"SAR {p['amount']:,.2f}"])
    pay_rows.append(["", "", "", "TOTAL", f"SAR {sum(p['amount'] for p in payments):,.2f}"])
    pt = Table(pay_rows, colWidths=[1.1*inch, 1.1*inch, 1.1*inch, 0.8*inch, 1.3*inch])
    pt.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#F5841F')), ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 7), ('GRID', (0,0), (-1,-1), 0.5, colors.grey), ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'), ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#FFF3E0'))]))
    elements.append(pt)
    elements.append(Spacer(1, 0.15*inch))
    
    # Leave summary
    annual_used = sum(l.get("days",0) for l in leaves if l.get("leave_type") == "annual" and l.get("status") == "approved")
    sick_used = sum(l.get("days",0) for l in leaves if l.get("leave_type") == "sick" and l.get("status") == "approved")
    elements.append(Paragraph(f"<b>Leave:</b> Annual: {annual_used}/{emp.get('annual_leave_entitled',30)} used | Sick: {sick_used}/{emp.get('sick_leave_entitled',15)} used", body_s))
    
    # Deductions
    if deductions:
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph(f"<b>Deductions:</b> Total SAR {sum(d['amount'] for d in deductions):,.2f}", body_s))
    
    doc.build(elements)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename=employee_report_{emp['name'].replace(' ','_')}.pdf"})

# Salary Acknowledgment
@api_router.post("/salary-payments/{payment_id}/acknowledge")
async def acknowledge_salary(payment_id: str, current_user: User = Depends(get_current_user)):
    payment = await db.salary_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    await db.salary_payments.update_one({"id": payment_id}, {"$set": {
        "acknowledged": True,
        "acknowledged_at": datetime.now(timezone.utc).isoformat()
    }})
    return {"message": "Payment acknowledged"}

# Notification Routes
@api_router.get("/notifications")
async def get_notifications(current_user: User = Depends(get_current_user)):
    notifs = await db.notifications.find({"user_id": current_user.id}, {"_id": 0}).sort("created_at", -1).to_list(50)
    for n in notifs:
        if isinstance(n.get('created_at'), str):
            n['created_at'] = datetime.fromisoformat(n['created_at'])
    return notifs

@api_router.get("/notifications/unread-count")
async def get_unread_count(current_user: User = Depends(get_current_user)):
    count = await db.notifications.count_documents({"user_id": current_user.id, "read": False})
    return {"count": count}

@api_router.post("/notifications/mark-read")
async def mark_notifications_read(current_user: User = Depends(get_current_user)):
    await db.notifications.update_many({"user_id": current_user.id, "read": False}, {"$set": {"read": True}})
    return {"message": "Notifications marked as read"}

# Attendance Routes
@api_router.post("/attendance/time-in")
async def time_in(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    existing = await db.attendance.find_one({"employee_id": emp["id"], "date": today}, {"_id": 0})
    if existing and existing.get("time_in"):
        raise HTTPException(status_code=400, detail="Already timed in today")
    now = datetime.now(timezone.utc)
    if existing:
        await db.attendance.update_one({"id": existing["id"]}, {"$set": {"time_in": now.isoformat()}})
    else:
        att = Attendance(employee_id=emp["id"], employee_name=emp["name"], date=today, time_in=now)
        a_dict = att.model_dump()
        a_dict["time_in"] = a_dict["time_in"].isoformat()
        a_dict["created_at"] = a_dict["created_at"].isoformat()
        await db.attendance.insert_one(a_dict)
    return {"message": "Timed in", "time": now.isoformat()}

@api_router.post("/attendance/time-out")
async def time_out(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    existing = await db.attendance.find_one({"employee_id": emp["id"], "date": today}, {"_id": 0})
    if not existing or not existing.get("time_in"):
        raise HTTPException(status_code=400, detail="Not timed in today")
    if existing.get("time_out"):
        raise HTTPException(status_code=400, detail="Already timed out today")
    now = datetime.now(timezone.utc)
    await db.attendance.update_one({"id": existing["id"]}, {"$set": {"time_out": now.isoformat()}})
    return {"message": "Timed out", "time": now.isoformat()}

@api_router.get("/attendance")
async def get_attendance(employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    records = await db.attendance.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return records

@api_router.get("/my/attendance")
async def get_my_attendance(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        return []
    return await db.attendance.find({"employee_id": emp["id"]}, {"_id": 0}).sort("date", -1).to_list(100)

# Employee Documents (multiple per employee)
@api_router.get("/employee-documents")
async def get_employee_documents(employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id:
        query["employee_id"] = employee_id
    docs = await db.employee_documents.find(query, {"_id": 0}).to_list(1000)
    now = datetime.now(timezone.utc)
    for d in docs:
        exp = d.get("expiry_date")
        if exp:
            if isinstance(exp, str): exp = datetime.fromisoformat(exp)
            if exp.tzinfo is None: exp = exp.replace(tzinfo=timezone.utc)
            d["days_until_expiry"] = (exp - now).days
    return docs

@api_router.post("/employee-documents")
async def create_employee_document(data: EmployeeDocumentCreate, current_user: User = Depends(get_current_user)):
    doc = EmployeeDocument(**data.model_dump())
    d_dict = doc.model_dump()
    for f in ['issue_date', 'expiry_date', 'created_at']:
        if d_dict.get(f): d_dict[f] = d_dict[f].isoformat()
    await db.employee_documents.insert_one(d_dict)
    return {k: v for k, v in d_dict.items() if k != '_id'}

@api_router.delete("/employee-documents/{doc_id}")
async def delete_employee_document(doc_id: str, current_user: User = Depends(get_current_user)):
    await db.employee_documents.delete_one({"id": doc_id})
    return {"message": "Document deleted"}

# Letter Generation
@api_router.post("/letters/generate")
async def generate_letter(body: dict, current_user: User = Depends(get_current_user)):
    emp_id = body.get("employee_id")
    letter_type = body.get("letter_type", "salary_certificate")
    emp = await db.employees.find_one({"id": emp_id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    buffer = BytesIO()
    doc_pdf = SimpleDocTemplate(buffer, pagesize=A4, topMargin=50, bottomMargin=50, leftMargin=60, rightMargin=60)
    elements = []
    styles = getSampleStyleSheet()
    
    title_s = ParagraphStyle('T', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#F5841F'), alignment=1, spaceAfter=5)
    body_s = ParagraphStyle('B', parent=styles['Normal'], fontSize=11, leading=18, spaceAfter=12)
    right_s = ParagraphStyle('R', parent=styles['Normal'], fontSize=10, alignment=2)
    
    logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.jpg"
    if not logo_path.exists(): logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.png"
    if logo_path.exists():
        from reportlab.platypus import Image as RLImage
        try:
            logo = RLImage(str(logo_path), width=1.5*inch, height=0.7*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
        except: pass
    
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    co_name = company.get("company_name", "Smart Standard Company")
    addr_parts = [company.get("address_line1",""), company.get("address_line2",""), company.get("city",""), company.get("country","")]
    co_addr = ", ".join([p for p in addr_parts if p])
    co_contact = " | ".join([p for p in [company.get("phone",""), company.get("email","")] if p])
    
    elements.append(Paragraph(co_name.upper(), title_s))
    if co_addr:
        elements.append(Paragraph(co_addr, ParagraphStyle('Addr', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)))
    if co_contact:
        elements.append(Paragraph(co_contact, ParagraphStyle('Contact', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#F5841F')))
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(f"Date: {datetime.now().strftime('%d %B %Y')}", right_s))
    elements.append(Spacer(1, 0.2*inch))
    
    name = emp["name"]
    doc_id = emp.get("document_id", "N/A")
    position = emp.get("position", "N/A")
    salary = emp.get("salary", 0)
    join = emp.get("join_date", "")
    if isinstance(join, str) and join: join = datetime.fromisoformat(join).strftime("%d %B %Y")
    elif hasattr(join, 'strftime'): join = join.strftime("%d %B %Y")
    else: join = "N/A"
    
    if letter_type == "salary_certificate":
        elements.append(Paragraph("<b>TO WHOM IT MAY CONCERN</b>", ParagraphStyle('C', parent=body_s, alignment=1, fontSize=13, spaceAfter=20)))
        elements.append(Paragraph(f"This is to certify that <b>{name}</b>, holding Document ID <b>{doc_id}</b>, is employed with Smart Standard Company as <b>{position}</b> since <b>{join}</b>.", body_s))
        elements.append(Paragraph(f"His/Her current monthly salary is <b>SAR {salary:,.2f}</b> (inclusive of all allowances).", body_s))
        elements.append(Paragraph("This certificate is issued upon the employee's request for whatever purpose it may serve.", body_s))
    elif letter_type == "employment":
        elements.append(Paragraph("<b>EMPLOYMENT CERTIFICATE</b>", ParagraphStyle('C', parent=body_s, alignment=1, fontSize=13, spaceAfter=20)))
        elements.append(Paragraph(f"This is to certify that <b>{name}</b>, Document ID: <b>{doc_id}</b>, has been employed with Smart Standard Company since <b>{join}</b> as <b>{position}</b>.", body_s))
        elements.append(Paragraph("The employee is currently active and in good standing with the company.", body_s))
        elements.append(Paragraph("This letter is issued upon the employee's request.", body_s))
    elif letter_type == "noc":
        elements.append(Paragraph("<b>NO OBJECTION CERTIFICATE</b>", ParagraphStyle('C', parent=body_s, alignment=1, fontSize=13, spaceAfter=20)))
        elements.append(Paragraph(f"This is to confirm that we have No Objection for our employee <b>{name}</b>, Document ID: <b>{doc_id}</b>, Position: <b>{position}</b>.", body_s))
        elements.append(Paragraph("This NOC is issued upon the employee's request.", body_s))
    elif letter_type == "experience":
        elements.append(Paragraph("<b>EXPERIENCE CERTIFICATE</b>", ParagraphStyle('C', parent=body_s, alignment=1, fontSize=13, spaceAfter=20)))
        elements.append(Paragraph(f"This is to certify that <b>{name}</b>, Document ID: <b>{doc_id}</b>, has worked with Smart Standard Company as <b>{position}</b> from <b>{join}</b>.", body_s))
        elements.append(Paragraph("During the tenure, the employee demonstrated professionalism and dedication. We wish them success in future endeavors.", body_s))
    
    elements.append(Spacer(1, 0.5*inch))
    elements.append(Paragraph("Authorized Signatory", ParagraphStyle('Sig', parent=body_s, fontSize=10)))
    elements.append(Paragraph("_________________________", body_s))
    elements.append(Paragraph("Smart Standard Company", ParagraphStyle('Co', parent=body_s, fontSize=9, textColor=colors.grey)))
    
    doc_pdf.build(elements)
    buffer.seek(0)
    fname = f"{letter_type}_{name.replace(' ','_')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})

# Employee Portal - My Data
@api_router.get("/my/employee-profile")
async def get_my_employee_profile(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    return emp

@api_router.get("/my/payments")
async def get_my_payments(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    payments = await db.salary_payments.find({"employee_id": emp["id"]}, {"_id": 0}).sort("date", -1).to_list(1000)
    for p in payments:
        for f in ['date', 'created_at', 'acknowledged_at']:
            if isinstance(p.get(f), str):
                p[f] = datetime.fromisoformat(p[f])
    return payments

@api_router.get("/my/leaves")
async def get_my_leaves(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    leaves = await db.leaves.find({"employee_id": emp["id"]}, {"_id": 0}).sort("start_date", -1).to_list(1000)
    for l in leaves:
        for f in ['start_date', 'end_date', 'created_at', 'approved_at']:
            if isinstance(l.get(f), str):
                l[f] = datetime.fromisoformat(l[f])
    return leaves

@api_router.post("/my/apply-leave")
async def apply_leave(data: LeaveCreate, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    data.employee_id = emp["id"]
    data.status = "pending"
    return await create_leave(data, current_user)

# Employee Requests (letter, loan, etc.)
@api_router.get("/employee-requests")
async def get_employee_requests(status: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    reqs = await db.employee_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    for r in reqs:
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return reqs

@api_router.post("/my/request")
async def create_employee_request(data: EmployeeRequestCreate, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    req = EmployeeRequest(**data.model_dump(), employee_id=emp["id"], employee_name=emp["name"])
    r_dict = req.model_dump()
    r_dict["created_at"] = r_dict["created_at"].isoformat()
    await db.employee_requests.insert_one(r_dict)
    # Notify admins
    admins = await db.users.find({"role": {"$in": ["admin", "manager"]}}, {"_id": 0}).to_list(100)
    for admin in admins:
        n = Notification(user_id=admin["id"], title=f"New Request: {data.request_type.replace('_',' ').title()}", message=f"{emp['name']}: {data.subject}", type="employee_request", related_id=req.id)
        n_dict = n.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    return {k: v for k, v in r_dict.items() if k != '_id'}

@api_router.get("/my/requests")
async def get_my_requests(current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"user_id": current_user.id}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="No employee profile linked")
    reqs = await db.employee_requests.find({"employee_id": emp["id"]}, {"_id": 0}).sort("created_at", -1).to_list(100)
    for r in reqs:
        if isinstance(r.get('created_at'), str):
            r['created_at'] = datetime.fromisoformat(r['created_at'])
    return reqs

@api_router.put("/employee-requests/{req_id}/respond")
async def respond_to_request(req_id: str, body: dict, current_user: User = Depends(get_current_user)):
    req = await db.employee_requests.find_one({"id": req_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    status = body.get("status", "approved")
    response = body.get("response", "")
    await db.employee_requests.update_one({"id": req_id}, {"$set": {"status": status, "response": response, "processed_by": current_user.id}})
    # Notify employee
    emp = await db.employees.find_one({"id": req["employee_id"]}, {"_id": 0})
    if emp and emp.get("user_id"):
        n = Notification(user_id=emp["user_id"], title=f"Request {status.title()}", message=f"Your {req['request_type'].replace('_',' ')} request: {response or status}", type="request_response", related_id=req_id)
        n_dict = n.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    return {"message": f"Request {status}"}

# Send Announcement to employees
@api_router.post("/announcements/send")
async def send_announcement(body: dict, current_user: User = Depends(get_current_user)):
    title = body.get("title", "Announcement")
    message = body.get("message", "")
    target = body.get("target", "all")  # "all" or specific employee_id
    
    if target == "all":
        employees = await db.employees.find({"user_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(1000)
        for emp in employees:
            n = Notification(user_id=emp["user_id"], title=title, message=message, type="announcement")
            n_dict = n.model_dump()
            n_dict["created_at"] = n_dict["created_at"].isoformat()
            await db.notifications.insert_one(n_dict)
        return {"message": f"Announcement sent to {len(employees)} employees"}
    else:
        emp = await db.employees.find_one({"id": target}, {"_id": 0})
        if emp and emp.get("user_id"):
            n = Notification(user_id=emp["user_id"], title=title, message=message, type="announcement")
            n_dict = n.model_dump()
            n_dict["created_at"] = n_dict["created_at"].isoformat()
            await db.notifications.insert_one(n_dict)
        return {"message": f"Announcement sent to {emp['name'] if emp else 'unknown'}"}

# Document Routes
@api_router.get("/documents")
async def get_documents(current_user: User = Depends(get_current_user)):
    docs = await db.documents.find({}, {"_id": 0}).to_list(1000)
    now = datetime.now(timezone.utc)
    for d in docs:
        for f in ['created_at', 'issue_date', 'expiry_date']:
            if isinstance(d.get(f), str):
                d[f] = datetime.fromisoformat(d[f])
        if d.get('expiry_date'):
            exp = d['expiry_date'] if isinstance(d['expiry_date'], datetime) else datetime.fromisoformat(str(d['expiry_date']))
            if exp.tzinfo is None:
                exp = exp.replace(tzinfo=timezone.utc)
            days_left = (exp - now).days
            d['days_until_expiry'] = days_left
            if days_left < 0:
                d['status'] = 'expired'
            elif days_left <= d.get('alert_days', 30):
                d['status'] = 'expiring_soon'
            else:
                d['status'] = 'active'
    return docs

@api_router.post("/documents")
async def create_document(data: DocumentCreate, current_user: User = Depends(get_current_user)):
    doc = Document(**data.model_dump())
    doc_dict = doc.model_dump()
    for f in ['created_at', 'issue_date', 'expiry_date']:
        if doc_dict.get(f):
            doc_dict[f] = doc_dict[f].isoformat()
    await db.documents.insert_one(doc_dict)
    return {k: v for k, v in doc_dict.items() if k != '_id'}

@api_router.put("/documents/{doc_id}")
async def update_document(doc_id: str, data: DocumentCreate, current_user: User = Depends(get_current_user)):
    result = await db.documents.find_one({"id": doc_id})
    if not result:
        raise HTTPException(status_code=404, detail="Document not found")
    update = data.model_dump()
    for f in ['issue_date', 'expiry_date']:
        if update.get(f):
            update[f] = update[f].isoformat()
    await db.documents.update_one({"id": doc_id}, {"$set": update})
    updated = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    return updated

@api_router.delete("/documents/{doc_id}")
async def delete_document(doc_id: str, current_user: User = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    if doc.get("file_path") and os.path.exists(doc["file_path"]):
        os.remove(doc["file_path"])
    await db.documents.delete_one({"id": doc_id})
    return {"message": "Document deleted"}

# Document File Upload
@api_router.post("/documents/{doc_id}/upload")
async def upload_document_file(doc_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    upload_dir = ROOT_DIR / "uploads" / "documents"
    upload_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix
    file_path = upload_dir / f"{doc_id}{ext}"
    
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    
    await db.documents.update_one({"id": doc_id}, {"$set": {"file_path": str(file_path), "file_name": file.filename}})
    return {"message": "File uploaded", "file_name": file.filename}

@api_router.get("/documents/{doc_id}/download")
async def download_document_file(doc_id: str, current_user: User = Depends(get_current_user)):
    doc = await db.documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc or not doc.get("file_path"):
        raise HTTPException(status_code=404, detail="No file attached")
    if not os.path.exists(doc["file_path"]):
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(doc["file_path"], filename=doc.get("file_name", "document"))

# Company Settings (address, etc.)
@api_router.get("/settings/company")
async def get_company_settings(current_user: User = Depends(get_current_user)):
    settings = await db.company_settings.find_one({}, {"_id": 0})
    return settings or {"company_name": "Smart Standard Company", "address_line1": "", "address_line2": "", "city": "", "country": "", "phone": "", "email": "", "cr_number": "", "vat_number": "", "vat_enabled": False, "vat_rate": 15}

@api_router.post("/settings/company")
async def save_company_settings(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.company_settings.find_one({})
    data = {k: body.get(k, "") for k in ["company_name", "address_line1", "address_line2", "city", "country", "phone", "email", "cr_number", "vat_number"]}
    data["vat_enabled"] = body.get("vat_enabled", False)
    data["vat_rate"] = float(body.get("vat_rate", 15) or 15)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    if existing:
        await db.company_settings.update_one({}, {"$set": data})
    else:
        await db.company_settings.insert_one(data)
    return {"message": "Company settings saved"}

# Company Logo Upload
@api_router.post("/settings/upload-logo")
async def upload_logo(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    upload_dir = ROOT_DIR / "uploads" / "logos"
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_path = upload_dir / "company_logo.png"
    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)
    return {"message": "Logo uploaded"}

# Payslip PDF Generation
@api_router.get("/salary-payments/{payment_id}/payslip")
async def generate_payslip(payment_id: str, current_user: User = Depends(get_current_user)):
    payment = await db.salary_payments.find_one({"id": payment_id}, {"_id": 0})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    emp = await db.employees.find_one({"id": payment["employee_id"]}, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get all payments for this employee in same period
    period_payments = await db.salary_payments.find(
        {"employee_id": payment["employee_id"], "period": payment["period"]}, {"_id": 0}
    ).to_list(100)
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    elements = []
    styles = getSampleStyleSheet()
    
    # Header with logo
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=20, textColor=colors.HexColor('#F5841F'), alignment=1, spaceAfter=5)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=1, spaceAfter=20)
    
    logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.jpg"
    if not logo_path.exists():
        logo_path = ROOT_DIR / "uploads" / "logos" / "company_logo.png"
    if logo_path.exists():
        from reportlab.platypus import Image as RLImage
        try:
            logo = RLImage(str(logo_path), width=1.5*inch, height=0.7*inch)
            logo.hAlign = 'CENTER'
            elements.append(logo)
            elements.append(Spacer(1, 0.1*inch))
        except:
            pass
    
    company = await db.company_settings.find_one({}, {"_id": 0}) or {}
    co_name = company.get("company_name", "Smart Standard Company")
    addr_parts = [company.get("address_line1",""), company.get("address_line2",""), company.get("city",""), company.get("country","")]
    co_addr = ", ".join([p for p in addr_parts if p])
    
    elements.append(Paragraph(co_name.upper(), title_style))
    if co_addr:
        elements.append(Paragraph(co_addr, ParagraphStyle('PAddr', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1, spaceAfter=3)))
    elements.append(Paragraph("Pay Slip", sub_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#F5841F')))
    elements.append(Spacer(1, 0.2*inch))
    
    # Employee Details
    emp_data = [
        ["Employee Name:", emp["name"], "Period:", payment.get("period", "-")],
        ["Position:", emp.get("position", "-"), "Document ID:", emp.get("document_id", "-")],
        ["Date:", datetime.fromisoformat(payment["date"]).strftime("%d %b %Y") if isinstance(payment["date"], str) else payment["date"].strftime("%d %b %Y"), "Payment Mode:", payment.get("payment_mode", "-").upper()],
    ]
    emp_table = Table(emp_data, colWidths=[1.2*inch, 2.3*inch, 1.2*inch, 2.3*inch])
    emp_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(emp_table)
    elements.append(Spacer(1, 0.2*inch))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(Spacer(1, 0.15*inch))
    
    # Payment Breakdown
    elements.append(Paragraph("Payment Details", ParagraphStyle('H2', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#F5841F'))))
    
    salary_total = sum(p["amount"] for p in period_payments if p.get("payment_type") == "salary")
    overtime = sum(p["amount"] for p in period_payments if p.get("payment_type") == "overtime")
    advance = sum(p["amount"] for p in period_payments if p.get("payment_type") == "advance")
    loan_repay = sum(p["amount"] for p in period_payments if p.get("payment_type") == "loan_repayment")
    tickets = sum(p["amount"] for p in period_payments if p.get("payment_type") == "tickets")
    id_card = sum(p["amount"] for p in period_payments if p.get("payment_type") == "id_card")
    
    pay_rows = [["Description", "Amount"]]
    pay_rows.append(["Monthly Salary", f"SAR {emp.get('salary', 0):.2f}"])
    if salary_total > 0: pay_rows.append(["Salary Paid", f"SAR {salary_total:.2f}"])
    if overtime > 0: pay_rows.append(["Overtime", f"SAR {overtime:.2f}"])
    if advance > 0: pay_rows.append(["Advance / Loan", f"SAR {advance:.2f}"])
    if loan_repay > 0: pay_rows.append(["Loan Repayment (Deduction)", f"-${loan_repay:.2f}"])
    if tickets > 0: pay_rows.append(["Tickets", f"SAR {tickets:.2f}"])
    if id_card > 0: pay_rows.append(["ID Card", f"SAR {id_card:.2f}"])
    
    net = salary_total + overtime - loan_repay
    balance = emp.get("salary", 0) - salary_total
    pay_rows.append(["", ""])
    pay_rows.append(["Net Payment", f"SAR {net:.2f}"])
    pay_rows.append(["Salary Balance", f"SAR {balance:.2f}"])
    
    pay_table = Table(pay_rows, colWidths=[4.5*inch, 2.5*inch])
    pay_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -2), (-1, -1), colors.HexColor('#F5F3FF')),
    ]))
    elements.append(pay_table)
    elements.append(Spacer(1, 0.4*inch))
    
    # Acknowledgment section
    ack_text = "I acknowledge receipt of the above payment."
    elements.append(Paragraph(ack_text, ParagraphStyle('Ack', parent=styles['Normal'], fontSize=10)))
    elements.append(Spacer(1, 0.3*inch))
    
    sig_data = [
        ["Employee Signature:", "____________________", "Date:", "____________________"],
        ["", "", "", ""],
        ["Authorized By:", "____________________", "Company Stamp:", ""],
    ]
    sig_table = Table(sig_data, colWidths=[1.3*inch, 2.2*inch, 1.3*inch, 2.2*inch])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(sig_table)
    
    if payment.get("acknowledged"):
        ack_at = payment.get("acknowledged_at", "")
        if isinstance(ack_at, str) and ack_at:
            ack_at = datetime.fromisoformat(ack_at).strftime("%d %b %Y %H:%M")
        elements.append(Spacer(1, 0.2*inch))
        elements.append(Paragraph(f"Digitally acknowledged on: {ack_at}", ParagraphStyle('Dig', parent=styles['Normal'], fontSize=8, textColor=colors.green)))
    
    doc.build(elements)
    buffer.seek(0)
    
    fname = f"payslip_{emp['name'].replace(' ', '_')}_{payment.get('period', '').replace(' ', '_')}.pdf"
    return StreamingResponse(buffer, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})

# Document Expiry Alerts
@api_router.get("/documents/alerts/upcoming")
async def get_expiry_alerts(current_user: User = Depends(get_current_user)):
    docs = await db.documents.find({}, {"_id": 0}).to_list(1000)
    employees = await db.employees.find({}, {"_id": 0}).to_list(1000)
    now = datetime.now(timezone.utc)
    alerts = []
    
    for d in docs:
        exp = d.get('expiry_date')
        if not exp:
            continue
        if isinstance(exp, str):
            exp = datetime.fromisoformat(exp)
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        days_left = (exp - now).days
        alert_days = d.get('alert_days', 30)
        if days_left <= alert_days:
            alerts.append({
                "type": "document",
                "name": d["name"],
                "related_to": d.get("related_to", "-"),
                "expiry_date": exp.isoformat(),
                "days_left": days_left,
                "status": "expired" if days_left < 0 else "expiring_soon",
                "id": d["id"]
            })
    
    for emp in employees:
        exp = emp.get('document_expiry')
        if not exp:
            continue
        if isinstance(exp, str):
            exp = datetime.fromisoformat(exp)
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        days_left = (exp - now).days
        if days_left <= 30:
            alerts.append({
                "type": "employee_document",
                "name": f"{emp['name']} - ID Document",
                "related_to": emp["name"],
                "expiry_date": exp.isoformat(),
                "days_left": days_left,
                "status": "expired" if days_left < 0 else "expiring_soon",
                "id": emp["id"]
            })
    
    alerts.sort(key=lambda x: x["days_left"])
    return alerts

# Supplier Payment Breakdown (per supplier, cash/bank by branch)
@api_router.get("/suppliers/{supplier_id}/payment-breakdown")
async def get_supplier_payment_breakdown(supplier_id: str, current_user: User = Depends(get_current_user)):
    supplier = await db.suppliers.find_one({"id": supplier_id}, {"_id": 0})
    if not supplier:
        raise HTTPException(status_code=404, detail="Supplier not found")
    
    payments = await db.supplier_payments.find({"supplier_id": supplier_id}, {"_id": 0}).to_list(10000)
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    
    total_cash = sum(p["amount"] for p in payments if p.get("payment_mode") == "cash")
    total_bank = sum(p["amount"] for p in payments if p.get("payment_mode") == "bank")
    total_credit = sum(p["amount"] for p in payments if p.get("payment_mode") == "credit")
    
    by_branch = {}
    for p in payments:
        bid = p.get("branch_id")
        bname = branches.get(bid, "No Branch") if bid else "No Branch"
        if bname not in by_branch:
            by_branch[bname] = {"cash": 0, "bank": 0, "credit": 0, "total": 0}
        mode = p.get("payment_mode", "cash")
        by_branch[bname][mode] = by_branch[bname].get(mode, 0) + p["amount"]
        by_branch[bname]["total"] += p["amount"]
    
    return {
        "supplier_id": supplier_id, "supplier_name": supplier["name"],
        "total_cash": total_cash, "total_bank": total_bank, "total_credit": total_credit,
        "total": total_cash + total_bank + total_credit,
        "by_branch": by_branch
    }

# All Suppliers Payment Summary (for supplier cards)
@api_router.get("/suppliers/payment-summaries")
async def get_all_supplier_summaries(current_user: User = Depends(get_current_user)):
    suppliers = await db.suppliers.find({}, {"_id": 0}).to_list(1000)
    payments = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    
    result = {}
    for sup in suppliers:
        sid = sup["id"]
        sp = [p for p in payments if p.get("supplier_id") == sid]
        by_branch = {}
        for p in sp:
            bid = p.get("branch_id")
            bname = branches.get(bid, "No Branch") if bid else "No Branch"
            if bname not in by_branch:
                by_branch[bname] = {"cash": 0, "bank": 0}
            if p.get("payment_mode") == "cash":
                by_branch[bname]["cash"] += p["amount"]
            elif p.get("payment_mode") == "bank":
                by_branch[bname]["bank"] += p["amount"]
        
        result[sid] = {
            "cash": sum(p["amount"] for p in sp if p.get("payment_mode") == "cash"),
            "bank": sum(p["amount"] for p in sp if p.get("payment_mode") == "bank"),
            "by_branch": by_branch
        }
    return result

# Branch-to-Branch Dues (ALL cross-branch payments: supplier, salary, expenses, transfers)
@api_router.get("/reports/branch-dues")
async def get_branch_dues(current_user: User = Depends(get_current_user)):
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}
    
    expenses = await db.expenses.find({"branch_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    sp = await db.supplier_payments.find({"supplier_id": {"$exists": True, "$ne": None}, "branch_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    salary_payments = await db.salary_payments.find({"branch_id": {"$exists": True, "$ne": None}}, {"_id": 0}).to_list(10000)
    employees = {e["id"]: e for e in await db.employees.find({}, {"_id": 0}).to_list(1000)}
    suppliers = {s["id"]: s for s in await db.suppliers.find({}, {"_id": 0}).to_list(1000)}
    transfers = await db.cash_transfers.find({}, {"_id": 0}).to_list(10000)
    
    dues = {}
    
    for p in sp:
        pay_b = p.get("branch_id")
        sup = suppliers.get(p.get("supplier_id"), {})
        sup_b = sup.get("branch_id")
        if pay_b and sup_b and pay_b != sup_b:
            key = f"{branch_map.get(pay_b, '?')} paid for {branch_map.get(sup_b, '?')} (supplier)"
            dues[key] = dues.get(key, 0) + p["amount"]
    
    for p in salary_payments:
        pay_b = p.get("branch_id")
        emp = employees.get(p.get("employee_id"), {})
        emp_b = emp.get("branch_id")
        if pay_b and emp_b and pay_b != emp_b:
            key = f"{branch_map.get(pay_b, '?')} paid for {branch_map.get(emp_b, '?')} (salary)"
            dues[key] = dues.get(key, 0) + p["amount"]
    
    for e in expenses:
        pay_b = e.get("branch_id")
        if e.get("supplier_id"):
            sup = suppliers.get(e["supplier_id"], {})
            sup_b = sup.get("branch_id")
            if pay_b and sup_b and pay_b != sup_b:
                key = f"{branch_map.get(pay_b, '?')} paid for {branch_map.get(sup_b, '?')} (expense)"
                dues[key] = dues.get(key, 0) + e["amount"]
    
    for t in transfers:
        from_b = t.get("from_branch_id")
        to_b = t.get("to_branch_id")
        if from_b and to_b and from_b != to_b:
            key = f"{branch_map.get(from_b, 'Office')} sent to {branch_map.get(to_b, 'Office')} (transfer)"
            dues[key] = dues.get(key, 0) + t["amount"]
    
    return {"dues": dues, "total_cross_branch": sum(dues.values())}

# Salary History Routes
@api_router.get("/salary-history/{emp_id}")
async def get_salary_history(emp_id: str, current_user: User = Depends(get_current_user)):
    history = await db.salary_history.find({"employee_id": emp_id}, {"_id": 0}).sort("effective_date", -1).to_list(100)
    for h in history:
        if isinstance(h.get("effective_date"), str): h["effective_date"] = datetime.fromisoformat(h["effective_date"])
    return history

@api_router.post("/salary-history")
async def add_salary_history(body: dict, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": body["employee_id"]}, {"_id": 0})
    if not emp: raise HTTPException(status_code=404, detail="Employee not found")
    old_salary = emp.get("salary", 0)
    new_salary = float(body["new_salary"])
    record = SalaryHistory(employee_id=body["employee_id"], old_salary=old_salary, new_salary=new_salary,
                           effective_date=datetime.fromisoformat(body["effective_date"]), reason=body.get("reason", ""))
    r_dict = record.model_dump()
    r_dict["effective_date"] = r_dict["effective_date"].isoformat()
    r_dict["created_at"] = r_dict["created_at"].isoformat()
    await db.salary_history.insert_one(r_dict)
    await db.employees.update_one({"id": body["employee_id"]}, {"$set": {"salary": new_salary}})
    return {k: v for k, v in r_dict.items() if k != '_id'}

# Branch Payback Routes
@api_router.get("/branch-paybacks")
async def get_branch_paybacks(current_user: User = Depends(get_current_user)):
    paybacks = await db.branch_paybacks.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for p in paybacks:
        for f in ['date', 'created_at']:
            if isinstance(p.get(f), str): p[f] = datetime.fromisoformat(p[f])
    return paybacks

@api_router.post("/branch-paybacks")
async def create_branch_payback(body: dict, current_user: User = Depends(get_current_user)):
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    payback = BranchPayback(
        from_branch_id=body["from_branch_id"], to_branch_id=body["to_branch_id"],
        from_branch_name=branches.get(body["from_branch_id"], "?"), to_branch_name=branches.get(body["to_branch_id"], "?"),
        amount=float(body["amount"]), payment_mode=body.get("payment_mode", "cash"),
        date=datetime.fromisoformat(body["date"]), notes=body.get("notes", ""), created_by=current_user.id
    )
    p_dict = payback.model_dump()
    p_dict["date"] = p_dict["date"].isoformat()
    p_dict["created_at"] = p_dict["created_at"].isoformat()
    await db.branch_paybacks.insert_one(p_dict)
    return {k: v for k, v in p_dict.items() if k != '_id'}

# Enhanced Branch Dues with payback deduction
@api_router.get("/reports/branch-dues-net")
async def get_branch_dues_net(current_user: User = Depends(get_current_user)):
    dues_resp = await get_branch_dues(current_user)
    paybacks = await db.branch_paybacks.find({}, {"_id": 0}).to_list(10000)
    
    # Build payback totals
    payback_totals = {}
    for p in paybacks:
        key = f"{p['from_branch_name']} paid back {p['to_branch_name']}"
        payback_totals[key] = payback_totals.get(key, 0) + p["amount"]
    
    return {"dues": dues_resp["dues"], "paybacks": payback_totals, "total_dues": dues_resp["total_cross_branch"], "total_paybacks": sum(payback_totals.values())}

# Partner Routes
@api_router.get("/partners")
async def get_partners(current_user: User = Depends(get_current_user)):
    partners = await db.partners.find({}, {"_id": 0}).to_list(100)
    transactions = await db.partner_transactions.find({}, {"_id": 0}).to_list(10000)
    for p in partners:
        pt = [t for t in transactions if t.get("partner_id") == p["id"]]
        invested = sum(t["amount"] for t in pt if t.get("transaction_type") in ["investment"])
        withdrawn = sum(t["amount"] for t in pt if t.get("transaction_type") in ["withdrawal", "profit_share", "expense"])
        p["total_invested"] = invested
        p["total_withdrawn"] = withdrawn
        p["balance"] = invested - withdrawn
    return partners

@api_router.post("/partners")
async def create_partner(data: PartnerCreate, current_user: User = Depends(get_current_user)):
    partner = Partner(**data.model_dump())
    p_dict = partner.model_dump()
    p_dict["created_at"] = p_dict["created_at"].isoformat()
    await db.partners.insert_one(p_dict)
    return {k: v for k, v in p_dict.items() if k != '_id'}

@api_router.put("/partners/{partner_id}")
async def update_partner(partner_id: str, data: PartnerCreate, current_user: User = Depends(get_current_user)):
    await db.partners.update_one({"id": partner_id}, {"$set": data.model_dump()})
    return await db.partners.find_one({"id": partner_id}, {"_id": 0})

@api_router.delete("/partners/{partner_id}")
async def delete_partner(partner_id: str, current_user: User = Depends(get_current_user)):
    await db.partners.delete_one({"id": partner_id})
    return {"message": "Partner deleted"}

@api_router.get("/partner-transactions")
async def get_partner_transactions(partner_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if partner_id: query["partner_id"] = partner_id
    txns = await db.partner_transactions.find(query, {"_id": 0}).sort("date", -1).to_list(10000)
    for t in txns:
        for f in ['date', 'created_at']:
            if isinstance(t.get(f), str): t[f] = datetime.fromisoformat(t[f])
    return txns

@api_router.post("/partner-transactions")
async def create_partner_transaction(data: PartnerTransactionCreate, current_user: User = Depends(get_current_user)):
    partner = await db.partners.find_one({"id": data.partner_id}, {"_id": 0})
    if not partner: raise HTTPException(status_code=404, detail="Partner not found")
    txn = PartnerTransaction(**data.model_dump(), partner_name=partner["name"], created_by=current_user.id)
    t_dict = txn.model_dump()
    t_dict["date"] = t_dict["date"].isoformat()
    t_dict["created_at"] = t_dict["created_at"].isoformat()
    if t_dict.get("branch_id") == '': t_dict["branch_id"] = None
    await db.partner_transactions.insert_one(t_dict)
    return {k: v for k, v in t_dict.items() if k != '_id'}

@api_router.delete("/partner-transactions/{txn_id}")
async def delete_partner_transaction(txn_id: str, current_user: User = Depends(get_current_user)):
    await db.partner_transactions.delete_one({"id": txn_id})
    return {"message": "Transaction deleted"}

# Fines & Penalties Routes
@api_router.get("/fines")
async def get_fines(current_user: User = Depends(get_current_user)):
    fines = await db.fines.find({}, {"_id": 0}).sort("fine_date", -1).to_list(1000)
    for f in fines:
        for k in ['fine_date', 'due_date', 'paid_date', 'created_at']:
            if isinstance(f.get(k), str): f[k] = datetime.fromisoformat(f[k])
    return fines

@api_router.post("/fines")
async def create_fine(data: FineCreate, current_user: User = Depends(get_current_user)):
    fine = Fine(**data.model_dump())
    f_dict = fine.model_dump()
    for k in ['fine_date', 'due_date', 'created_at']:
        if f_dict.get(k): f_dict[k] = f_dict[k].isoformat()
    for k in ['branch_id', 'employee_id']:
        if f_dict.get(k) == '': f_dict[k] = None
    await db.fines.insert_one(f_dict)
    
    # Notify employee if fine is charged to them
    if data.employee_id:
        emp = await db.employees.find_one({"id": data.employee_id}, {"_id": 0})
        if emp and emp.get("user_id"):
            msg = f"SAR {data.amount:.2f} fine ({data.fine_type}) from {data.department}. {data.description}"
            if data.deduct_from_salary and data.monthly_deduction:
                msg += f" | SAR {data.monthly_deduction:.2f}/month will be deducted from salary."
            n = Notification(user_id=emp["user_id"], title="Fine Charged to You", message=msg, type="fine_charged", related_id=fine.id)
            n_dict = n.model_dump()
            n_dict["created_at"] = n_dict["created_at"].isoformat()
            await db.notifications.insert_one(n_dict)
    
    return {k: v for k, v in f_dict.items() if k != '_id'}

@api_router.put("/fines/{fine_id}/pay")
async def pay_fine(fine_id: str, body: dict, current_user: User = Depends(get_current_user)):
    fine = await db.fines.find_one({"id": fine_id}, {"_id": 0})
    if not fine: raise HTTPException(status_code=404, detail="Fine not found")
    amount = float(body.get("amount", 0))
    mode = body.get("payment_mode", "cash")
    new_paid = fine.get("paid_amount", 0) + amount
    status = "paid" if new_paid >= fine["amount"] else "partial"
    await db.fines.update_one({"id": fine_id}, {"$set": {"paid_amount": new_paid, "payment_status": status, "payment_mode": mode, "paid_date": datetime.now(timezone.utc).isoformat()}})
    return {"message": f"Fine payment recorded", "paid_amount": new_paid, "status": status}

@api_router.delete("/fines/{fine_id}")
async def delete_fine(fine_id: str, current_user: User = Depends(get_current_user)):
    fine = await db.fines.find_one({"id": fine_id}, {"_id": 0})
    if fine and fine.get("file_path") and os.path.exists(fine["file_path"]): os.remove(fine["file_path"])
    await db.fines.delete_one({"id": fine_id})
    return {"message": "Fine deleted"}

@api_router.post("/fines/{fine_id}/upload")
async def upload_fine_proof(fine_id: str, file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    upload_dir = ROOT_DIR / "uploads" / "fines"
    upload_dir.mkdir(parents=True, exist_ok=True)
    ext = Path(file.filename).suffix
    file_path = upload_dir / f"{fine_id}{ext}"
    with open(file_path, "wb") as f: f.write(await file.read())
    await db.fines.update_one({"id": fine_id}, {"$set": {"file_path": str(file_path), "file_name": file.filename}})
    return {"message": "Proof uploaded", "file_name": file.filename}

@api_router.get("/fines/{fine_id}/download")
async def download_fine_proof(fine_id: str, current_user: User = Depends(get_current_user)):
    fine = await db.fines.find_one({"id": fine_id}, {"_id": 0})
    if not fine or not fine.get("file_path"): raise HTTPException(status_code=404, detail="No file")
    if not os.path.exists(fine["file_path"]): raise HTTPException(status_code=404, detail="File missing")
    return FileResponse(fine["file_path"], filename=fine.get("file_name", "proof"))

# Capital Expense / Goodwill (branch acquisition costs)
@api_router.get("/capital-expenses")
async def get_capital_expenses(current_user: User = Depends(get_current_user)):
    return await db.capital_expenses.find({}, {"_id": 0}).sort("date", -1).to_list(1000)

@api_router.post("/capital-expenses")
async def create_capital_expense(body: dict, current_user: User = Depends(get_current_user)):
    doc = {"id": str(uuid.uuid4()), "title": body.get("title",""), "category": body.get("category","goodwill"),
           "description": body.get("description",""), "amount": float(body.get("amount",0)),
           "branch_id": body.get("branch_id") or None, "payment_mode": body.get("payment_mode","cash"),
           "date": body.get("date", datetime.now(timezone.utc).isoformat()),
           "notes": body.get("notes",""), "created_by": current_user.id,
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.capital_expenses.insert_one(doc)
    return {k: v for k, v in doc.items() if k != '_id'}

@api_router.delete("/capital-expenses/{cap_id}")
async def delete_capital_expense(cap_id: str, current_user: User = Depends(get_current_user)):
    await db.capital_expenses.delete_one({"id": cap_id})
    return {"message": "Deleted"}

# Salary Deductions Routes
@api_router.get("/salary-deductions")
async def get_salary_deductions(employee_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if employee_id: query["employee_id"] = employee_id
    deductions = await db.salary_deductions.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    for d in deductions:
        for k in ['date', 'created_at']:
            if isinstance(d.get(k), str): d[k] = datetime.fromisoformat(d[k])
    return deductions

@api_router.post("/salary-deductions")
async def create_salary_deduction(data: SalaryDeductionCreate, current_user: User = Depends(get_current_user)):
    emp = await db.employees.find_one({"id": data.employee_id}, {"_id": 0})
    if not emp: raise HTTPException(status_code=404, detail="Employee not found")
    deduction = SalaryDeduction(**data.model_dump(), employee_name=emp["name"], created_by=current_user.id)
    d_dict = deduction.model_dump()
    d_dict["date"] = d_dict["date"].isoformat()
    d_dict["created_at"] = d_dict["created_at"].isoformat()
    for k in ['branch_id', 'fine_id']:
        if d_dict.get(k) == '': d_dict[k] = None
    await db.salary_deductions.insert_one(d_dict)
    
    # Notify employee about deduction
    if emp.get("user_id"):
        type_label = data.deduction_type.replace('_', ' ').title()
        n = Notification(
            user_id=emp["user_id"],
            title=f"Salary Deduction: {type_label}",
            message=f"SAR {data.amount:.2f} deducted from your salary ({data.period}). Reason: {data.reason}",
            type="salary_deduction",
            related_id=deduction.id
        )
        n_dict = n.model_dump()
        n_dict["created_at"] = n_dict["created_at"].isoformat()
        await db.notifications.insert_one(n_dict)
    
    return {k: v for k, v in d_dict.items() if k != '_id'}

@api_router.delete("/salary-deductions/{ded_id}")
async def delete_salary_deduction(ded_id: str, current_user: User = Depends(get_current_user)):
    await db.salary_deductions.delete_one({"id": ded_id})
    return {"message": "Deduction deleted"}

# Items Master (Products/Services)
@api_router.get("/items")
async def get_items(current_user: User = Depends(get_current_user)):
    items = await db.items.find({}, {"_id": 0}).to_list(1000)
    return items

@api_router.post("/items")
async def create_item(data: ItemCreate, current_user: User = Depends(get_current_user)):
    item = Item(**data.model_dump())
    i_dict = item.model_dump()
    i_dict["created_at"] = i_dict["created_at"].isoformat()
    await db.items.insert_one(i_dict)
    return {k: v for k, v in i_dict.items() if k != '_id'}

@api_router.put("/items/{item_id}")
async def update_item(item_id: str, data: ItemCreate, current_user: User = Depends(get_current_user)):
    await db.items.update_one({"id": item_id}, {"$set": data.model_dump()})
    updated = await db.items.find_one({"id": item_id}, {"_id": 0})
    return updated

@api_router.delete("/items/{item_id}")
async def delete_item(item_id: str, current_user: User = Depends(get_current_user)):
    await db.items.delete_one({"id": item_id})
    return {"message": "Item deleted"}

# Recurring Expenses
@api_router.get("/recurring-expenses")
async def get_recurring_expenses(current_user: User = Depends(get_current_user)):
    recs = await db.recurring_expenses.find({}, {"_id": 0}).to_list(100)
    now = datetime.now(timezone.utc)
    for r in recs:
        for f in ['next_due_date', 'created_at']:
            if isinstance(r.get(f), str):
                r[f] = datetime.fromisoformat(r[f])
        due = r.get('next_due_date')
        if due:
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
            r['days_until_due'] = (due - now).days
    return recs

@api_router.post("/recurring-expenses")
async def create_recurring_expense(data: RecurringExpenseCreate, current_user: User = Depends(get_current_user)):
    rec = RecurringExpense(**data.model_dump())
    r_dict = rec.model_dump()
    r_dict["next_due_date"] = r_dict["next_due_date"].isoformat()
    r_dict["created_at"] = r_dict["created_at"].isoformat()
    if r_dict.get("branch_id") == '':
        r_dict["branch_id"] = None
    await db.recurring_expenses.insert_one(r_dict)
    return {k: v for k, v in r_dict.items() if k != '_id'}

@api_router.delete("/recurring-expenses/{rec_id}")
async def delete_recurring_expense(rec_id: str, current_user: User = Depends(get_current_user)):
    await db.recurring_expenses.delete_one({"id": rec_id})
    return {"message": "Recurring expense deleted"}

@api_router.post("/recurring-expenses/{rec_id}/renew-pay")
async def renew_pay_recurring(rec_id: str, body: dict, current_user: User = Depends(get_current_user)):
    rec = await db.recurring_expenses.find_one({"id": rec_id}, {"_id": 0})
    if not rec: raise HTTPException(status_code=404, detail="Not found")
    
    amount = float(body.get("amount", rec["amount"]))
    mode = body.get("payment_mode", "cash")
    branch_id = body.get("branch_id") or rec.get("branch_id")
    if branch_id == '': branch_id = None
    
    # Create expense record
    expense = Expense(category=rec.get("category", "other"), description=f"{rec['name']} - Renewed", amount=amount,
                      payment_mode=mode, branch_id=branch_id, date=datetime.now(timezone.utc), notes=f"Recurring: {rec['name']}", created_by=current_user.id)
    e_dict = expense.model_dump()
    e_dict["date"] = e_dict["date"].isoformat()
    e_dict["created_at"] = e_dict["created_at"].isoformat()
    await db.expenses.insert_one(e_dict)
    
    # Update next due date
    freq = rec.get("frequency", "monthly")
    due = rec.get("next_due_date")
    if isinstance(due, str): due = datetime.fromisoformat(due)
    if due is None: due = datetime.now(timezone.utc)
    if freq == "monthly": new_due = due.replace(month=due.month % 12 + 1) if due.month < 12 else due.replace(year=due.year + 1, month=1)
    elif freq == "quarterly":
        m = due.month + 3
        new_due = due.replace(year=due.year + m // 12, month=m % 12 or 12) if m > 12 else due.replace(month=m)
    elif freq == "yearly": new_due = due.replace(year=due.year + 1)
    else: new_due = due.replace(month=due.month % 12 + 1) if due.month < 12 else due.replace(year=due.year + 1, month=1)
    
    await db.recurring_expenses.update_one({"id": rec_id}, {"$set": {"next_due_date": new_due.isoformat(), "amount": amount}})
    
    return {"message": f"Paid SAR {amount:.2f} & renewed. Next due: {new_due.strftime('%d %b %Y')}"}

# Employee Pending Salary Summary
@api_router.get("/employees/pending-summary")
async def get_employees_pending(current_user: User = Depends(get_current_user)):
    employees = await db.employees.find({"active": {"$ne": False}}, {"_id": 0}).to_list(1000)
    payments = await db.salary_payments.find({}, {"_id": 0}).to_list(10000)
    leaves = await db.leaves.find({}, {"_id": 0}).to_list(10000)
    
    now = datetime.now(timezone.utc)
    current_period = now.strftime("%b %Y")
    
    result = []
    for emp in employees:
        eid = emp["id"]
        emp_payments = [p for p in payments if p.get("employee_id") == eid and p.get("period") == current_period]
        salary_paid = sum(p["amount"] for p in emp_payments if p.get("payment_type") == "salary")
        pending = emp.get("salary", 0) - salary_paid
        
        emp_leaves = [l for l in leaves if l.get("employee_id") == eid]
        annual_used = sum(l.get("days", 0) for l in emp_leaves if l.get("leave_type") == "annual" and l.get("status") == "approved")
        sick_used = sum(l.get("days", 0) for l in emp_leaves if l.get("leave_type") == "sick" and l.get("status") == "approved")
        pending_leaves = sum(1 for l in emp_leaves if l.get("status") == "pending")
        
        # Check if currently on leave
        on_leave = None
        for l in emp_leaves:
            if l.get("status") == "approved":
                start = l.get("start_date")
                end = l.get("end_date")
                if isinstance(start, str): start = datetime.fromisoformat(start)
                if isinstance(end, str): end = datetime.fromisoformat(end)
                if start and end:
                    if start.tzinfo is None: start = start.replace(tzinfo=timezone.utc)
                    if end.tzinfo is None: end = end.replace(tzinfo=timezone.utc)
                    if start <= now <= end:
                        on_leave = {"from": start.strftime("%d %b"), "to": end.strftime("%d %b %Y"), "type": l.get("leave_type", "")}
                        break
        
        result.append({
            "id": eid, "name": emp["name"], "position": emp.get("position", ""),
            "branch_id": emp.get("branch_id"),
            "salary": emp.get("salary", 0), "salary_paid": salary_paid, "pending_salary": max(0, pending),
            "loan_balance": emp.get("loan_balance", 0),
            "on_leave": on_leave,
            "annual_leave_remaining": emp.get("annual_leave_entitled", 30) - annual_used,
            "sick_leave_remaining": emp.get("sick_leave_entitled", 15) - sick_used,
            "pending_leave_requests": pending_leaves
        })
    
    # Branch-wise salary summary
    branch_map = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    branch_summary = {}
    for emp in result:
        bid = emp.get("branch_id")
        bname = branch_map.get(bid, "No Branch") if bid else "No Branch"
        if bname not in branch_summary:
            branch_summary[bname] = {"total_salary": 0, "total_paid": 0, "total_pending": 0, "count": 0}
        branch_summary[bname]["total_salary"] += emp["salary"]
        branch_summary[bname]["total_paid"] += emp["salary_paid"]
        branch_summary[bname]["total_pending"] += emp["pending_salary"]
        branch_summary[bname]["count"] += 1
    
    total_salary = sum(e["salary"] for e in result)
    total_paid = sum(e["salary_paid"] for e in result)
    total_pending = sum(e["pending_salary"] for e in result)
    
    return {
        "employees": result,
        "branch_summary": branch_summary,
        "totals": {"total_salary": total_salary, "total_paid": total_paid, "total_pending": total_pending, "employee_count": len(result)},
        "period": current_period
    }

# Cash Transfer Routes
@api_router.get("/cash-transfers")
async def get_cash_transfers(current_user: User = Depends(get_current_user)):
    transfers = await db.cash_transfers.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for t in transfers:
        for f in ['date', 'created_at']:
            if isinstance(t.get(f), str):
                t[f] = datetime.fromisoformat(t[f])
    return transfers

@api_router.post("/cash-transfers")
async def create_cash_transfer(data: CashTransferCreate, current_user: User = Depends(get_current_user)):
    branches = {b["id"]: b["name"] for b in await db.branches.find({}, {"_id": 0}).to_list(100)}
    from_name = branches.get(data.from_branch_id, "Office") if data.from_branch_id else "Office"
    to_name = branches.get(data.to_branch_id, "Office") if data.to_branch_id else "Office"
    
    transfer = CashTransfer(
        **data.model_dump(),
        from_branch_name=from_name, to_branch_name=to_name,
        created_by=current_user.id
    )
    t_dict = transfer.model_dump()
    t_dict["date"] = t_dict["date"].isoformat()
    t_dict["created_at"] = t_dict["created_at"].isoformat()
    for f in ['from_branch_id', 'to_branch_id']:
        if t_dict.get(f) == '':
            t_dict[f] = None
    await db.cash_transfers.insert_one(t_dict)
    return {k: v for k, v in t_dict.items() if k != '_id'}

@api_router.delete("/cash-transfers/{transfer_id}")
async def delete_cash_transfer(transfer_id: str, current_user: User = Depends(get_current_user)):
    result = await db.cash_transfers.delete_one({"id": transfer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transfer not found")
    return {"message": "Transfer deleted"}

# Invoice Routes
@api_router.get("/invoices")
async def get_invoices(current_user: User = Depends(get_current_user)):
    invoices = await db.invoices.find({}, {"_id": 0}).sort("date", -1).to_list(1000)
    for inv in invoices:
        for f in ['date', 'created_at']:
            if isinstance(inv.get(f), str):
                inv[f] = datetime.fromisoformat(inv[f])
    return invoices

@api_router.post("/invoices")
async def create_invoice(data: InvoiceCreate, current_user: User = Depends(get_current_user)):
    # Get customer name
    customer_name = None
    if data.customer_id:
        cust = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
        customer_name = cust["name"] if cust else None
    
    # Calculate totals
    items = []
    subtotal = 0
    for item in data.items:
        qty = float(item.get("quantity", 1))
        price = float(item.get("unit_price", 0))
        item_total = qty * price
        items.append({"description": item["description"], "quantity": qty, "unit_price": price, "total": item_total})
        subtotal += item_total
    
    discount = data.discount or 0
    total = subtotal - discount
    
    # Generate invoice number
    count = await db.invoices.count_documents({})
    inv_number = f"INV-{count + 1:05d}"
    
    # Build payment details
    if data.payment_details:
        payment_details = data.payment_details
    else:
        payment_details = [{"mode": data.payment_mode, "amount": total}]
    
    invoice = Invoice(
        invoice_number=inv_number,
        branch_id=data.branch_id or None,
        customer_id=data.customer_id or None,
        customer_name=customer_name,
        items=items,
        subtotal=subtotal,
        discount=discount,
        total=total,
        payment_mode=data.payment_mode,
        payment_details=payment_details,
        date=data.date,
        notes=data.notes,
        created_by=current_user.id
    )
    inv_dict = invoice.model_dump()
    inv_dict["date"] = inv_dict["date"].isoformat()
    inv_dict["created_at"] = inv_dict["created_at"].isoformat()
    await db.invoices.insert_one(inv_dict)
    
    # Auto-create sale entry
    cash_bank_paid = sum(p["amount"] for p in payment_details if p.get("mode") in ["cash", "bank"])
    credit_in_details = sum(p["amount"] for p in payment_details if p.get("mode") == "credit")
    credit_amount = max(0, credit_in_details if credit_in_details > 0 else total - cash_bank_paid)
    
    sale = Sale(
        sale_type="branch" if not data.customer_id else "online",
        branch_id=data.branch_id or None,
        customer_id=data.customer_id or None,
        amount=subtotal,
        discount=discount,
        final_amount=total,
        payment_details=payment_details,
        credit_amount=credit_amount,
        credit_received=0,
        date=data.date,
        notes=f"Invoice {inv_number}",
        created_by=current_user.id
    )
    sale_dict = sale.model_dump()
    sale_dict["date"] = sale_dict["date"].isoformat()
    sale_dict["created_at"] = sale_dict["created_at"].isoformat()
    await db.sales.insert_one(sale_dict)
    
    # Link sale to invoice
    await db.invoices.update_one({"id": invoice.id}, {"$set": {"sale_id": sale.id}})
    inv_dict["sale_id"] = sale.id
    
    return {k: v for k, v in inv_dict.items() if k != '_id'}

@api_router.delete("/invoices/{invoice_id}")
async def delete_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    inv = await db.invoices.find_one({"id": invoice_id}, {"_id": 0})
    if inv and inv.get("sale_id"):
        await db.sales.delete_one({"id": inv["sale_id"]})
    result = await db.invoices.delete_one({"id": invoice_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Invoice and linked sale deleted"}

# Database Backup
@api_router.get("/backup/database")
async def backup_database(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    import json as json_module
    
    collections = ["users", "branches", "customers", "suppliers", "sales", "expenses", 
                    "supplier_payments", "salary_payments", "employees", "leaves", "documents",
                    "categories", "invoices", "items", "cash_transfers", "notifications",
                    "employee_requests", "recurring_expenses", "email_settings", "whatsapp_config",
                    "notification_prefs", "whatsapp_settings"]
    
    backup_data = {"backup_date": datetime.now(timezone.utc).isoformat(), "app": "SSC Track", "collections": {}}
    
    for col_name in collections:
        try:
            col = db[col_name]
            docs = await col.find({}, {"_id": 0}).to_list(100000)
            backup_data["collections"][col_name] = docs
        except:
            backup_data["collections"][col_name] = []
    
    json_str = json_module.dumps(backup_data, default=str, indent=2)
    buffer = BytesIO(json_str.encode('utf-8'))
    buffer.seek(0)
    
    fname = f"dataentry_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    return StreamingResponse(buffer, media_type="application/json",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})

@api_router.post("/backup/restore")
async def restore_database(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    # This is a placeholder - actual restore requires file upload
    return {"message": "Use the backup JSON file to restore. Contact support for restore assistance."}

# Database Import from XLS/CSV
@api_router.post("/import/data")
async def import_data(file: UploadFile = File(...), data_type: str = Form(...), current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    content = await file.read()
    
    try:
        if file.filename.endswith('.csv'):
            df = pd.read_csv(BytesIO(content))
        else:
            df = pd.read_excel(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read file: {str(e)}")
    
    records = df.where(df.notna(), None).to_dict('records')
    imported = 0
    errors = []
    
    for i, row in enumerate(records):
        try:
            clean = {k: (None if v is None or str(v).strip() == '' else v) for k, v in row.items()}
            
            if data_type == "customers":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "branch_id": None, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.customers.insert_one(doc)
                    imported += 1
            elif data_type == "suppliers":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "category": str(clean.get("category", "")) if clean.get("category") else None, "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "branch_id": None, "credit_limit": float(clean.get("credit_limit", 0) or 0), "current_credit": 0, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.suppliers.insert_one(doc)
                    imported += 1
            elif data_type == "employees":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "document_id": str(clean.get("document_id", "")) if clean.get("document_id") else None, "phone": str(clean.get("phone", "")) if clean.get("phone") else None, "email": str(clean.get("email", "")) if clean.get("email") else None, "position": str(clean.get("position", "")) if clean.get("position") else None, "salary": float(clean.get("salary", 0) or 0), "branch_id": None, "loan_balance": 0, "annual_leave_entitled": 30, "sick_leave_entitled": 15, "ticket_entitled": 1, "ticket_years": 2, "ticket_used": 0, "active": True, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.employees.insert_one(doc)
                    imported += 1
            elif data_type == "items":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "unit_price": float(clean.get("unit_price", clean.get("price", 0)) or 0), "category": str(clean.get("category", "")) if clean.get("category") else None, "active": True, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.items.insert_one(doc)
                    imported += 1
            elif data_type == "branches":
                doc = {"id": str(uuid.uuid4()), "name": str(clean.get("name", "")), "location": str(clean.get("location", "")) if clean.get("location") else None, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["name"]:
                    await db.branches.insert_one(doc)
                    imported += 1
            elif data_type == "sales":
                amt = float(clean.get("amount", 0) or 0)
                disc = float(clean.get("discount", 0) or 0)
                mode = str(clean.get("payment_mode", "cash") or "cash").lower()
                doc = {"id": str(uuid.uuid4()), "sale_type": str(clean.get("sale_type", "branch")), "branch_id": None, "customer_id": None, "amount": amt, "discount": disc, "final_amount": amt - disc, "payment_details": [{"mode": mode, "amount": amt - disc}], "credit_amount": 0, "credit_received": 0, "date": str(clean.get("date", datetime.now(timezone.utc).isoformat())), "notes": str(clean.get("notes", "") or ""), "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()}
                await db.sales.insert_one(doc)
                imported += 1
            elif data_type == "expenses_import":
                amt = float(clean.get("amount", 0) or 0)
                doc = {"id": str(uuid.uuid4()), "category": str(clean.get("category", "other")), "description": str(clean.get("description", "")), "amount": amt, "payment_mode": str(clean.get("payment_mode", "cash") or "cash"), "branch_id": None, "supplier_id": None, "date": str(clean.get("date", datetime.now(timezone.utc).isoformat())), "notes": "", "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()}
                if doc["description"]:
                    await db.expenses.insert_one(doc)
                    imported += 1
        except Exception as e:
            errors.append(f"Row {i+1}: {str(e)[:50]}")
    
    return {"message": f"Imported {imported} records", "imported": imported, "total_rows": len(records), "errors": errors[:10]}

# Download Import Template
@api_router.get("/import/template/{data_type}")
async def download_import_template(data_type: str, current_user: User = Depends(get_current_user)):
    templates = {
        "customers": ["name", "phone", "email"],
        "suppliers": ["name", "category", "phone", "email", "credit_limit"],
        "employees": ["name", "document_id", "phone", "email", "position", "salary"],
        "items": ["name", "unit_price", "category"],
        "branches": ["name", "location"],
        "sales": ["date", "sale_type", "amount", "discount", "payment_mode", "notes"],
        "expenses_import": ["date", "category", "description", "amount", "payment_mode"],
    }
    if data_type not in templates:
        raise HTTPException(status_code=400, detail="Invalid type")
    
    wb = Workbook()
    ws = wb.active
    ws.title = data_type.capitalize()
    ws.append(templates[data_type])
    for col in ws[1]:
        col.font = Font(bold=True, color="FFFFFF")
        col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
    ws.append(["Example " + templates[data_type][0]] + ["" for _ in templates[data_type][1:]])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={data_type}_import_template.xlsx"})

# ===== SETTINGS & NOTIFICATION ROUTES =====

# Email Settings
@api_router.get("/settings/email")
async def get_email_settings(current_user: User = Depends(get_current_user)):
    settings = await db.email_settings.find_one({}, {"_id": 0})
    if settings and settings.get("password"):
        settings["password"] = "••••••••"  # mask password
    return settings

@api_router.post("/settings/email")
async def save_email_settings(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.email_settings.find_one({})
    data = {
        "smtp_host": body.get("smtp_host", ""),
        "smtp_port": int(body.get("smtp_port", 587)),
        "username": body.get("username", ""),
        "from_email": body.get("from_email", ""),
        "use_tls": body.get("use_tls", True),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if body.get("password") and body["password"] != "••••••••":
        data["password"] = body["password"]
    if existing:
        await db.email_settings.update_one({}, {"$set": data})
    else:
        data["password"] = body.get("password", "")
        await db.email_settings.insert_one(data)
    return {"message": "Email settings saved"}

@api_router.post("/settings/email/test")
async def test_email(body: dict, current_user: User = Depends(get_current_user)):
    import aiosmtplib
    from email.mime.text import MIMEText
    settings = await db.email_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("smtp_host"):
        raise HTTPException(status_code=400, detail="Email not configured. Save settings first.")
    try:
        to_email = body.get("to_email", current_user.email)
        msg = MIMEText("This is a test email from SSC Track. Your email settings are working correctly!")
        msg["Subject"] = "SSC Track - Test Email"
        msg["From"] = settings.get("from_email", settings["username"])
        msg["To"] = to_email
        await aiosmtplib.send(msg, hostname=settings["smtp_host"], port=settings["smtp_port"],
                              username=settings["username"], password=settings["password"],
                              use_tls=settings.get("use_tls", True))
        return {"message": f"Test email sent to {to_email}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Email failed: {str(e)}")

# WhatsApp Settings
@api_router.get("/settings/whatsapp")
async def get_whatsapp_settings(current_user: User = Depends(get_current_user)):
    settings = await db.whatsapp_config.find_one({}, {"_id": 0})
    if settings and settings.get("auth_token"):
        settings["auth_token"] = "••••••••"
    return settings

@api_router.post("/settings/whatsapp")
async def save_whatsapp_settings(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.whatsapp_config.find_one({})
    data = {
        "account_sid": body.get("account_sid", ""),
        "phone_number": body.get("phone_number", ""),
        "recipient_number": body.get("recipient_number", ""),
        "enabled": body.get("enabled", True),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if body.get("auth_token") and body["auth_token"] != "••••••••":
        data["auth_token"] = body["auth_token"]
    if existing:
        await db.whatsapp_config.update_one({}, {"$set": data})
    else:
        data["auth_token"] = body.get("auth_token", "")
        await db.whatsapp_config.insert_one(data)
    return {"message": "WhatsApp settings saved"}

# Notification Preferences
@api_router.get("/settings/notifications")
async def get_notification_prefs(current_user: User = Depends(get_current_user)):
    prefs = await db.notification_prefs.find_one({}, {"_id": 0})
    return prefs or {
        "email_daily_sales": False,
        "email_document_expiry": True,
        "email_leave_updates": False,
        "whatsapp_daily_sales": False,
        "whatsapp_document_expiry": False,
    }

@api_router.post("/settings/notifications")
async def save_notification_prefs(body: dict, current_user: User = Depends(get_current_user)):
    existing = await db.notification_prefs.find_one({})
    data = {
        "email_daily_sales": body.get("email_daily_sales", False),
        "email_document_expiry": body.get("email_document_expiry", True),
        "email_leave_updates": body.get("email_leave_updates", False),
        "whatsapp_daily_sales": body.get("whatsapp_daily_sales", False),
        "whatsapp_document_expiry": body.get("whatsapp_document_expiry", False),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if existing:
        await db.notification_prefs.update_one({}, {"$set": data})
    else:
        await db.notification_prefs.insert_one(data)
    return {"message": "Notification preferences saved"}

# Send email notification helper
async def send_email_notification(subject: str, body_text: str, to_email: str = None):
    import aiosmtplib
    from email.mime.text import MIMEText
    settings = await db.email_settings.find_one({}, {"_id": 0})
    if not settings or not settings.get("smtp_host") or not settings.get("password"):
        return False
    try:
        recipient = to_email or settings.get("from_email", settings["username"])
        msg = MIMEText(body_text)
        msg["Subject"] = subject
        msg["From"] = settings.get("from_email", settings["username"])
        msg["To"] = recipient
        await aiosmtplib.send(msg, hostname=settings["smtp_host"], port=settings["smtp_port"],
                              username=settings["username"], password=settings["password"],
                              use_tls=settings.get("use_tls", True))
        return True
    except:
        return False

# Send test WhatsApp
@api_router.post("/settings/whatsapp/test")
async def test_whatsapp(current_user: User = Depends(get_current_user)):
    config = await db.whatsapp_config.find_one({}, {"_id": 0})
    if not config or not config.get("account_sid") or not config.get("auth_token"):
        raise HTTPException(status_code=400, detail="WhatsApp not configured")
    try:
        client = Client(config["account_sid"], config["auth_token"])
        message = client.messages.create(
            from_=f'whatsapp:{config["phone_number"]}',
            body="Test message from SSC Track. Your WhatsApp settings are working!",
            to=f'whatsapp:{config["recipient_number"]}'
        )
        return {"message": "Test WhatsApp sent", "sid": message.sid}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"WhatsApp failed: {str(e)}")

# Send daily sales report via configured channels
@api_router.post("/send-daily-report")
async def send_daily_report(current_user: User = Depends(get_current_user)):
    prefs = await db.notification_prefs.find_one({}, {"_id": 0}) or {}
    
    today_start = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
    today_end = today_start + timedelta(days=1)
    
    sales = await db.sales.find({"date": {"$gte": today_start.isoformat(), "$lt": today_end.isoformat()}}, {"_id": 0}).to_list(1000)
    expenses = await db.expenses.find({"date": {"$gte": today_start.isoformat(), "$lt": today_end.isoformat()}}, {"_id": 0}).to_list(1000)
    branches = await db.branches.find({}, {"_id": 0}).to_list(100)
    
    total_sales = sum(s.get("final_amount", s["amount"]) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    
    branch_lines = ""
    for b in branches:
        bt = sum(s.get("final_amount", s["amount"]) for s in sales if s.get("branch_id") == b["id"])
        if bt > 0:
            branch_lines += f"  {b['name']}: ${bt:.2f}\n"
    
    report = f"Daily Sales Report - {datetime.now().strftime('%d %b %Y')}\n\nTotal Sales: ${total_sales:.2f}\nTotal Expenses: ${total_expenses:.2f}\nNet: ${(total_sales - total_expenses):.2f}\n"
    if branch_lines:
        report += f"\nBranch-wise Sales:\n{branch_lines}"
    
    results = []
    
    if prefs.get("email_daily_sales"):
        sent = await send_email_notification("SSC Track - Daily Sales Report", report)
        results.append(f"Email: {'sent' if sent else 'failed (check email settings)'}")
    
    if prefs.get("whatsapp_daily_sales"):
        config = await db.whatsapp_config.find_one({}, {"_id": 0})
        if config and config.get("account_sid") and config.get("auth_token"):
            try:
                client = Client(config["account_sid"], config["auth_token"])
                client.messages.create(from_=f'whatsapp:{config["phone_number"]}', body=report, to=f'whatsapp:{config["recipient_number"]}')
                results.append("WhatsApp: sent")
            except Exception as e:
                results.append(f"WhatsApp: failed ({str(e)[:50]})")
        else:
            results.append("WhatsApp: not configured")
    
    if not results:
        return {"message": "No notification channels enabled. Go to Settings to configure."}
    
    return {"message": "Report sent", "details": results}

# Export Routes
@api_router.post("/export/reports")
async def export_reports(export_request: ExportRequest, current_user: User = Depends(get_current_user)):
    try:
        # Fetch data
        sales = await db.sales.find({}, {"_id": 0}).to_list(10000)
        expenses = await db.expenses.find({}, {"_id": 0}).to_list(10000)
        supplier_payments = await db.supplier_payments.find({}, {"_id": 0}).to_list(10000)
        branches = await db.branches.find({}, {"_id": 0}).to_list(100)
        customers = await db.customers.find({}, {"_id": 0}).to_list(1000)
        
        # Filter by date range
        if export_request.start_date:
            start = datetime.fromisoformat(export_request.start_date)
            sales = [s for s in sales if datetime.fromisoformat(s["date"]) >= start]
            expenses = [e for e in expenses if datetime.fromisoformat(e["date"]) >= start]
            supplier_payments = [p for p in supplier_payments if datetime.fromisoformat(p["date"]) >= start]
        
        if export_request.end_date:
            end = datetime.fromisoformat(export_request.end_date)
            sales = [s for s in sales if datetime.fromisoformat(s["date"]) <= end]
            expenses = [e for e in expenses if datetime.fromisoformat(e["date"]) <= end]
            supplier_payments = [p for p in supplier_payments if datetime.fromisoformat(p["date"]) <= end]
        
        # Filter by branch
        if export_request.branch_id and export_request.branch_id != "all":
            sales = [s for s in sales if s.get("branch_id") == export_request.branch_id]
        
        if export_request.format == "pdf":
            return generate_pdf_report(sales, expenses, supplier_payments, branches, customers)
        elif export_request.format == "excel":
            return generate_excel_report(sales, expenses, supplier_payments, branches, customers)
        else:
            raise HTTPException(status_code=400, detail="Invalid export format")
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

def generate_pdf_report(sales, expenses, supplier_payments, branches, customers):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#F5841F'),
        spaceAfter=30,
        alignment=1
    )
    elements.append(Paragraph("SSC Track - Sales Report", title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    total_sales = sum(s.get("final_amount", s["amount"] - s.get("discount", 0)) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    total_supplier = sum(p["amount"] for p in supplier_payments)
    net_profit = total_sales - total_expenses - total_supplier
    
    summary_data = [
        ["Metric", "Amount"],
        ["Total Sales", f"SAR {total_sales:.2f}"],
        ["Total Expenses", f"SAR {total_expenses:.2f}"],
        ["Supplier Payments", f"SAR {total_supplier:.2f}"],
        ["Net Profit", f"SAR {net_profit:.2f}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5841F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    elements.append(Paragraph("Sales Transactions", styles['Heading2']))
    elements.append(Spacer(1, 0.1*inch))
    
    sales_data = [["Date", "Type", "Amount", "Payment"]]
    for sale in sales[:20]:
        date_str = datetime.fromisoformat(sale["date"]).strftime("%Y-%m-%d")
        modes = ", ".join(p["mode"].capitalize() for p in sale.get("payment_details", []))
        if not modes:
            modes = sale.get("payment_mode", "N/A").capitalize()
        sales_data.append([
            date_str,
            sale["sale_type"].capitalize(),
            f"SAR {sale['amount']:.2f}",
            modes
        ])
    
    sales_table = Table(sales_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(sales_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=sales_report.pdf"}
    )

def generate_excel_report(sales, expenses, supplier_payments, branches, customers):
    buffer = BytesIO()
    wb = Workbook()
    
    # Summary Sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary['A1'] = "SSC Track - Sales Report"
    ws_summary['A1'].font = Font(size=16, bold=True, color="F5841F")
    
    total_sales = sum(s.get("final_amount", s["amount"] - s.get("discount", 0)) for s in sales)
    total_expenses = sum(e["amount"] for e in expenses)
    total_supplier = sum(p["amount"] for p in supplier_payments)
    net_profit = total_sales - total_expenses - total_supplier
    
    ws_summary['A3'] = "Metric"
    ws_summary['B3'] = "Amount"
    ws_summary['A3'].font = Font(bold=True)
    ws_summary['B3'].font = Font(bold=True)
    
    summary_data = [
        ("Total Sales", total_sales),
        ("Total Expenses", total_expenses),
        ("Supplier Payments", total_supplier),
        ("Net Profit", net_profit)
    ]
    
    for idx, (metric, amount) in enumerate(summary_data, start=4):
        ws_summary[f'A{idx}'] = metric
        ws_summary[f'B{idx}'] = f"SAR {amount:.2f}"
    
    # Sales Sheet
    ws_sales = wb.create_sheet("Sales")
    sales_headers = ["Date", "Type", "Branch/Customer", "Amount", "Payment Mode", "Status"]
    ws_sales.append(sales_headers)
    
    for col in ws_sales[1]:
        col.font = Font(bold=True)
        col.fill = PatternFill(start_color="F5841F", end_color="F5841F", fill_type="solid")
        col.font = Font(bold=True, color="FFFFFF")
    
    for sale in sales:
        date_str = datetime.fromisoformat(sale["date"]).strftime("%Y-%m-%d")
        branch_name = next((b["name"] for b in branches if b["id"] == sale.get("branch_id")), "-")
        customer_name = next((c["name"] for c in customers if c["id"] == sale.get("customer_id")), "-")
        ref = branch_name if sale["sale_type"] == "branch" else customer_name
        modes = ", ".join(p["mode"].capitalize() for p in sale.get("payment_details", []))
        if not modes:
            modes = sale.get("payment_mode", "N/A").capitalize()
        has_credit = sale.get("credit_amount", 0) > 0
        remaining = sale.get("credit_amount", 0) - sale.get("credit_received", 0)
        status = "Paid" if remaining <= 0 else ("Partial" if sale.get("credit_received", 0) > 0 else "Pending") if has_credit else "Received"
        
        ws_sales.append([
            date_str,
            sale["sale_type"].capitalize(),
            ref,
            sale["amount"],
            modes,
            status
        ])
    
    # Expenses Sheet
    ws_expenses = wb.create_sheet("Expenses")
    expense_headers = ["Date", "Category", "Description", "Amount", "Payment Mode"]
    ws_expenses.append(expense_headers)
    
    for col in ws_expenses[1]:
        col.font = Font(bold=True)
        col.fill = PatternFill(start_color="F43F5E", end_color="F43F5E", fill_type="solid")
        col.font = Font(bold=True, color="FFFFFF")
    
    for expense in expenses:
        date_str = datetime.fromisoformat(expense["date"]).strftime("%Y-%m-%d")
        ws_expenses.append([
            date_str,
            expense["category"].capitalize(),
            expense["description"],
            expense["amount"],
            expense["payment_mode"].capitalize()
        ])
    
    # Supplier Payments Sheet
    ws_supplier = wb.create_sheet("Supplier Payments")
    supplier_headers = ["Date", "Supplier Name", "Amount", "Payment Mode", "Notes"]
    ws_supplier.append(supplier_headers)
    
    for col in ws_supplier[1]:
        col.font = Font(bold=True)
        col.fill = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")
        col.font = Font(bold=True, color="FFFFFF")
    
    for payment in supplier_payments:
        date_str = datetime.fromisoformat(payment["date"]).strftime("%Y-%m-%d")
        ws_supplier.append([
            date_str,
            payment["supplier_name"],
            payment["amount"],
            payment["payment_mode"].capitalize(),
            payment.get("notes", "")
        ])
    
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sales_report.xlsx"}
    )

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()